import openpyxl
from openpyxl import load_workbook
from typing import Optional, Tuple, List, Dict, Any
import os
import sys
import argparse
from datetime import datetime



def compare_worksheets(workbook1_path: str, worksheet1_name: str, 
                      workbook2_path: str, worksheet2_name: str,
                      max_row: Optional[int] = None, max_col: Optional[int] = None,
                      case_sensitive: bool = False) -> Dict[str, Any]:
    """
    Compare two worksheets cell by cell and return differences.
    
    Args:
        workbook1_path (str): Path to the first Excel workbook
        worksheet1_name (str): Name of worksheet in first workbook
        workbook2_path (str): Path to the second Excel workbook  
        worksheet2_name (str): Name of worksheet in second workbook
        max_row (int, optional): Maximum row to compare (if None, compares all rows)
        max_col (int, optional): Maximum column to compare (if None, compares all columns)
        case_sensitive (bool): Whether comparison should be case sensitive (default: False)
        
    Returns:
        Dict containing comparison results and statistics
    """
    
    # Load workbooks
    try:
        wb1 = load_workbook(workbook1_path, data_only=True)
        wb2 = load_workbook(workbook2_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbooks: {e}")
        return {"error": str(e)}
    
    # Get worksheets
    try:
        ws1 = wb1[worksheet1_name]
        ws2 = wb2[worksheet2_name]
    except KeyError as e:
        print(f"Error accessing worksheet: {e}")
        return {"error": f"Worksheet not found: {e}"}
    
    # Determine comparison range
    if max_row is None:
        max_row = max(ws1.max_row, ws2.max_row)
    if max_col is None:
        max_col = max(ws1.max_column, ws2.max_column)
    
    differences = []
    total_cells_compared = 0
    identical_cells = 0
    different_cells = 0
    
    print(f"Comparing worksheets:")
    print(f"  Workbook 1: {os.path.basename(workbook1_path)} -> {worksheet1_name}")
    print(f"  Workbook 2: {os.path.basename(workbook2_path)} -> {worksheet2_name}")
    print(f"  Range: A1 to {openpyxl.utils.get_column_letter(max_col)}{max_row}")
    print(f"  Case sensitive: {case_sensitive}")
    print("-" * 80)
    
    # Compare cell by cell
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            total_cells_compared += 1
            
            # Get cell values
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)
            
            value1 = cell1.value
            value2 = cell2.value
            
            # Convert to string for comparison (handle None values)
            str1 = str(value1) if value1 is not None else ""
            str2 = str(value2) if value2 is not None else ""

            # Apply case sensitivity setting
            if not case_sensitive:
                str1 = str1.lower()
                str2 = str2.lower()

            # Special rule for cell (1,1)
            if row == 1 and col == 1:
                import re
                def extract_special_parts(s):
                    # Match: first two letters, year (4 digits), and "Run"
                    m = re.match(r"([a-zA-Z]{2}).*?(\d{4}).*?(run)", s)
                    if m:
                        return m.group(1), m.group(2), m.group(3)
                    return None
                parts1 = extract_special_parts(str1)
                parts2 = extract_special_parts(str2)
                if parts1 and parts2 and parts1 == parts2:
                    identical_cells += 1
                    continue  # Considered identical, skip to next cell

            # Treat as identical if only difference is a single leading apostrophe
            def strip_leading_apostrophe(s):
                ans = s[1:] if s.startswith("'") else s
                return ans[1:] if ans.startswith("=") or ans.startswith("+") else ans

            if str1 != str2:
                # Check for leading apostrophe difference
                if strip_leading_apostrophe(str1) == strip_leading_apostrophe(str2):
                    identical_cells += 1
                    continue  # Considered identical, skip to next cell

                different_cells += 1
                cell_address = f"{openpyxl.utils.get_column_letter(col)}{row}"

                difference = {
                    "cell": cell_address,
                    "workbook1_value": value1,
                    "workbook2_value": value2,
                    "workbook1_display": str(value1) if value1 is not None else "None",
                    "workbook2_display": str(value2) if value2 is not None else "None"
                }
                differences.append(difference)
                
                # Print difference immediately
                print(f"Cell {cell_address:6}: '{difference['workbook1_display']:20}' != '{difference['workbook2_display']:20}'")
            else:
                identical_cells += 1
    
    # Print summary
    print("-" * 80)
    print(f"Comparison Summary:")
    print(f"  Total cells compared: {total_cells_compared:,}")
    print(f"  Identical cells: {identical_cells:,}")
    print(f"  Different cells: {different_cells:,}")
    print(f"  Match percentage: {(identical_cells/total_cells_compared)*100:.2f}%")
    
    # Write results to ComparisonResult.txt in the same folder as workbook1
    try:
        workbook1_dir = os.path.dirname(os.path.abspath(workbook1_path))
        output_file = os.path.join(workbook1_dir, "ComparisonResult.txt")
        
        with open(output_file, 'a', encoding='utf-8') as f:
            f.write(f"Excel Worksheet Comparison Report\n")
            f.write(f"Generated: {datetime.now()}\n")
            f.write(f"="*80 + "\n\n")
            
            f.write(f"Source Files:\n")
            f.write(f"  Workbook 1: {workbook1_path} -> {worksheet1_name}\n")
            f.write(f"  Workbook 2: {workbook2_path} -> {worksheet2_name}\n\n")
            
            f.write(f"Comparison Settings:\n")
            f.write(f"  Max Row: {max_row}\n")
            f.write(f"  Max Col: {max_col}\n")
            f.write(f"  Case Sensitive: {case_sensitive}\n\n")
            
            f.write(f"Summary:\n")
            f.write(f"  Total cells compared: {total_cells_compared:,}\n")
            f.write(f"  Identical cells: {identical_cells:,}\n")
            f.write(f"  Different cells: {different_cells:,}\n")
            f.write(f"  Match percentage: {(identical_cells/total_cells_compared)*100:.2f}%\n\n")
            
            if differences:
                f.write(f"Differences Found ({len(differences)}):\n")
                f.write(f"{'Cell':8} {'Workbook 1':30} {'Workbook 2':30}\n")
                f.write(f"{'-'*8} {'-'*30} {'-'*30}\n")
                
                for diff in differences:
                    wb1_val = str(diff['workbook1_display'])[:29]
                    wb2_val = str(diff['workbook2_display'])[:29]
                    f.write(f"{diff['cell']:8} {wb1_val:30} {wb2_val:30}\n")
            else:
                f.write("No differences found - worksheets are identical!\n")
        
        print(f"Comparison results saved to: {output_file}")
        
    except Exception as e:
        print(f"Error saving comparison results: {e}")
    
    # Return results
    return {
        "success": True,
        "workbook1_path": workbook1_path,
        "worksheet1_name": worksheet1_name,
        "workbook2_path": workbook2_path,
        "worksheet2_name": worksheet2_name,
        "total_cells_compared": total_cells_compared,
        "identical_cells": identical_cells,
        "different_cells": different_cells,
        "match_percentage": (identical_cells/total_cells_compared)*100,
        "differences": differences,
        "case_sensitive": case_sensitive,
        "max_row": max_row,
        "max_col": max_col,
        "output_file": output_file if 'output_file' in locals() else None
    }



def compare_two_workbooks(workbook1_path: str, 
                           workbook2_path: str, 
                           case_sensitive: bool = False,
                           sheet_name: str = None):
    """
    Compare two Excel workbooks and return the comparison results.
    
    Args:
        workbook1_path (str): Path to the first workbook
        workbook2_path (str): Path to the second workbook  
        case_sensitive (bool): Whether comparison should be case sensitive
        sheet_name (str, optional): If provided, only compare this specific sheet
    """
    
    # Use provided workbook paths instead of hardcoded ones
    source_wb1 = os.path.abspath(workbook1_path)
    target_wb2 = os.path.abspath(workbook2_path)
    
    # Load the target workbook to get sheet names
    try:
        wb2 = load_workbook(target_wb2, data_only=True)
    except Exception as e:
        print(f"Error loading target workbook: {e}")
        return []

    # Remove ComparisonResult.txt if it exists (in same directory as workbook1)
    workbook1_dir = os.path.dirname(source_wb1)
    comparison_result_path = os.path.join(workbook1_dir, "ComparisonResult.txt")
    if os.path.exists(comparison_result_path):
        os.remove(comparison_result_path)
        
    # Get sheet names to compare
    if sheet_name:
        # Only compare the specified sheet
        sheetnames = [sheet_name]
        print(f"Comparing only sheet: {sheet_name}")
    else:
        # Compare all sheets
        sheetnames = wb2.sheetnames
        print(f"Comparing all sheets: {sheetnames}")

    all_results = []
    total_cells = 0
    total_identical = 0
    total_different = 0
    sheet_summaries = []

    for sheetname in sheetnames:
        print(f"\nComparing worksheet: {sheetname}")
        if os.path.exists(source_wb1) and os.path.exists(target_wb2):
            # Check if the sheet exists in both workbooks
            try:
                wb1_check = load_workbook(source_wb1, data_only=True)
                if sheetname not in wb1_check.sheetnames:
                    print(f"Warning: Sheet '{sheetname}' not found in {source_wb1}")
                    continue
                if sheetname not in wb2.sheetnames:
                    print(f"Warning: Sheet '{sheetname}' not found in {target_wb2}")
                    continue
            except Exception as e:
                print(f"Error checking sheets: {e}")
                continue
                
            result = compare_worksheets(
                workbook1_path=source_wb1,
                worksheet1_name=sheetname, 
                workbook2_path=target_wb2,
                worksheet2_name=sheetname,
                case_sensitive=case_sensitive  # Use the parameter instead of hardcoded False
            )
            print(f"\nComparison completed. Success: {result.get('success', False)}")
            all_results.append(result)
            total_cells += result.get('total_cells_compared', 0)
            total_identical += result.get('identical_cells', 0)
            total_different += result.get('different_cells', 0)
            sheet_summaries.append(
                f"Sheet '{sheetname}': {result.get('identical_cells', 0)} identical, {result.get('different_cells', 0)} different, Match: {result.get('match_percentage', 0):.2f}%"
            )
        else:
            print("One or both Excel files not found:")
            print(f"  File 1 exists: {os.path.exists(source_wb1)} - {source_wb1}")
            print(f"  File 2 exists: {os.path.exists(target_wb2)} - {target_wb2}")

    # Print overall summary
    print("\n" + "="*80)
    if sheet_name:
        print(f"Comparison Summary for Sheet '{sheet_name}':")
    else:
        print("Overall Comparison Summary:")
    for summary in sheet_summaries:
        print(summary)
    print(f"\nTotal sheets compared: {len(sheetnames)}")
    print(f"Number of identical sheets: {len([s for s in sheet_summaries if 'different, Match: 100.00%' in s])}")
    print(f"Number of different sheets: {len(sheetnames) - len([s for s in sheet_summaries if 'different, Match: 100.00%' in s])}")
    print(f"\nTotal cells compared: {total_cells:,}")
    print(f"Total identical cells: {total_identical:,}")
    print(f"Total different cells: {total_different:,}")
    if total_cells > 0:
        print(f"Overall match percentage: {(total_identical/total_cells)*100:.2f}%")
    print("="*80)

    # Write overall summary to ComparisonResult.txt
    try:
        with open(comparison_result_path, 'a', encoding='utf-8') as f:
            f.write("\n" + "="*80 + "\n")
            if sheet_name:
                f.write(f"Comparison Summary for Sheet '{sheet_name}':\n")
            else:
                f.write("Overall Comparison Summary:\n")
            for summary in sheet_summaries:
                f.write(summary + "\n")
            f.write(f"\nTotal sheets compared: {len(sheetnames)}")
            f.write(f"\nNumber of identical sheets: {len([s for s in sheet_summaries if 'different, Match: 100.00%' in s])}")
            f.write(f"\nNumber of different sheets: {len(sheetnames) - len([s for s in sheet_summaries if 'different, Match: 100.00%' in s])}\n")
            f.write(f"\nTotal cells compared: {total_cells:,}\n")
            f.write(f"Total identical cells: {total_identical:,}\n")
            f.write(f"Total different cells: {total_different:,}\n")
            if total_cells > 0:
                f.write(f"Overall match percentage: {(total_identical/total_cells)*100:.2f}%\n")
            f.write("="*80 + "\n")
    except Exception as e:
        print(f"Error writing overall summary: {e}")
    return all_results

# Example usage and test function
if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(
        description="Excel Worksheet Comparison Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python CompareTwoSheets.py                    # Compare default workbooks (all sheets)
  python CompareTwoSheets.py A1P9               # Compare default workbooks (only A1P9 sheet)
  python CompareTwoSheets.py --wb1 file1.xlsx --wb2 file2.xlsx    # Compare custom workbooks (all sheets)
  python CompareTwoSheets.py A1P9 --wb1 file1.xlsx --wb2 file2.xlsx    # Compare custom workbooks (only A1P9 sheet)
  uv run CompareTwoSheets.py                    # Compare default workbooks (all sheets)
  uv run CompareTwoSheets.py A1P9               # Compare default workbooks (only A1P9 sheet)
        """
    )
    parser.add_argument(
        'sheet_name', 
        nargs='?', 
        default=None,
        help='Optional sheet name to compare. If not provided, all sheets will be compared.'
    )
    parser.add_argument(
        '--case-sensitive', 
        action='store_true',
        help='Enable case-sensitive comparison (default: False)'
    )
    parser.add_argument(
        '--wb1', '--workbook1',
        type=str,
        help='Path to the first workbook to compare'
    )
    parser.add_argument(
        '--wb2', '--workbook2', 
        type=str,
        help='Path to the second workbook to compare'
    )
    
    # Parse arguments
    args = parser.parse_args()
    
    # Determine workbook paths
    if args.wb1 and args.wb2:
        # Use custom workbook paths
        source_wb1 = os.path.abspath(args.wb1)
        target_wb2 = os.path.abspath(args.wb2)
        print(f"Using custom workbooks:")
        print(f"  Workbook 1: {source_wb1}")
        print(f"  Workbook 2: {target_wb2}")
    else:
        # Use default workbook paths
        reports_dir = os.path.join(os.path.dirname(__file__), "../..", "reports")
        source_wb1 = os.path.abspath(os.path.join(reports_dir, "CN2023.xlsx"))
        target_wb2 = os.path.abspath(os.path.join(reports_dir, "CN-2023_report.xlsx"))
        print(f"Using default workbooks:")
        print(f"  Workbook 1: {source_wb1}")
        print(f"  Workbook 2: {target_wb2}")
    
    # Example usage
    print("\nExcel Worksheet Comparison Tool")
    print("=" * 50)
    
    if args.sheet_name:
        print(f"Comparing single sheet: {args.sheet_name}")
    else:
        print("Comparing all sheets")
    
    if args.case_sensitive:
        print("Case-sensitive comparison enabled")

    # Check if files exist
    if not os.path.exists(source_wb1):
        print(f"Source workbook not found: {source_wb1}")
        sys.exit(1)
    if not os.path.exists(target_wb2):
        print(f"Target workbook not found: {target_wb2}")
        sys.exit(1)
        
    # Run comparison with command line arguments
    compare_two_workbooks(
        source_wb1, 
        target_wb2, 
        case_sensitive=args.case_sensitive,
        sheet_name=args.sheet_name
    )
    
    print("\nComparison completed successfully!")
