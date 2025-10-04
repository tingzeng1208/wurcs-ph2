"""
Formula Utilities for Excel Report Generation
This module provides utilities for handling different types of formulas in Excel reports.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

class ExcelFormulaHandler:
    """Handles different types of formula writing to Excel cells"""
    
    @staticmethod
    def add_text_formula(ws, row, col, formula_text, formatting=None):
        """
        Add a formula as display text (not executable) - will show exactly as written
        
        Args:
            ws: Worksheet object
            row: Row number (1-based)
            col: Column number (1-based) or letter
            formula_text: The formula text to display
            formatting: Optional cell formatting dictionary
        
        Returns:
            Cell object
        """
        cell = ws.cell(row=row, column=col)
        cell.number_format = '@'  # Text format
        cell.value = formula_text
        
        if formatting:
            ExcelFormulaHandler._apply_formatting(cell, formatting)
        
        return cell
    
    @staticmethod
    def add_executable_formula(ws, row, col, formula_text, formatting=None):
        """
        Add an executable formula that Excel will calculate
        
        Args:
            ws: Worksheet object
            row: Row number (1-based)
            col: Column number (1-based) or letter
            formula_text: The formula text (with or without leading =)
            formatting: Optional cell formatting dictionary
        
        Returns:
            Cell object
        """
        cell = ws.cell(row=row, column=col)
        
        # Ensure formula starts with = for Excel to execute it
        if formula_text and not formula_text.startswith('='):
            formula_text = '=' + formula_text
        
        cell.value = formula_text
        
        if formatting:
            ExcelFormulaHandler._apply_formatting(cell, formatting)
        
        return cell
    
    @staticmethod
    def smart_formula_handler(ws, row, col, text_value, force_executable=False, force_text=False, formatting=None):
        """
        Smart formula handler that decides whether to make formula executable or display as text
        
        Args:
            ws: Worksheet object
            row: Row number (1-based)
            col: Column number (1-based) or letter
            text_value: The text/formula to write
            force_executable: Force the formula to be executable (overrides auto-detection)
            force_text: Force the formula to be text display (overrides auto-detection)
            formatting: Optional cell formatting dictionary
        
        Returns:
            Cell object
        """
        if not text_value:
            cell = ws.cell(row=row, column=col, value="")
            return cell
        
        text_str = str(text_value).strip()
        
        # Force text display
        if force_text:
            return ExcelFormulaHandler.add_text_formula(ws, row, col, text_str, formatting)
        
        # Force executable
        if force_executable:
            return ExcelFormulaHandler.add_executable_formula(ws, row, col, text_str, formatting)
        
        # Auto-detection logic
        if text_str.startswith('='):
            # Looks like a formula
            # Check for common patterns that should be executable vs display
            if any(func in text_str.upper() for func in ['SUM(', 'AVERAGE(', 'COUNT(', 'IF(', 'VLOOKUP(', 'INDEX(', 'MATCH(']):
                # Common Excel functions - make executable
                return ExcelFormulaHandler.add_executable_formula(ws, row, col, text_str, formatting)
            else:
                # Other formulas - display as text by default (can be changed as needed)
                return ExcelFormulaHandler.add_text_formula(ws, row, col, text_str, formatting)
        
        elif text_str.startswith('+'):
            # Mathematical expression starting with + - make executable
            return ExcelFormulaHandler.add_executable_formula(ws, row, col, text_str, formatting)
        
        else:
            # Regular text
            cell = ws.cell(row=row, column=col, value=text_str)
            if formatting:
                ExcelFormulaHandler._apply_formatting(cell, formatting)
            return cell
    
    @staticmethod
    def _apply_formatting(cell, formatting):
        """Apply formatting to a cell"""
        if 'font' in formatting:
            cell.font = Font(**formatting['font'])
        if 'alignment' in formatting:
            cell.alignment = Alignment(**formatting['alignment'])
        if 'number_format' in formatting:
            cell.number_format = formatting['number_format']
    
    @staticmethod
    def enable_workbook_calculation(wb):
        """
        Enable automatic calculation for the workbook
        This ensures Excel will calculate formulas when the file is opened
        """
        try:
            wb.calculation.calcMode = 'automatic'
            wb.calculation.fullCalcOnLoad = True
            print("✓ Excel automatic calculation enabled")
            return True
        except AttributeError:
            print("⚠ Warning: Automatic calculation setting not available in this openpyxl version")
            return False
    
    @staticmethod
    def add_formula_comment(ws, row, col, comment_text):
        """Add a comment to a cell explaining the formula"""
        try:
            cell = ws.cell(row=row, column=col)
            # Note: Comments require openpyxl.comments.Comment
            from openpyxl.comments import Comment
            cell.comment = Comment(comment_text, "System")
        except ImportError:
            print(f"⚠ Warning: Could not add comment to cell {row},{col}")


def demo_formula_usage():
    """Demonstrate different formula handling approaches"""
    
    print("Creating Excel demo workbook with different formula types...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formula Demo"
    
    # Headers
    ws.cell(row=1, column=1, value="Formula Type")
    ws.cell(row=1, column=2, value="Formula Text")
    ws.cell(row=1, column=3, value="Result")
    ws.cell(row=1, column=4, value="Notes")
    
    # Examples
    examples = [
        ("Text Display", "=SUM(A1:A10)", "Shows as text", "Will display exactly as written"),
        ("Executable", "SUM(B2:B5)", "Calculates", "Excel will calculate this"),
        ("Smart Auto", "=AVERAGE(C1:C3)", "Auto-detected", "Smart handler decides"),
        ("Mathematical", "+5*2", "Calculates", "Simple math expression"),
        ("Regular Text", "Hello World", "Regular text", "Just regular text")
    ]
    
    # Add sample data for calculations
    for i in range(2, 6):
        ws.cell(row=i, column=2, value=i * 10)  # B2:B5 = 20, 30, 40, 50
        ws.cell(row=i, column=3, value=i * 5)   # C1:C3 will be 10, 15, 20
    
    row_num = 7
    for formula_type, formula_text, _, notes in examples:
        ws.cell(row=row_num, column=1, value=formula_type)
        ws.cell(row=row_num, column=4, value=notes)
        
        if formula_type == "Text Display":
            ExcelFormulaHandler.add_text_formula(ws, row_num, 3, formula_text)
        elif formula_type == "Executable":
            ExcelFormulaHandler.add_executable_formula(ws, row_num, 3, formula_text)
        elif formula_type == "Smart Auto":
            ExcelFormulaHandler.smart_formula_handler(ws, row_num, 3, formula_text)
        elif formula_type == "Mathematical":
            ExcelFormulaHandler.add_executable_formula(ws, row_num, 3, formula_text)
        else:
            ws.cell(row=row_num, column=3, value=formula_text)
        
        row_num += 1
    
    # Enable calculation
    ExcelFormulaHandler.enable_workbook_calculation(wb)
    
    # Save the demo file
    output_path = "formula_demo.xlsx"
    wb.save(output_path)
    print(f"✓ Demo saved as {output_path}")
    
    return output_path


if __name__ == "__main__":
    # Run the demo
    demo_path = demo_formula_usage()
    print(f"""
Formula Demo Complete!

The demo file '{demo_path}' shows different formula handling approaches:

1. Text Display: Formulas shown as text (won't calculate)
2. Executable: Formulas that Excel will calculate
3. Smart Auto: Automatic detection of formula type
4. Mathematical: Simple math expressions
5. Regular Text: Just regular text values

Open the file in Excel to see the results!
""")
