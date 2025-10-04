import os
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from services.DBManager import DBManager
from data.db_data import db_data
from utils.excel_formula_handler import ExcelFormulaHandler

class CreateReports:
    
    def __init__(self, year, db_manager: DBManager= None):
        if db_manager is None:
            db_manager = DBManager()
        self.o_db = db_manager
        self.db_data = db_manager.db_data
        self.current_year = year
        db_manager.db_data.s_current_year = year
        self.current_railroad = None
        self.current_df = None
        self.sTitle_RR_YEAR = ""

        # Initialize all dt*** variables to empty DataFrames with None
        self.dtTitles = pd.DataFrame(None)
        
        self.dtFootnotes = pd.DataFrame(None)
        self.dtAValue = pd.DataFrame(None)
        self.dtAValue0_RR = pd.DataFrame(None)
        self.dtAValueRegion_RR = pd.DataFrame(None)
        self.dtPriceIndexes = pd.DataFrame(None)
        self.dtPriceIndexReferences = pd.DataFrame(None)
        self.dtRegression = pd.DataFrame(None)
        self.dtRegressionDependentVariable = pd.DataFrame(None)
        self.dtRegressionDistribution = pd.DataFrame(None)
        self.dtDefinitions = pd.DataFrame(None)
        self.dtCarTypeStatistics = pd.DataFrame(None)
        self.dtCarTypeStatisticsPart2 = pd.DataFrame(None)
        self.dtCarTypeStatisticsPart3 = pd.DataFrame(None)
        self.dtLineSourceText = pd.DataFrame(None)
        self.dtDataDictionary = pd.DataFrame(None)
        self.dtECodes = pd.DataFrame(None)
        self.rr_icc = None
        self.s_output_folder = os.path.join(parent_dir, "..", "reports")
    
    def initialize_report_for_a_railroad(self, rr_no):
        self.dtAValueRegion_RR = self.o_db.get_a_value_region_rr(rr_no, self.current_year)
        print("Loaded dtAValueRegion_RR. number of rows:", len(self.dtAValueRegion_RR))
        self.dtPriceIndexes = self.o_db.get_price_indexes(rr_no, str(self.current_year))
        print("Loaded dtPriceIndexes. number of rows:", len(self.dtPriceIndexes))        
        self.dtAValue = self.o_db.get_a_value(rr_no, self.current_year)
        print("Loaded dtAValue, number of rows:", len(self.dtAValue))
        self.dtCarTypeStatistics = self.o_db.get_car_type_statistics(rr_no)        
        print("Loaded dtCarTypeStatistics, number of rows:", len(self.dtCarTypeStatistics))
        self.dtCarTypeStatisticsPart2 = self.o_db.get_car_type_statistics_part2(rr_no)
        print("Loaded dtCarTypeStatisticsPart2, number of rows:", len(self.dtCarTypeStatisticsPart2))

    def initialize_report_format(self):
        print("Initializing report format...")
        self.dtTitles = self.o_db.get_custom_data("REPORT_TITLES")
        print("Loaded dtTitles. number of rows:", len(self.dtTitles))
        self.dtFootnotes = self.o_db.get_custom_data("REPORT_FOOTNOTES")
        print("Loaded dtFootnotes. number of rows:", len(self.dtFootnotes))
        self.dtAValue0_RR = self.o_db.get_a_value0_rr(self.current_year)
        print("Loaded dtAValue0_RR. number of rows:", len(self.dtAValue0_RR))
        self.dtPriceIndexReferences = self.o_db.get_custom_data("PRICE_INDEX_REFERENCE")
        print("Loaded dtPriceIndexReferences. number of rows:", len(self.dtPriceIndexReferences))
        self.dtRegression = self.o_db.get_custom_data("REGR")
        print("Loaded dtRegression. number of rows:", len(self.dtRegression))
        self.dtRegressionDependentVariable = self.o_db.get_custom_data("REGR_DEP_VAR")
        print("Loaded dtRegressionDependentVariable. number of rows:", len(self.dtRegressionDependentVariable))
        self.dtRegressionDistribution = self.o_db.get_custom_data("REGR_DISTRIB")
        print("Loaded dtRegressionDistribution. number of rows:", len(self.dtRegressionDistribution))
        self.dtDefinitions = self.o_db.get_custom_data("URCSDEF")
        print("Loaded dtDefinitions. number of rows:", len(self.dtDefinitions))
        self.dtCarTypeStatisticsPart3 = self.o_db.get_car_type_statistics_part3()
        print("Loaded dtCarTypeStatisticsPart3. number of rows:", len(self.dtCarTypeStatisticsPart3))
        self.dtLineSourceText = self.o_db.get_line_source_text()
        print("Loaded dtLineSourceText. number of rows:", len(self.dtLineSourceText))
        self.dtDataDictionary = self.o_db.get_data_dictionary(self.current_year)
        print("Loaded dtDataDictionary. number of rows:", len(self.dtDataDictionary))
        self.dtECodes = self.o_db.get_custom_data("ECODES")
        print("Loaded dtECodes. number of rows:", len(self.dtECodes))

    def print_initialized_dataframes(self):
        print("dtTitles:")
        print(self.dtTitles)
        print("dtFootnotes:")
        print(self.dtFootnotes)
        print("dtAValue:")
        print(self.dtAValue)
        print("dtAValue0_RR:")
        print(self.dtAValue0_RR)
        print("dtAValueRegion_RR:")
        print(self.dtAValueRegion_RR)
        print("dtPriceIndexes:")
        print(self.dtPriceIndexes)
        print("dtPriceIndexReferences:")
        print(self.dtPriceIndexReferences)
        print("dtRegression:")
        print(self.dtRegression)
        print("dtRegressionDependentVariable:")
        print(self.dtRegressionDependentVariable)
        print("dtRegressionDistribution:")
        print(self.dtRegressionDistribution)
        print("dtDefinitions:")
        print(self.dtDefinitions)
        print("dtCarTypeStatistics:")
        print(self.dtCarTypeStatistics)
        print("dtCarTypeStatisticsPart2:")
        print(self.dtCarTypeStatisticsPart2)
        print("dtCarTypeStatisticsPart3:")
        print(self.dtCarTypeStatisticsPart3)
        print("dtLineSourceText:")
        print(self.dtLineSourceText)
        print("dtDataDictionary:")
        print(self.dtDataDictionary)
        print("dtECodes:")
        print(self.dtECodes)

    def setup_report(self, output_folder, current_year, rr_no, rricc, rr_name, rr_longname,
                    null_string_replace, run_years, capital_cost, variability_flow, account76,
                    account80, account90, ssac, create_log, db):
        """
        Sets up report parameters and returns them as a dictionary (no class used).
        Equivalent to the C# constructor logic.
        """
        from datetime import datetime
        sFileName = f"{output_folder}\\{rr_name}{current_year}"
        sLogFileName = f"{output_folder}\\Log\\{rr_name}{current_year}.log"
        sTitle_RR_YEAR = f"{rr_name} {current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
        return {
            "oDB": db,
            "sFileName": sFileName,
            "sLogFileName": sLogFileName,
            "bCreateLog": create_log,
            "sNullValueString": null_string_replace,
            "iRunYears": run_years,
            "iRailroadNumber": rr_no,
            "iRRICC": rricc,
            "sRailRoadShortName": rr_name,
            "sRailRoadName": rr_longname,
            "bCapitalCost": capital_cost,
            "bVarialabilityFlow": variability_flow,
            "bAccount76": account76,
            "bAccount80": account80,
            "bAccount90": account90,
            "bSSAC": ssac,
            "iCurrentYear": int(current_year),
            "sTitle_RR_YEAR": sTitle_RR_YEAR
        }

    def create_workbook(self, workbookname, path):

        if not os.path.exists(path):
            os.makedirs(path)
        full_path = os.path.join(path, workbookname)
        wb = Workbook()
        wb.save(full_path)
        return wb, full_path

    def add_worksheet(self, workbook, sheet_name):
        """
        Adds a worksheet with the specified name to the provided openpyxl Workbook object.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
            sheet_name (str): The name of the worksheet to add.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        return workbook.create_sheet(title=sheet_name)

    def create_reports(self):

        df = self.o_db.get_class1_rail_list()
        count = 0
        self.initialize_report_format()
        for row in df.itertuples(index=False):
            count += 1
            if (count>1):
                break
            self.current_df = row
            short_name = row.short_name
            self.initialize_report_for_a_railroad(row.rr_id)
            self.create_a_report(short_name)

    def set_cell_with_format_and_name(self, ws, wb, row, col, value, number_format, alignment, named_range):
        # Convert value to int if possible for number formatting
        try:
            value = int(value)
        except (ValueError, TypeError):
            pass
        cell = ws.cell(row=row, column=col, value=value)
        cell.number_format = number_format
        cell.alignment = alignment
        wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{ws.title}'!${cell.column_letter}${cell.row}")
        return cell

    def add_formula_as_text(self, ws, row, col, formula_text):
        """Add a formula as display text (not executable) without quote marks"""
        cell = ws.cell(row=row, column=col)
        cell.number_format = '@'  # Text format
        cell.value = formula_text
        return cell

    def add_executable_formula(self, ws, row, col, formula_text):
        """Add an executable formula that Excel will calculate"""
        cell = ws.cell(row=row, column=col)
        # Ensure formula starts with = for Excel to execute it
        if not formula_text.startswith('='):
            formula_text = '=' + formula_text
        cell.value = formula_text
        return cell

    def force_excel_recalculation(self, wb):
        """
        Forces Excel to recalculate all formulas when the file is opened.
        This ensures that executable formulas are properly calculated.
        """
        try:
            # Set calculation mode to automatic
            wb.calculation.calcMode = 'automatic'
            # Force full calculation
            wb.calculation.fullCalcOnLoad = True
            print("Excel recalculation enabled for workbook")
        except AttributeError:
            # Fallback for older openpyxl versions
            print("Note: Automatic recalculation setting not available in this openpyxl version")

    def create_a_report(self, railroad_shortname):

        workbookname = f"{railroad_shortname}-{self.o_db.db_data.s_current_year}_report.xlsx"
        wb, full_path = self.create_workbook(workbookname,self.s_output_folder)
        print(f"Workbook {workbookname} created at: {full_path}")
        
        # Configure workbook for formula calculation
        self.force_excel_recalculation(wb)
        
        self.build_index_worksheet(wb)
        self.A1P1_worksheet(wb)
        wb.save(full_path)
        print(f"Workbook {workbookname} saved at: {full_path}")


    def build_index_worksheet(self, wb: Workbook):
        """
        Builds the 'INDEX' worksheet in the provided workbook, sets titles, and populates user input values with formatting.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        try:
            sTitle_WORKTABLE = "User Inputs and Index"
            sSheetTitle = "INDEX"
            ws = wb.create_sheet(title=sSheetTitle)

            # Set font for all cells
            default_font = Font(name="Consolas", size=10)
            for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=2):
                for cell in row:
                    cell.font = default_font
            sTitle_RR_YEAR = f"{self.current_df.name} {self.o_db.db_data.s_current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
            self.sTitle_RR_YEAR = sTitle_RR_YEAR
            ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
            ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)

            # Worktable title
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
            ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

            # User Input Table header
            ws.cell(row=4, column=1, value="User Input values:")
            ws.cell(row=4, column=1).font = Font(name="Consolas", size=12, bold=True)
            ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
            ws.merge_cells("A4:B4")
            # Add border and double bottom border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
            for cell in ws["A4:B4"]:
                for c in cell:
                    c.border = border

            # User input values
            user_inputs = [
                ("RRICC:", str(self.current_df.rricc), "RRICC"),
                ("Railroad Short Name:", str(self.current_df.short_name), "RR_SHORT_NAME"),
                ("Railroad Name:", str(self.current_df.name), "RR_NAME"),
                ("Railroad ID:", str(self.current_df.rr_id), "RR_ID"),
                ("Current Year:", str(self.o_db.db_data.s_current_year), "CURRENT_YEAR"),
                ("Run Years:", str(self.o_db.db_data.RUN_YEARS), "RUN_YEARS"),
                ("Embedded Cost of Capital for ROI:", "Y" if self.o_db.db_data.CapitalCost else "N", "EMBEDDED_COC"),
                ("100% Variability Flow Thru:", "Y" if self.o_db.db_data.VarialabilityFlow else "N", "FLOW_THRU_100PCT"),
                ("Include Account 76:", "Y" if self.o_db.db_data.Account76 else "N", "INCL_ACCT_76"),
                ("Include Account 80:", "Y" if self.o_db.db_data.Account80 else "N", "INCL_ACCT_80"),
                ("Include Account 90:", "Y" if self.o_db.db_data.Account90 else "N", "INCL_ACCT_90"),
                ("SSAC:", "Y" if self.o_db.db_data.SSAC else "N", "INCL_SSAC"),
                ("Replace Null Values With:", str(self.o_db.db_data.NullStringReplace), "NULL_VALUE"),
            ]
            for idx, (label, value, name) in enumerate(user_inputs, start=5):
                ws.cell(row=idx, column=1, value=label)
                cell = ws.cell(row=idx, column=2, value=value)
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(horizontal="right")
                ws.column_dimensions[cell.column_letter].width = 20
                # Add named range for this cell
                wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Set column width for A2 (worktable title)
            ws.column_dimensions['A'].width = 33

            # Save workbook after worksheet creation/modification
            if hasattr(wb, 'filename') and wb.filename:
                wb.save(wb.filename)
            else:
                default_path = os.path.join(self.s_output_folder, "temp_report.xlsx")
                wb.save(default_path)
                print(f"Workbook saved to {default_path}")

            return ws

        except Exception as ex:
            print(f"Error in build_index_worksheet: {ex}")
            return None

    def A1P1_worksheet(self, wb, index=2):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            iColumnCount = 21
            sSheetTitle = "A1P1"
            iLineNumberOffset = 93

            dtaValue = self.dtAValue
            dtaValue0_RR = self.dtAValue0_RR
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)
            oDB = self.o_db

            # Update status
            print(f"Processing {sSheetTitle}")

            
            # Select worktable and string rows
            try:
                sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]}'"
            except Exception:
                sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1]}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            # Get worksheet
            ws = wb.create_sheet(title=sSheetTitle)


            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)  
            # Freeze panes
            ws.freeze_panes = ws['D8']

            # Write titles and column headers (implement these as needed)
            # self.write_titles_and_column_headers(ws)
            # self.write_first3_columns_and_page_layout(ws)

            # Write hard values from dtaValue for all lines but 158
            count = 0
            col_map = {
                        iCurrentYear: 5,
                        iCurrentYear - 1: 7,
                        iCurrentYear - 2: 9,
                        iCurrentYear - 3: 11,
                        iCurrentYear - 4: 13
                    }
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] != 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                    count += 1
                    print(f"Processing row {count}, Line: {draValues['aline']}, Year: {iProcessYear}, Value: {draValues['value']}")  # Debug print
                    if iProcessYear in col_map:
                        col = col_map[iProcessYear]
                        # Calculate the 'C' number for the named range, similar to the original VB
                        c_num = iCurrentYear - iProcessYear + 1
                        ws.cell(row=iROW_COUNT, column=col, value=str(draValues["value"]))
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}c{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"
                        # cell.name = sNamedRange  # openpyxl does not support cell naming directly

            # Write hard values from dtaValue0_RR for line 158
            for _, draValues in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] == 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset

                    
                    if iProcessYear in col_map:
                        c_num = iCurrentYear - iProcessYear + 1
                        col = col_map[iProcessYear]
                        value = str(draValues["value"]) if iProcessYear == iCurrentYear else "0"
                        ws.cell(row=iROW_COUNT, column=col, value=value)
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}c{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"

            # Write sources and values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                
                # Sources first
                source_cols = {
                    4: "c1", 6: "c2", 8: "c3", 10: "c4", 12: "c5",
                    14: "c6", 16: "c7", 18: "c8", 20: "c9"
                }
                for col, c_name in source_cols.items():
                    source_text = self.scrub_year(str(drSource.get(c_name, "")))
                    # Check if this should be an executable formula or display text
                    if source_text.startswith('=') or source_text.startswith('+'):
                        # This looks like a formula - decide if executable or text display
                        self.add_formula_as_text(ws, iLine, col, source_text)
                    else:
                        # Regular text
                        ws.cell(row=iLine, column=col, value=source_text)

                # C6 - C9 Values
                value_cols_c6_c9 = {15: "c6", 17: "c7", 19: "c8",  21: "c9"}
                for col, c_name in value_cols_c6_c9.items():
                    value = "0" if drSource["line"] == 158 else str(drSource[c_name])
                    c_num = (col - 15) // 2 + 6
                    ws.cell(row=iLine, column=col, value=value)
                    cell = ws.cell(row=iLine, column=col)
                    cell.number_format = "#,##0"
                    named_range_name = f"A1L{drSource['line']}c{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    

                # Random values not from Data Dictionary
                random_lines = {111, 122, 131, 133, 137, 144, 145, 146, 151, 154, 155}
                if int(drSource["line"]) in random_lines:
                    random_value_cols = {5: "c1", 7: "c2", 9: "c3", 11: "c4", 13: "c5"} # This corresponds to the VB code block for "Random Values"
                    for col, c_name in random_value_cols.items():
                        c_num = (col - 5) // 2 + 1
                        ws.cell(row=iLine, column=col, value=str(drSource.get(c_name, "")))
                        cell = ws.cell(row=iLine, column=col)
                        cell.number_format = "#,##0"
                        named_range_name = f"A1L{drSource['line']}C{c_num}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        

            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in A1P1: {ex}")
            import traceback
            print(traceback.format_exc())

    def add_formula_as_text(self, ws, row, col, formula_text):
        """Add a formula as display text (not executable) without quote marks"""
        cell = ws.cell(row=row, column=col)
        cell.number_format = '@'  # Text format
        cell.value = formula_text
        return cell

    def add_executable_formula(self, ws, row, col, formula_text):
        """Add an executable formula that Excel will calculate"""
        cell = ws.cell(row=row, column=col)
        # Ensure formula starts with = for Excel to execute it
        if not formula_text.startswith('='):
            formula_text = '=' + formula_text
        cell.value = formula_text
        return cell

    def force_excel_recalculation(self, wb):
        """
        Forces Excel to recalculate all formulas when the file is opened.
        This ensures that executable formulas are properly calculated.
        """
        try:
            # Set calculation mode to automatic
            wb.calculation.calcMode = 'automatic'
            # Force full calculation
            wb.calculation.fullCalcOnLoad = True
            print("Excel recalculation enabled for workbook")
        except AttributeError:
            # Fallback for older openpyxl versions
            print("Note: Automatic recalculation setting not available in this openpyxl version")

    def create_a_report(self, railroad_shortname):

        workbookname = f"{railroad_shortname}-{self.o_db.db_data.s_current_year}_report.xlsx"
        wb, full_path = self.create_workbook(workbookname,self.s_output_folder)
        print(f"Workbook {workbookname} created at: {full_path}")
        
        # Configure workbook for formula calculation
        self.force_excel_recalculation(wb)
        
        self.build_index_worksheet(wb)
        self.A1P1_worksheet(wb)
        wb.save(full_path)
        print(f"Workbook {workbookname} saved at: {full_path}")


    def build_index_worksheet(self, wb: Workbook):
        """
        Builds the 'INDEX' worksheet in the provided workbook, sets titles, and populates user input values with formatting.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        try:
            sTitle_WORKTABLE = "User Inputs and Index"
            sSheetTitle = "INDEX"
            ws = wb.create_sheet(title=sSheetTitle)

            # Set font for all cells
            default_font = Font(name="Consolas", size=10)
            for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=2):
                for cell in row:
                    cell.font = default_font
            sTitle_RR_YEAR = f"{self.current_df.name} {self.o_db.db_data.s_current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
            self.sTitle_RR_YEAR = sTitle_RR_YEAR
            ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
            ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)

            # Worktable title
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
            ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

            # User Input Table header
            ws.cell(row=4, column=1, value="User Input values:")
            ws.cell(row=4, column=1).font = Font(name="Consolas", size=12, bold=True)
            ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
            ws.merge_cells("A4:B4")
            # Add border and double bottom border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
            for cell in ws["A4:B4"]:
                for c in cell:
                    c.border = border

            # User input values
            user_inputs = [
                ("RRICC:", str(self.current_df.rricc), "RRICC"),
                ("Railroad Short Name:", str(self.current_df.short_name), "RR_SHORT_NAME"),
                ("Railroad Name:", str(self.current_df.name), "RR_NAME"),
                ("Railroad ID:", str(self.current_df.rr_id), "RR_ID"),
                ("Current Year:", str(self.o_db.db_data.s_current_year), "CURRENT_YEAR"),
                ("Run Years:", str(self.o_db.db_data.RUN_YEARS), "RUN_YEARS"),
                ("Embedded Cost of Capital for ROI:", "Y" if self.o_db.db_data.CapitalCost else "N", "EMBEDDED_COC"),
                ("100% Variability Flow Thru:", "Y" if self.o_db.db_data.VarialabilityFlow else "N", "FLOW_THRU_100PCT"),
                ("Include Account 76:", "Y" if self.o_db.db_data.Account76 else "N", "INCL_ACCT_76"),
                ("Include Account 80:", "Y" if self.o_db.db_data.Account80 else "N", "INCL_ACCT_80"),
                ("Include Account 90:", "Y" if self.o_db.db_data.Account90 else "N", "INCL_ACCT_90"),
                ("SSAC:", "Y" if self.o_db.db_data.SSAC else "N", "INCL_SSAC"),
                ("Replace Null Values With:", str(self.o_db.db_data.NullStringReplace), "NULL_VALUE"),
            ]
            for idx, (label, value, name) in enumerate(user_inputs, start=5):
                ws.cell(row=idx, column=1, value=label)
                cell = ws.cell(row=idx, column=2, value=value)
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(horizontal="right")
                ws.column_dimensions[cell.column_letter].width = 20
                # Add named range for this cell
                wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Set column width for A2 (worktable title)
            ws.column_dimensions['A'].width = 33

            # Save workbook after worksheet creation/modification
            if hasattr(wb, 'filename') and wb.filename:
                wb.save(wb.filename)
            else:
                default_path = os.path.join(self.s_output_folder, "temp_report.xlsx")
                wb.save(default_path)
                print(f"Workbook saved to {default_path}")

            return ws

        except Exception as ex:
            print(f"Error in build_index_worksheet: {ex}")
            return None

    def A1P1_worksheet(self, wb, index=2):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            iColumnCount = 21
            sSheetTitle = "A1P1"
            iLineNumberOffset = 93

            dtaValue = self.dtAValue
            dtaValue0_RR = self.dtAValue0_RR
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)
            oDB = self.o_db

            # Update status
            print(f"Processing {sSheetTitle}")

            
            # Select worktable and string rows
            try:
                sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]}'"
            except Exception:
                sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1]}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            # Get worksheet
            ws = wb.create_sheet(title=sSheetTitle)


            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)  
            # Freeze panes
            ws.freeze_panes = ws['D8']

            # Write titles and column headers (implement these as needed)
            # self.write_titles_and_column_headers(ws)
            # self.write_first3_columns_and_page_layout(ws)

            # Write hard values from dtaValue for all lines but 158
            count = 0
            col_map = {
                        iCurrentYear: 5,
                        iCurrentYear - 1: 7,
                        iCurrentYear - 2: 9,
                        iCurrentYear - 3: 11,
                        iCurrentYear - 4: 13
                    }
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] != 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                    count += 1
                    print(f"Processing row {count}, Line: {draValues['aline']}, Year: {iProcessYear}, Value: {draValues['value']}")  # Debug print
                    if iProcessYear in col_map:
                        col = col_map[iProcessYear]
                        # Calculate the 'C' number for the named range, similar to the original VB
                        c_num = iCurrentYear - iProcessYear + 1
                        ws.cell(row=iROW_COUNT, column=col, value=str(draValues["value"]))
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}c{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"
                        # cell.name = sNamedRange  # openpyxl does not support cell naming directly

            # Write hard values from dtaValue0_RR for line 158
            for _, draValues in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] == 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset

                    
                    if iProcessYear in col_map:
                        c_num = iCurrentYear - iProcessYear + 1
                        col = col_map[iProcessYear]
                        value = str(draValues["value"]) if iProcessYear == iCurrentYear else "0"
                        ws.cell(row=iROW_COUNT, column=col, value=value)
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}c{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"

            # Write sources and values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                
                # Sources first
                source_cols = {
                    4: "c1", 6: "c2", 8: "c3", 10: "c4", 12: "c5",
                    14: "c6", 16: "c7", 18: "c8", 20: "c9"
                }
                for col, c_name in source_cols.items():
                    source_text = self.scrub_year(str(drSource.get(c_name, "")))
                    # Check if this should be an executable formula or display text
                    if source_text.startswith('=') or source_text.startswith('+'):
                        # This looks like a formula - decide if executable or text display
                        self.add_formula_as_text(ws, iLine, col, source_text)
                    else:
                        # Regular text
                        ws.cell(row=iLine, column=col, value=source_text)

                # C6 - C9 Values
                value_cols_c6_c9 = {15: "c6", 17: "c7", 19: "c8",  21: "c9"}
                for col, c_name in value_cols_c6_c9.items():
                    value = "0" if drSource["line"] == 158 else str(drSource[c_name])
                    c_num = (col - 15) // 2 + 6
                    ws.cell(row=iLine, column=col, value=value)
                    cell = ws.cell(row=iLine, column=col)
                    cell.number_format = "#,##0"
                    named_range_name = f"A1L{drSource['line']}c{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    

                # Random values not from Data Dictionary
                random_lines = {111, 122, 131, 133, 137, 144, 145, 146, 151, 154, 155}
                if int(drSource["line"]) in random_lines:
                    random_value_cols = {5: "c1", 7: "c2", 9: "c3", 11: "c4", 13: "c5"} # This corresponds to the VB code block for "Random Values"
                    for col, c_name in random_value_cols.items():
                        c_num = (col - 5) // 2 + 1
                        ws.cell(row=iLine, column=col, value=str(drSource.get(c_name, "")))
                        cell = ws.cell(row=iLine, column=col)
                        cell.number_format = "#,##0"
                        named_range_name = f"A1L{drSource['line']}C{c_num}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        

            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in A1P1: {ex}")
            import traceback
            print(traceback.format_exc())

    def add_formula_as_text(self, ws, row, col, formula_text):
        """Add a formula as display text (not executable) without quote marks"""
        cell = ws.cell(row=row, column=col)
        cell.number_format = '@'  # Text format
        cell.value = formula_text
        return cell

    def add_executable_formula(self, ws, row, col, formula_text):
        """Add an executable formula that Excel will calculate"""
        cell = ws.cell(row=row, column=col)
        # Ensure formula starts with = for Excel to execute it
        if not formula_text.startswith('='):
            formula_text = '=' + formula_text
        cell.value = formula_text
        return cell

    def force_excel_recalculation(self, wb):
        """
        Forces Excel to recalculate all formulas when the file is opened.
        This ensures that executable formulas are properly calculated.
        """
        try:
            # Set calculation mode to automatic
            wb.calculation.calcMode = 'automatic'
            # Force full calculation
            wb.calculation.fullCalcOnLoad = True
            print("Excel recalculation enabled for workbook")
        except AttributeError:
            # Fallback for older openpyxl versions
            print("Note: Automatic recalculation setting not available in this openpyxl version")

    def create_a_report(self, railroad_shortname):

        workbookname = f"{railroad_shortname}-{self.o_db.db_data.s_current_year}_report.xlsx"
        wb, full_path = self.create_workbook(workbookname,self.s_output_folder)
        print(f"Workbook {workbookname} created at: {full_path}")
        
        # Configure workbook for formula calculation
        self.force_excel_recalculation(wb)
        
        self.build_index_worksheet(wb)
        self.A1P1_worksheet(wb)
        wb.save(full_path)
        print(f"Workbook {workbookname} saved at: {full_path}")


    def build_index_worksheet(self, wb: Workbook):
        """
        Builds the 'INDEX' worksheet in the provided workbook, sets titles, and populates user input values with formatting.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        try:
            sTitle_WORKTABLE = "User Inputs and Index"
            sSheetTitle = "INDEX"
            ws = wb.create_sheet(title=sSheetTitle)

            # Set font for all cells
            default_font = Font(name="Consolas", size=10)
            for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=2):
                for cell in row:
                    cell.font = default_font
            sTitle_RR_YEAR = f"{self.current_df.name} {self.o_db.db_data.s_current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
            self.sTitle_RR_YEAR = sTitle_RR_YEAR
            ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
            ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)

            # Worktable title
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
            ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

            # User Input Table header
            ws.cell(row=4, column=1, value="User Input values:")
            ws.cell(row=4, column=1).font = Font(name="Consolas", size=12, bold=True)
            ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
            ws.merge_cells("A4:B4")
            # Add border and double bottom border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
            for cell in ws["A4:B4"]:
                for c in cell:
                    c.border = border

            # User input values
            user_inputs = [
                ("RRICC:", str(self.current_df.rricc), "RRICC"),
                ("Railroad Short Name:", str(self.current_df.short_name), "RR_SHORT_NAME"),
                ("Railroad Name:", str(self.current_df.name), "RR_NAME"),
                ("Railroad ID:", str(self.current_df.rr_id), "RR_ID"),
                ("Current Year:", str(self.o_db.db_data.s_current_year), "CURRENT_YEAR"),
                ("Run Years:", str(self.o_db.db_data.RUN_YEARS), "RUN_YEARS"),
                ("Embedded Cost of Capital for ROI:", "Y" if self.o_db.db_data.CapitalCost else "N", "EMBEDDED_COC"),
                ("100% Variability Flow Thru:", "Y" if self.o_db.db_data.VarialabilityFlow else "N", "FLOW_THRU_100PCT"),
                ("Include Account 76:", "Y" if self.o_db.db_data.Account76 else "N", "INCL_ACCT_76"),
                ("Include Account 80:", "Y" if self.o_db.db_data.Account80 else "N", "INCL_ACCT_80"),
                ("Include Account 90:", "Y" if self.o_db.db_data.Account90 else "N", "INCL_ACCT_90"),
                ("SSAC:", "Y" if self.o_db.db_data.SSAC else "N", "INCL_SSAC"),
                ("Replace Null Values With:", str(self.o_db.db_data.NullStringReplace), "NULL_VALUE"),
            ]
            for idx, (label, value, name) in enumerate(user_inputs, start=5):
                ws.cell(row=idx, column=1, value=label)
                cell = ws.cell(row=idx, column=2, value=value)
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(horizontal="right")
                ws.column_dimensions[cell.column_letter].width = 20
                # Add named range for this cell
                wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Set column width for A2 (worktable title)
            ws.column_dimensions['A'].width = 33

            # Save workbook after worksheet creation/modification
            if hasattr(wb, 'filename') and wb.filename:
                wb.save(wb.filename)
            else:
                default_path = os.path.join(self.s_output_folder, "temp_report.xlsx")
                wb.save(default_path)
                print(f"Workbook saved to {default_path}")

            return ws

        except Exception as ex:
            print(f"Error in build_index_worksheet: {ex}")
            return None

    def A1P1_worksheet(self, wb, index=2):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            iColumnCount = 21
            sSheetTitle = "A1P1"
            iLineNumberOffset = 93

            dtaValue = self.dtAValue
            dtaValue0_RR = self.dtAValue0_RR
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)
            oDB = self.o_db

            # Update status
            print(f"Processing {sSheetTitle}")

            
            # Select worktable and string rows
            try:
                sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]}'"
            except Exception:
                sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1]}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            # Get worksheet
            ws = wb.create_sheet(title=sSheetTitle)


            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)  
            # Freeze panes
            ws.freeze_panes = ws['D8']

            # Write titles and column headers (implement these as needed)
            # self.write_titles_and_column_headers(ws)
            # self.write_first3_columns_and_page_layout(ws)

            # Write hard values from dtaValue for all lines but 158
            count = 0
            col_map = {
                        iCurrentYear: 5,
                        iCurrentYear - 1: 7,
                        iCurrentYear - 2: 9,
                        iCurrentYear - 3: 11,
                        iCurrentYear - 4: 13
                    }
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] != 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                    count += 1
                    print(f"Processing row {count}, Line: {draValues['aline']}, Year: {iProcessYear}, Value: {draValues['value']}")  # Debug print
                    if iProcessYear in col_map:
                        col = col_map[iProcessYear]
                        # Calculate the 'C' number for the named range, similar to the original VB
                        c_num = iCurrentYear - iProcessYear + 1
                        ws.cell(row=iROW_COUNT, column=col, value=str(draValues["value"]))
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}c{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"
                        # cell.name = sNamedRange  # openpyxl does not support cell naming directly

            # Write hard values from dtaValue0_RR for line 158
            for _, draValues in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] == 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset

                    
                    if iProcessYear in col_map:
                        c_num = iCurrentYear - iProcessYear + 1
                        col = col_map[iProcessYear]
                        value = str(draValues["value"]) if iProcessYear == iCurrentYear else "0"
                        ws.cell(row=iROW_COUNT, column=col, value=value)
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}c{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"

            # Write sources and values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                
                # Sources first
                source_cols = {
                    4: "c1", 6: "c2", 8: "c3", 10: "c4", 12: "c5",
                    14: "c6", 16: "c7", 18: "c8", 20: "c9"
                }
                for col, c_name in source_cols.items():
                    source_text = self.scrub_year(str(drSource.get(c_name, "")))
                    # Check if this should be an executable formula or display text
                    if source_text.startswith('=') or source_text.startswith('+'):
                        # This looks like a formula - decide if executable or text display
                        self.add_formula_as_text(ws, iLine, col, source_text)
                    else:
                        # Regular text
                        ws.cell(row=iLine, column=col, value=source_text)

                # C6 - C9 Values
                value_cols_c6_c9 = {15: "c6", 17: "c7", 19: "c8",  21: "c9"}
                for col, c_name in value_cols_c6_c9.items():
                    value = "0" if drSource["line"] == 158 else str(drSource[c_name])
                    c_num = (col - 15) // 2 + 6
                    ws.cell(row=iLine, column=col, value=value)
                    cell = ws.cell(row=iLine, column=col)
                    cell.number_format = "#,##0"
                    named_range_name = f"A1L{drSource['line']}c{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    

                # Random values not from Data Dictionary
                random_lines = {111, 122, 131, 133, 137, 144, 145, 146, 151, 154, 155}
                if int(drSource["line"]) in random_lines:
                    random_value_cols = {5: "c1", 7: "c2", 9: "c3", 11: "c4", 13: "c5"} # This corresponds to the VB code block for "Random Values"
                    for col, c_name in random_value_cols.items():
                        c_num = (col - 5) // 2 + 1
                        ws.cell(row=iLine, column=col, value=str(drSource.get(c_name, "")))
                        cell = ws.cell(row=iLine, column=col)
                        cell.number_format = "#,##0"
                        named_range_name = f"A1L{drSource['line']}C{c_num}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        

            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in A1P1: {ex}")
            import traceback
            print(traceback.format_exc())
            
            
            
if __name__ == "__main__":
    
    create_reports = CreateReports(2023)
    create_reports.create_reports()  # Call the method to create reports