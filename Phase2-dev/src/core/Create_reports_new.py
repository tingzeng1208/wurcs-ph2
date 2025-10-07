import os
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
from dotenv import load_dotenv
import win32com.client

load_dotenv()

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from services.DBManager import DBManager
from data.db_data import db_data
from utils.utility import to_str

class CreateReports:
    
    def __init__(self, year, db_manager: DBManager= None):
        if db_manager is None:
            db_manager = DBManager()
        self.o_db = db_manager
        self.db_data = db_manager.db_data
        self.current_year = year
        db_manager.db_data.s_current_year = year
        self.current_railroad = None
        self.current_df = None
        self.sTitle_RR_YEAR = ""

        # Initialize all dt*** variables to empty DataFrames with None
        self.dtTitles = pd.DataFrame(None)
        
        self.dtFootnotes = pd.DataFrame(None)
        self.dtAValue = pd.DataFrame(None)
        self.dtAValue0_RR = pd.DataFrame(None)
        self.dtAValueRegion_RR = pd.DataFrame(None)
        self.dtPriceIndexes = pd.DataFrame(None)
        self.dtPriceIndexReferences = pd.DataFrame(None)
        self.dtRegression = pd.DataFrame(None)
        self.dtRegressionDependentVariable = pd.DataFrame(None)
        self.dtRegressionDistribution = pd.DataFrame(None)
        self.dtDefinitions = pd.DataFrame(None)
        self.dtCarTypeStatistics = pd.DataFrame(None)
        self.dtCarTypeStatisticsPart2 = pd.DataFrame(None)
        self.dtCarTypeStatisticsPart3 = pd.DataFrame(None)
        self.dtLineSourceText = pd.DataFrame(None)
        self.dtDataDictionary = pd.DataFrame(None)
        self.dtECodes = pd.DataFrame(None)
        self.rr_icc = None
        self.s_output_folder = os.path.join(parent_dir, "..", "reports")
        self.fontname = "Consolas"
        self.fontsize = 10

    def initialize_report_for_a_railroad(self, rr_no):
        self.dtAValueRegion_RR = self.o_db.get_a_value_region_rr(rr_no, self.current_year)
        print("Loaded dtAValueRegion_RR. number of rows:", len(self.dtAValueRegion_RR))
        self.dtPriceIndexes = self.o_db.get_price_indexes(rr_no, str(self.current_year))
        print("Loaded dtPriceIndexes. number of rows:", len(self.dtPriceIndexes))        
        self.dtAValue = self.o_db.get_a_value(rr_no, self.current_year)
        print("Loaded dtAValue, number of rows:", len(self.dtAValue))
        self.dtCarTypeStatistics = self.o_db.get_car_type_statistics(rr_no)        
        print("Loaded dtCarTypeStatistics, number of rows:", len(self.dtCarTypeStatistics))
        self.dtCarTypeStatisticsPart2 = self.o_db.get_car_type_statistics_part2(rr_no)
        print("Loaded dtCarTypeStatisticsPart2, number of rows:", len(self.dtCarTypeStatisticsPart2))

    def initialize_report_format(self):
        print("Initializing report format...")
        self.dtTitles = self.o_db.get_custom_data("REPORT_TITLES")
        print("Loaded dtTitles. number of rows:", len(self.dtTitles))
        self.dtFootnotes = self.o_db.get_custom_data("REPORT_FOOTNOTES")
        print("Loaded dtFootnotes. number of rows:", len(self.dtFootnotes))
        self.dtAValue0_RR = self.o_db.get_a_value0_rr(self.current_year)
        print("Loaded dtAValue0_RR. number of rows:", len(self.dtAValue0_RR))
        self.dtPriceIndexReferences = self.o_db.get_custom_data("PRICE_INDEX_REFERENCE")
        print("Loaded dtPriceIndexReferences. number of rows:", len(self.dtPriceIndexReferences))
        self.dtRegression = self.o_db.get_custom_data("REGR")
        print("Loaded dtRegression. number of rows:", len(self.dtRegression))
        self.dtRegressionDependentVariable = self.o_db.get_custom_data("REGR_DEP_VAR")
        print("Loaded dtRegressionDependentVariable. number of rows:", len(self.dtRegressionDependentVariable))
        self.dtRegressionDistribution = self.o_db.get_custom_data("REGR_DISTRIB")
        print("Loaded dtRegressionDistribution. number of rows:", len(self.dtRegressionDistribution))
        self.dtDefinitions = self.o_db.get_custom_data("URCSDEF")
        print("Loaded dtDefinitions. number of rows:", len(self.dtDefinitions))
        self.dtCarTypeStatisticsPart3 = self.o_db.get_car_type_statistics_part3()
        print("Loaded dtCarTypeStatisticsPart3. number of rows:", len(self.dtCarTypeStatisticsPart3))
        self.dtLineSourceText = self.o_db.get_line_source_text()
        print("Loaded dtLineSourceText. number of rows:", len(self.dtLineSourceText))
        self.dtDataDictionary = self.o_db.get_data_dictionary(self.current_year)
        print("Loaded dtDataDictionary. number of rows:", len(self.dtDataDictionary))
        self.dtECodes = self.o_db.get_custom_data("ECODES")
        print("Loaded dtECodes. number of rows:", len(self.dtECodes))

    def print_initialized_dataframes(self):
        print("dtTitles:")
        print(self.dtTitles)
        print("dtFootnotes:")
        print(self.dtFootnotes)
        print("dtAValue:")
        print(self.dtAValue)
        print("dtAValue0_RR:")
        print(self.dtAValue0_RR)
        print("dtAValueRegion_RR:")
        print(self.dtAValueRegion_RR)
        print("dtPriceIndexes:")
        print(self.dtPriceIndexes)
        print("dtPriceIndexReferences:")
        print(self.dtPriceIndexReferences)
        print("dtRegression:")
        print(self.dtRegression)
        print("dtRegressionDependentVariable:")
        print(self.dtRegressionDependentVariable)
        print("dtRegressionDistribution:")
        print(self.dtRegressionDistribution)
        print("dtDefinitions:")
        print(self.dtDefinitions)
        print("dtCarTypeStatistics:")
        print(self.dtCarTypeStatistics)
        print("dtCarTypeStatisticsPart2:")
        print(self.dtCarTypeStatisticsPart2)
        print("dtCarTypeStatisticsPart3:")
        print(self.dtCarTypeStatisticsPart3)
        print("dtLineSourceText:")
        print(self.dtLineSourceText)
        print("dtDataDictionary:")
        print(self.dtDataDictionary)
        print("dtECodes:")
        print(self.dtECodes)

    def setup_report(self, output_folder, current_year, rr_no, rricc, rr_name, rr_longname,
                    null_string_replace, run_years, capital_cost, variability_flow, account76,
                    account80, account90, ssac, create_log, db):
        """
        Sets up report parameters and returns them as a dictionary (no class used).
        Equivalent to the C# constructor logic.
        """
        from datetime import datetime
        sFileName = f"{output_folder}\\{rr_name}{current_year}"
        sLogFileName = f"{output_folder}\\Log\\{rr_name}{current_year}.log"
        sTitle_RR_YEAR = f"{rr_name} {current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
        return {
            "oDB": db,
            "sFileName": sFileName,
            "sLogFileName": sLogFileName,
            "bCreateLog": create_log,
            "sNullValueString": null_string_replace,
            "iRunYears": run_years,
            "iRailroadNumber": rr_no,
            "iRRICC": rricc,
            "sRailRoadShortName": rr_name,
            "sRailRoadName": rr_longname,
            "bCapitalCost": capital_cost,
            "bVarialabilityFlow": variability_flow,
            "bAccount76": account76,
            "bAccount80": account80,
            "bAccount90": account90,
            "bSSAC": ssac,
            "iCurrentYear": int(current_year),
            "sTitle_RR_YEAR": sTitle_RR_YEAR
        }
        
    def remove_default_sheet(self, wb):
        # Remove the default sheet if it exists
        default_sheetnames = ["Sheet", "Sheet1"]
        for sheetname in default_sheetnames:
            if sheetname in wb.sheetnames:
                std = wb[sheetname]
                wb.remove(std)

    def create_workbook(self, workbookname, path):

        if not os.path.exists(path):
            os.makedirs(path)
        full_path = os.path.join(path, workbookname)
        wb = Workbook()
        # The default sheet "Sheet" is created. It will be removed later.
        wb.save(full_path)
        return wb, full_path

    def add_worksheet(self, workbook, sheet_name):
        """
        Adds a worksheet with the specified name to the provided openpyxl Workbook object.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
            sheet_name (str): The name of the worksheet to add.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        return workbook.create_sheet(title=sheet_name)

    def create_reports(self):

        df = self.o_db.get_class1_rail_list()
        count = 0
        self.initialize_report_format()
        for row in df.itertuples(index=False):
            count += 1
            if (count>1):
                break
            self.current_df = row
            short_name = row.short_name
            self.initialize_report_for_a_railroad(row.rr_id)
            self.create_a_report(short_name)

    def format_all_cells(self, ws, fontname=None, fontsize=None):
        """
        Format all cells in the worksheet with the given font name and size.
        If fontname or fontsize is None, use self.fontname and self.fontsize.
        """
        if fontname is None:
            fontname = getattr(self, "fontname", "Consolas")
        if fontsize is None:
            fontsize = getattr(self, "fontsize", 10)
        font = Font(name=fontname, size=fontsize)
        for row in ws.iter_rows():
            if row and row[0].row > 5:
                for cell in row:
                    cell.font = font

    def create_a_report(self, railroad_shortname):

        workbookname = f"{railroad_shortname}-{self.o_db.db_data.s_current_year}_report.xlsx"
        wb, full_path = self.create_workbook(workbookname,self.s_output_folder)
        print(f"Workbook {workbookname} created at: {full_path}")
        
        # Configure workbook for formula calculation
        
        self.build_index_worksheet(wb)
        self.A1P1_worksheet(wb)
        self.A1P2A_worksheet(wb)
        self.A1P2B_worksheet(wb)
        self.A1P2C_worksheet(wb)
        self.A1P3A_worksheet(wb)
        self.A1P3B_worksheet(wb)
        self.A1P4_worksheet(wb)
        self.A1P5A_worksheet(wb)

        self.A1P5B_worksheet(wb)
        self.A1P6_worksheet(wb)
        self.A1P7_worksheet(wb)
        self.A1P8_worksheet(wb)
        self.A1P9_worksheet(wb)
        self.A2P1_worksheet(wb)

        self.remove_default_sheet(wb)
        wb.save(full_path)
        self.recalculate_excel_formulas(full_path)
        print(f"Workbook {workbookname} saved at: {full_path}")


    def build_index_worksheet(self, wb: Workbook):
        """
        Builds the 'INDEX' worksheet in the provided workbook, sets titles, and populates user input values with formatting.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        try:
            sTitle_WORKTABLE = "User Inputs and Index"
            sSheetTitle = "INDEX"
            ws = wb.create_sheet(title=sSheetTitle)

            # Set font for all cells
            default_font = Font(name="Consolas", size=10)
            for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=2):
                for cell in row:
                    cell.font = default_font
            sTitle_RR_YEAR = f"{self.current_df.short_name} {self.o_db.db_data.s_current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
            self.sTitle_RR_YEAR = sTitle_RR_YEAR
            ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
            ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)

            # Worktable title
            ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
            ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

            # User Input Table header
            ws.cell(row=4, column=1, value="User Input values:")
            ws.cell(row=4, column=1).font = Font(name="Consolas", size=12, bold=True)
            ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
            ws.merge_cells("A4:B4")
            # Add border and double bottom border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
            for cell in ws["A4:B4"]:
                for c in cell:
                    c.border = border

            # User input values
            user_inputs = [
                ("RRICC:", str(self.current_df.rricc), "RRICC"),
                ("Railroad Short Name:", str(self.current_df.short_name), "RR_SHORT_NAME"),
                ("Railroad Name:", str(self.current_df.name), "RR_NAME"),
                ("Railroad ID:", str(self.current_df.rr_id), "RR_ID"),
                ("Current Year:", str(self.o_db.db_data.s_current_year), "CURRENT_YEAR"),
                ("Run Years:", str(self.o_db.db_data.RUN_YEARS), "RUN_YEARS"),
                ("Embedded Cost of Capital for ROI:", "Y" if self.o_db.db_data.CapitalCost else "N", "EMBEDDED_COC"),
                ("100% Variability Flow Thru:", "Y" if self.o_db.db_data.VarialabilityFlow else "N", "FLOW_THRU_100PCT"),
                ("Include Account 76:", "Y" if self.o_db.db_data.Account76 else "N", "INCL_ACCT_76"),
                ("Include Account 80:", "Y" if self.o_db.db_data.Account80 else "N", "INCL_ACCT_80"),
                ("Include Account 90:", "Y" if self.o_db.db_data.Account90 else "N", "INCL_ACCT_90"),
                ("SSAC:", "Y" if self.o_db.db_data.SSAC else "N", "INCL_SSAC"),
                ("Replace Null Values With:", str(self.o_db.db_data.NullStringReplace), "NULL_VALUE"),
            ]
            for idx, (label, value, name) in enumerate(user_inputs, start=5):
                ws.cell(row=idx, column=1, value=label)
                cell = ws.cell(row=idx, column=2, value=value)
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(horizontal="right")
                ws.column_dimensions[cell.column_letter].width = 20
                # Add named range for this cell
                wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Set column width for A2 (worktable title)
            ws.column_dimensions['A'].width = 33

            return ws

        except Exception as ex:
            print(f"Error in build_index_worksheet: {ex}")
            return None

    def WriteFirst3ColumnsAndPageLayout(self, ws, dtLineSourceText, dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable):
        # Set page layout
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.zoom = 70
        ws.oddFooter.left.text = "Page &P of &N"
        ws.print_title_rows = '$1:$7'
        ws.print_title_cols = 'A:C'

        # Write first 3 columns from dtLineSourceText
        iROW_COUNT = 8
        for _, drRow in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            ws.cell(row=iROW_COUNT, column=1, value=to_str(drRow["line"]))
            ws.cell(row=iROW_COUNT, column=2, value=to_str(drRow["code"]))
            ws.cell(row=iROW_COUNT, column=3, value=to_str(drRow["identification"]))
            iROW_COUNT += 1

        # Leave two rows, then write footnotes
        iROW_COUNT += 2
        for _, drRow in dtFootnotes[
            (dtFootnotes["worktable"] == sSheetTitle[:2]) &
            (dtFootnotes["part"] == sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3])
].iterrows():
            ws.cell(row=iROW_COUNT, column=2, value=to_str(drRow["no"]))
            ws.cell(row=iROW_COUNT, column=3, value=to_str(drRow["text"]))
            iROW_COUNT += 1
            
    
            
    def write_titles_and_column_headers(self, ws, dtTitles, sSelectWorktable, sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle):
        # Set font for all cells
        for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=iColumnCount):
            for cell in row:
                cell.font = Font(name="Consolas", size=10)

        # Title rows
        ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
        ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)
        ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
        ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

        # Get matching title rows
        part_str = sSheetTitle[sSheetTitle.find('P')+1:]
        filtered_titles = dtTitles[(dtTitles["worktable"] == sSheetTitle[:2]) & (dtTitles["part"] == part_str)]

        for _, drTitles in filtered_titles.iterrows():
            # Worktable titles
            ws.cell(row=3, column=2, value=to_str(drTitles["title1"]))
            ws.cell(row=3, column=2).font = Font(bold=True)
            ws.cell(row=4, column=2, value=to_str(drTitles["title2"]))
            ws.cell(row=4, column=2).font = Font(bold=True)
            ws.cell(row=5, column=2, value=to_str(drTitles["title3"]))
            ws.cell(row=5, column=2).font = Font(bold=True)

            sWorktableColumn = ""
            # Column headers
            for iCurrentCol in range(1, iColumnCount + 1):
                sColumnName = f"rpt_col{iCurrentCol}"
                col_val = str(drTitles.get(sColumnName, ""))
                if "Source" in col_val:
                    sWorktableColumn = col_val.split(" ")[0]
                else:
                    if sWorktableColumn:
                        ws.cell(row=6, column=iCurrentCol, value=sWorktableColumn)
                        ws.cell(row=6, column=iCurrentCol).alignment = Alignment(horizontal="center")
                ws.cell(row=7, column=iCurrentCol, value=col_val.replace("|", "\n"))

            # Column header formatting
            # Calculate Excel column range string for formatting
            def excel_col(n):
                result = ""
                while n > 0:
                    n, rem = divmod(n - 1, 26)
                    result = chr(65 + rem) + result
                return result

            start_col = "A"
            end_col = excel_col(iColumnCount)
            sColumnRange = f"{start_col}7:{end_col}7"
            for row in ws[sColumnRange]:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    cell.border = Border(bottom=Side(style='double'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
                    cell.font = Font(bold=True)
            # Set row height for header row
            ws.row_dimensions[7].height = 40
            # Set column width for header columns
            for iCurrentCol in range(1, iColumnCount + 1):
                col_letter = excel_col(iCurrentCol)
                ws.column_dimensions[col_letter].width = 25

            # Setup line, code, and identification columns
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 50
            for col in ['A', 'B', 'C']:
                for cell in ws[col]:
                    cell.alignment = Alignment(horizontal="left")
            # SET PAGE LAYOUT (openpyxl syntax)
            ws.page_setup.orientation = "landscape"
            ws.page_setup.zoom = 70
            ws.oddFooter.left.text = "Page &P of &N"
            ws.print_title_rows = '1:7'
            ws.print_title_cols = 'A:C'

            # Write the line source text
            iROW_COUNT = 8

            for _, drRow in self.dtLineSourceText[self.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                # Write The Line Number
                ws.cell(row=iROW_COUNT, column=1, value=getattr(drRow, "line", getattr(drRow, "Line", None)))
                # Write Code
                ws.cell(row=iROW_COUNT, column=2, value="'" + str(getattr(drRow, "code", getattr(drRow, "Code", ""))))
                # Write Identification
                ws.cell(row=iROW_COUNT, column=3, value=str(getattr(drRow, "identification", getattr(drRow, "Identification", ""))))
                # increment iROW_COUNT
                iROW_COUNT += 1
            iROW_COUNT += 2
            # Write footnotes
            filtered_footnotes = self.dtFootnotes[(self.dtFootnotes["worktable"] == sSheetTitle[:2]) & (self.dtFootnotes["part"] == part_str)]
            for drRow in filtered_footnotes.itertuples(index=False):
                # Write The Footnote number
                ws.cell(row=iROW_COUNT, column=2, value=str(getattr(drRow, "no", getattr(drRow, "No", ""))))
                # Write Code
                ws.cell(row=iROW_COUNT, column=3, value="'" + str(getattr(drRow, "text", getattr(drRow, "Text", ""))))
                # increment iROW_COUNT
                iROW_COUNT += 1

    def scrub_year(self, s: str, current_year: int) -> str:
        s = s.replace("#Y#", str(current_year))
        s = s.replace("#Y-1#", str(current_year - 1))
        s = s.replace("#Y-2#", str(current_year - 2))
        s = s.replace("#Y-3#", str(current_year - 3))
        s = s.replace("#Y-4#", str(current_year - 4))
        s = s.replace("#Y-5#", str(current_year - 4))
        return s
    
    
    def add_formula_as_text(self, ws, row, col, formula):
        """
        Writes a formula as a string (not executable) in the specified cell.
        This is useful for displaying the formula as text in Excel, not as a calculated value.
        """
        cell = ws.cell(row=row, column=col, value=f"'{formula}")
        cell.alignment = Alignment(horizontal="left")
    
    def A1P1_worksheet(self, wb, index=2):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
            iColumnCount = 21
            sSheetTitle = "A1P1"
            iLineNumberOffset = 93

            dtaValue = self.dtAValue
            dtaValue0_RR = self.dtAValue0_RR
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            # Update status
            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            # Get worksheet
            ws = wb.create_sheet(title=sSheetTitle)


            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)  
            # Freeze panes
            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue for all lines but 158
            col_map = {
                        iCurrentYear: 5,
                        iCurrentYear - 1: 7,
                        iCurrentYear - 2: 9,
                        iCurrentYear - 3: 11,
                        iCurrentYear - 4: 13
                    }
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] != 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                    if iProcessYear in col_map:
                        col = col_map[iProcessYear]
                        # Calculate the 'C' number for the named range, similar to the original VB
                        c_num = iCurrentYear - iProcessYear + 1
                        ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"

            # Write hard values from dtaValue0_RR for line 158
            for _, draValues in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
                if draValues["aline"] == 158:
                    iProcessYear = int(draValues["year"])
                    iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset

                    if iProcessYear in col_map:
                        c_num = iCurrentYear - iProcessYear + 1
                        col = col_map[iProcessYear]
                        value = draValues["value"] if iProcessYear == iCurrentYear else "0"
                        ws.cell(row=iROW_COUNT, column=col, value=value)
                        cell = ws.cell(row=iROW_COUNT, column=col)
                        sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                        cell.alignment = Alignment(horizontal="right")
                        wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        
                        cell.number_format = "#,##0"

            # Write sources and values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                
                # Sources first
                source_cols = {
                    4: "c1", 6: "c2", 8: "c3", 10: "c4", 12: "c5",
                    14: "c6", 16: "c7", 18: "c8", 20: "c9"
                }
                for col, c_name in source_cols.items():
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), self.current_year)
                    if source_text.startswith('=') or source_text.startswith('+'):
                        self.add_formula_as_text(ws, iLine, col, source_text)
                    else:
                        ws.cell(row=iLine, column=col, value=source_text)

                # C6 - C9 Values
                value_cols_c6_c9 = {15: "c6", 17: "c7", 19: "c8",  21: "c9"}
                for col, c_name in value_cols_c6_c9.items():
                    value = "0" if drSource["line"] == 158 else str(drSource[c_name])
                    c_num = (col - 15) // 2 + 6
                    ws.cell(row=iLine, column=col, value=value)
                    cell = ws.cell(row=iLine, column=col)
                    cell.number_format = "#,##0"
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    

                # Random values not from Data Dictionary
                random_lines = {111, 122, 131, 133, 137, 144, 145, 146, 151, 154, 155}
                if int(drSource["line"]) in random_lines:
                    random_value_cols = {5: "c1", 7: "c2", 9: "c3", 11: "c4", 13: "c5"}
                    for col, c_name in random_value_cols.items():
                        c_num = (col - 5) // 2 + 1
                        ws.cell(row=iLine, column=col, value=drSource.get(c_name, ""))
                        cell = ws.cell(row=iLine, column=col)
                        cell.number_format = "#,##0"
                        named_range_name = f"A1L{drSource['line']}C{c_num}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in A1P1: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P2A_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 2A"
            iColumnCount = 57
            sSheetTitle = "A1P2A"
            iLineNumberOffset = 193

            dtaValue = self.dtAValue
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except Exception:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+2]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                aCode_id = int(draValues["acode_id"]) if "acode_id" in draValues else int(draValues["aCode_id"]) if "aCode_id" in draValues else 0

                # Even/odd column mapping
                if aCode_id % 2 == 0:
                    col_map = {
                        iCurrentYear: 5,
                        iCurrentYear - 1: 11,
                        iCurrentYear - 2: 17,
                        iCurrentYear - 3: 23,
                        iCurrentYear - 4: 29
                    }
                    c_map = {
                        iCurrentYear: 1,
                        iCurrentYear - 1: 4,
                        iCurrentYear - 2: 7,
                        iCurrentYear - 3: 10,
                        iCurrentYear - 4: 13
                    }
                else:
                    col_map = {
                        iCurrentYear: 7,
                        iCurrentYear - 1: 13,
                        iCurrentYear - 2: 19,
                        iCurrentYear - 3: 25,
                        iCurrentYear - 4: 31
                    }
                    c_map = {
                        iCurrentYear: 2,
                        iCurrentYear - 1: 5,
                        iCurrentYear - 2: 8,
                        iCurrentYear - 3: 11,
                        iCurrentYear - 4: 14
                    }
                if iProcessYear in col_map:
                    col = col_map[iProcessYear]
                    c_num = c_map[iProcessYear]
                    cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                    sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources and derived values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                # Write all 27 source columns (C1 to C27) to columns 4,6,8,...,56
                for n in range(1, 28):
                    col = 2 * n + 2  # 4,6,8,...,56
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=source_text if not (source_text.startswith('=') or source_text.startswith('+')) else "'" + source_text)

                # Write derived values to columns 9, 15, 21, ..., 57 (every 6th column starting from 9)
                derived_cols = [
                    (9, 3), (15, 6), (21, 9), (27, 12), (33, 15), (35, 16), (37, 17), (39, 18),
                    (41, 19), (43, 20), (45, 21), (47, 22), (49, 23), (51, 24), (53, 25), (55, 26), (57, 27)
                ]
                for col, c_num in derived_cols:
                    c_name = f"c{c_num}"
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=col, value=value)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P2B_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 2B"
            iColumnCount = 57
            sSheetTitle = "A1P2B"
            iLineNumberOffset = 209

            dtaValue = self.dtAValue
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                aCode_id = int(draValues.get("acode_id") or draValues.get("aCode_id", 0))

                # Even/odd column mapping (same as A1P2A)
                if aCode_id % 2 == 0:
                    col_map = {
                        iCurrentYear: 5, iCurrentYear - 1: 11, iCurrentYear - 2: 17,
                        iCurrentYear - 3: 23, iCurrentYear - 4: 29
                    }
                    c_map = {
                        iCurrentYear: 1, iCurrentYear - 1: 4, iCurrentYear - 2: 7,
                        iCurrentYear - 3: 10, iCurrentYear - 4: 13
                    }
                else:
                    col_map = {
                        iCurrentYear: 7, iCurrentYear - 1: 13, iCurrentYear - 2: 19,
                        iCurrentYear - 3: 25, iCurrentYear - 4: 31
                    }
                    c_map = {
                        iCurrentYear: 2, iCurrentYear - 1: 5, iCurrentYear - 2: 8,
                        iCurrentYear - 3: 11, iCurrentYear - 4: 14
                    }

                if iProcessYear in col_map:
                    col = col_map[iProcessYear]
                    c_num = c_map[iProcessYear]
                    cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                    sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources and derived values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                # Write all 27 source columns (C1 to C27) to columns 4,6,8,...,56
                for n in range(1, 28):
                    col = 2 * n + 2
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                # Write derived values (same as A1P2A)
                derived_cols = [
                    (9, 3), (15, 6), (21, 9), (27, 12), (33, 15), (35, 16), (37, 17), (39, 18),
                    (41, 19), (43, 20), (45, 21), (47, 22), (49, 23), (51, 24), (53, 25), (55, 26), (57, 27)
                ]
                for col, c_num in derived_cols:
                    c_name = f"c{c_num}"
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=col, value=value)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

                # Random Values for line 234
                if int(drSource["line"]) == 234:
                    random_value_cols = {
                        5: "c1", 7: "c2", 11: "c4", 13: "c5", 17: "c7", 19: "c8",
                        23: "c10", 25: "c11", 29: "c13", 31: "c14"
                    }
                    c_num_map = {
                        5: 1, 7: 2, 11: 4, 13: 5, 17: 7, 19: 8,
                        23: 10, 25: 11, 29: 13, 31: 14
                    }
                    for col, c_name in random_value_cols.items():
                        c_num = c_num_map[col]
                        value = drSource.get(c_name, "")
                        cell = ws.cell(row=iLine, column=col, value=value)
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                        named_range_name = f"A1L{drSource['line']}C{c_num}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())
            
    
    def A1P2C_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 2C"
            iColumnCount = 57
            iWorkTableColumnCount = 27
            sSheetTitle = "A1P2C"
            iLineNumberOffset = 228
            sNamedRangePrefix = "A1L"

            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # This sheet has no values from the Data Dictionary (dtAValue)
            # It writes sources and values from dtLineSourceText
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                
                # Loop through C1 to C27 columns from dtLineSourceText
                for i in range(1, iWorkTableColumnCount + 1):
                    c_name = f"c{i}"
                    
                    # Source column (4, 6, 8, ...)
                    source_col = 2 * i + 2
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=source_col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                    # Value column (5, 7, 9, ...)
                    value_col = source_col + 1
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=value_col, value=value)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

                    # Create named range
                    named_range_name = f"{sNamedRangePrefix}{drSource['line']}{c_name.upper()}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
            
            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P3A_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 3A"
            iColumnCount = 41
            sSheetTitle = "A1P3A"
            iLineNumberOffset = 293

            dtaValue = self.dtAValue
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue (similar to A1P2A/B)
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                aCode_id = int(draValues.get("acode_id") or draValues.get("aCode_id", 0))

                if aCode_id % 2 == 0: # Even
                    col_map = { iCurrentYear: 5, iCurrentYear - 1: 11, iCurrentYear - 2: 17, iCurrentYear - 3: 23, iCurrentYear - 4: 29 }
                    c_map = { iCurrentYear: 1, iCurrentYear - 1: 4, iCurrentYear - 2: 7, iCurrentYear - 3: 10, iCurrentYear - 4: 13 }
                else: # Odd
                    col_map = { iCurrentYear: 7, iCurrentYear - 1: 13, iCurrentYear - 2: 19, iCurrentYear - 3: 25, iCurrentYear - 4: 31 }
                    c_map = { iCurrentYear: 2, iCurrentYear - 1: 5, iCurrentYear - 2: 8, iCurrentYear - 3: 11, iCurrentYear - 4: 14 }

                if iProcessYear in col_map:
                    col, c_num = col_map[iProcessYear], c_map[iProcessYear]
                    cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                    sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources and derived values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                # Write source columns (C1 to C19)
                for n in range(1, 20):
                    col = 2 * n + 2
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                # Write derived values
                derived_cols = [ (9, 3), (15, 6), (21, 9), (27, 12), (33, 15), (35, 16), (37, 17), (39, 18), (41, 19) ]
                for col, c_num in derived_cols:
                    c_name = f"c{c_num}"
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=col, value=value)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

                # Random Values for specific lines
                if int(drSource["line"]) in {304, 308, 312, 323, 324}:
                    random_value_cols = { 5: "c1", 7: "c2", 11: "c4", 13: "c5", 17: "c7", 19: "c8", 23: "c10", 25: "c11", 29: "c13", 31: "c14" }
                    c_num_map = { 5: 1, 7: 2, 11: 4, 13: 5, 17: 7, 19: 8, 23: 10, 25: 11, 29: 13, 31: 14 }
                    for col, c_name in random_value_cols.items():
                        value = drSource.get(c_name, "")
                        cell = ws.cell(row=iLine, column=col, value=value)
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                        named_range_name = f"A1L{drSource['line']}C{c_num_map[col]}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
            
            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P3B_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 3B"
            iColumnCount = 41
            iWorkTableColumnCount = 19
            sSheetTitle = "A1P3B"
            iLineNumberOffset = 333
            sNamedRangePrefix = "A1L"

            dtaValue = self.dtAValue
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                aCode_id = int(draValues.get("acode_id") or draValues.get("aCode_id", 0))

                if aCode_id % 2 == 0: # Even
                    col_map = { iCurrentYear: 5, iCurrentYear - 1: 11, iCurrentYear - 2: 17, iCurrentYear - 3: 23, iCurrentYear - 4: 29 }
                    c_map = { iCurrentYear: 1, iCurrentYear - 1: 4, iCurrentYear - 2: 7, iCurrentYear - 3: 10, iCurrentYear - 4: 13 }
                else: # Odd
                    col_map = { iCurrentYear: 7, iCurrentYear - 1: 13, iCurrentYear - 2: 19, iCurrentYear - 3: 25, iCurrentYear - 4: 31 }
                    c_map = { iCurrentYear: 2, iCurrentYear - 1: 5, iCurrentYear - 2: 8, iCurrentYear - 3: 11, iCurrentYear - 4: 14 }

                if iProcessYear in col_map:
                    col, c_num = col_map[iProcessYear], c_map[iProcessYear]
                    cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                    sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources and derived values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                # Write source columns (C1 to C19)
                for n in range(1, iWorkTableColumnCount + 1):
                    col = 2 * n + 2
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                # Write derived values
                derived_cols = [ (9, 3), (15, 6), (21, 9), (27, 12), (33, 15), (35, 16), (37, 17), (39, 18), (41, 19) ]
                for col, c_num in derived_cols:
                    c_name = f"c{c_num}"
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=col, value=value)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

                # Random Values for specific lines
                if int(drSource["line"]) in {344, 348, 352, 363, 364}:
                    random_value_cols = { 5: "c1", 7: "c2", 11: "c4", 13: "c5", 17: "c7", 19: "c8", 23: "c10", 25: "c11", 29: "c13", 31: "c14" }
                    c_num_map = { 5: 1, 7: 2, 11: 4, 13: 5, 17: 7, 19: 8, 23: 10, 25: 11, 29: 13, 31: 14 }
                    for col, c_name in random_value_cols.items():
                        value = drSource.get(c_name, "")
                        cell = ws.cell(row=iLine, column=col, value=value)
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                        named_range_name = f"A1L{drSource['line']}C{c_num_map[col]}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
            
            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P4_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 4"
            iColumnCount = 9
            iWorkTableColumnCount = 3
            sSheetTitle = "A1P4"
            iLineNumberOffset = 393
            sNamedRangePrefix = "A1L"

            dtaValue0_RR = self.dtAValue0_RR
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue0_RR
            for _, draValue0_RR in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValue0_RR["year"])
                if iProcessYear == iCurrentYear:
                    iROW_COUNT = int(draValue0_RR["aline"]) - iLineNumberOffset
                    aCode_id = int(draValue0_RR.get("acode_id", 0))

                    if aCode_id % 2 == 0: # Even
                        col, c_num = 5, 1
                    else: # Odd
                        col, c_num = 7, 2

                    cell = ws.cell(row=iROW_COUNT, column=col, value=draValue0_RR["value"])
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"
                    sNamedRange = f"A1L{draValue0_RR['aline']}C{c_num}"
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources and derived values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                # Write source columns (C1 to C3)
                for n in range(1, iWorkTableColumnCount + 1):
                    col = 2 * n + 2
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                # Write derived value for C3
                cell = ws.cell(row=iLine, column=9, value=drSource.get("c3", ""))
                named_range_name = f"A1L{drSource['line']}C3"
                wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

                # Random Values for specific lines
                if int(drSource["line"]) in {406, 426, 431, 439, 445, 449, 453, 457, 460, 465, 470, 475, 480}:
                    # C1 value
                    cell1 = ws.cell(row=iLine, column=5, value=drSource.get("c1", ""))
                    cell1.number_format = "#,##0"
                    wb.defined_names[f"A1L{drSource['line']}C1"] = DefinedName(name=f"A1L{drSource['line']}C1", attr_text=f"'{sSheetTitle}'!${cell1.column_letter}${cell1.row}")
                    # C2 value
                    cell2 = ws.cell(row=iLine, column=7, value=drSource.get("c2", ""))
                    cell2.number_format = "#,##0"
                    wb.defined_names[f"A1L{drSource['line']}C2"] = DefinedName(name=f"A1L{drSource['line']}C2", attr_text=f"'{sSheetTitle}'!${cell2.column_letter}${cell2.row}")
            
            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P5A_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 5A"
            iColumnCount = 25
            iWorkTableColumnCount = 11
            sSheetTitle = "A1P5A"
            iLineNumberOffset = 493
            sNamedRangePrefix = "A1L"

            dtCarTypeStatistics = self.dtCarTypeStatistics
            dtaValueRegion_RR = self.dtAValueRegion_RR
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtCarTypeStatistics
            for _, drCarStats in dtCarTypeStatistics.iterrows():
                iROW_COUNT = int(drCarStats["line"]) - iLineNumberOffset
                col_map = { 5: "c1", 7: "c2", 9: "c3", 13: "c5", 15: "c6", 17: "c7", 19: "c8", 21: "c9", 23: "c10", 25: "c11" }
                c_num_map = { 5: 1, 7: 2, 9: 3, 13: 5, 15: 6, 17: 7, 19: 8, 21: 9, 23: 10, 25: 11 }
                for col, c_name in col_map.items():
                    cell = ws.cell(row=iROW_COUNT, column=col, value=drCarStats.get(c_name, ""))
                    cell.number_format = "#,##0.#####"
                    if c_name == "c6":
                        cell.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
                    named_range = f"A1L{drCarStats['line']}C{c_num_map[col]}"
                    wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write hard values from dtaValueRegion_RR
            filtered_region_values = dtaValueRegion_RR[
                (dtaValueRegion_RR["rpt_sheet"] == sSheetTitle) &
                (dtaValueRegion_RR["year"] == iCurrentYear) &
                (dtaValueRegion_RR["code"] == "C4")
            ]
            for _, draValueRegion in filtered_region_values.iterrows():
                iROW_COUNT = int(draValueRegion["aline"]) - iLineNumberOffset
                cell = ws.cell(row=iROW_COUNT, column=11, value=draValueRegion["value"])
                named_range = f"A1L{draValueRegion['aline']}C4"
                wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                for n in range(1, iWorkTableColumnCount + 1):
                    col = 2 * n + 2
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)
            
            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P5B_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 5B"
            iColumnCount = 31
            iWorkTableColumnCount = 14
            sSheetTitle = "A1P5B"
            iLineNumberOffset = 513
            sNamedRangePrefix = "A1L"

            dtCarTypeStatisticsPart2 = self.dtCarTypeStatisticsPart2
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from R_OP_STATS_BY_CAR_TYPE_2
            for _, drCarStats in dtCarTypeStatisticsPart2.iterrows():
                iROW_COUNT = int(drCarStats["line"]) - iLineNumberOffset
                for i in range(1, iWorkTableColumnCount + 1):
                    col = 2 * i + 3 # 5, 7, 9, ...
                    c_name = f"c{i}"
                    cell = ws.cell(row=iROW_COUNT, column=col, value=drCarStats.get(c_name, ""))
                    named_range = f"{sNamedRangePrefix}{drCarStats['line']}C{i}"
                    wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                for n in range(1, iWorkTableColumnCount + 1):
                    col = 2 * n + 2 # 4, 6, 8, ...
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P6_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 6"
            iColumnCount = 165
            iWorkTableColumnCount = 81
            sSheetTitle = "A1P6"
            iLineNumberOffset = 533
            sNamedRangePrefix = "A1L"

            dtaValue = self.dtAValue
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)
            iCodeOffset = 738

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from dtaValue
            for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                iLineAdjustment = int(draValues["aline"]) - iLineNumberOffset - 8
                acode_id = int(draValues.get("acode_id", 0))

                year_diff = iCurrentYear - iProcessYear
                if not (0 <= year_diff <= 4):
                    continue

                col_offset = year_diff * 18
                c_offset = year_diff * 9

                acode_check = acode_id - (iCodeOffset + iLineAdjustment * 4)

                if acode_check == 0: # C1, C10, C19, C28, C37
                    col, c_num = 5 + col_offset, 1 + c_offset
                elif acode_check == 1: # C2, C11, C20, C29, C38
                    col, c_num = 7 + col_offset, 2 + c_offset
                elif acode_check == 2: # C3, C12, C21, C30, C39
                    col, c_num = 9 + col_offset, 3 + c_offset
                elif acode_check == 3: # C4, C13, C22, C31, C40
                    col, c_num = 11 + col_offset, 4 + c_offset
                else:
                    continue

                cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                cell.alignment = Alignment(horizontal="right")
                sNamedRange = f"{sNamedRangePrefix}{draValues['aline']}C{c_num}"
                wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources and derived values
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                # Write source columns (C1 to C81)
                for n in range(1, iWorkTableColumnCount + 1):
                    col = 2 * n + 2
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                # Write derived values
                derived_cols = [
                    (13, 5), (15, 6), (17, 7), (19, 8), (21, 9), (31, 14), (33, 15), (35, 16), (37, 17), (39, 18),
                    (49, 23), (51, 24), (53, 25), (55, 26), (57, 27), (67, 32), (69, 33), (71, 34), (73, 35), (75, 36),
                    (85, 41), (87, 42), (89, 43), (91, 44), (93, 45), (95, 46), (97, 47), (99, 48), (101, 49), (103, 50),
                    (105, 51), (107, 52), (109, 53), (111, 54), (113, 55), (115, 56), (117, 57), (119, 58), (121, 59),
                    (123, 60), (125, 61), (127, 62), (129, 63), (131, 64), (133, 65), (135, 66), (137, 67), (139, 68),
                    (141, 69), (143, 70), (145, 71), (147, 72), (149, 73), (151, 74), (153, 75), (155, 76), (157, 77),
                    (159, 78), (161, 79), (163, 80), (165, 81)
                ]
                for col, c_num in derived_cols:
                    c_name = f"c{c_num}"
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=col, value=value)
                    if c_num in {8, 9, 17, 18, 26, 27, 35, 36} or c_num >= 44:
                        cell.number_format = "#,##0.#####"
                    named_range_name = f"{sNamedRangePrefix}{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P7_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 7"
            iColumnCount = 13
            iWorkTableColumnCount = 5
            sSheetTitle = "A1P7"
            iLineNumberOffset = 553
            sNamedRangePrefix = "A1L"

            dtCarTypeStatisticsPart3 = self.dtCarTypeStatisticsPart3
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write hard values from R_OP_STATS_BY_CAR_TYPE_3
            for _, drCarStats in dtCarTypeStatisticsPart3.iterrows():
                iROW_COUNT = int(drCarStats["line"]) - iLineNumberOffset
                for i in range(1, iWorkTableColumnCount + 1):
                    col = 2 * i + 3 # 5, 7, 9, ...
                    c_name = f"c{i}"
                    cell = ws.cell(row=iROW_COUNT, column=col, value=drCarStats.get(c_name, ""))
                    named_range = f"{sNamedRangePrefix}{drCarStats['line']}C{i}"
                    wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                for n in range(1, iWorkTableColumnCount + 1):
                    col = 2 * n + 2 # 4, 6, 8, ...
                    c_name = f"c{n}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P8_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 8"
            iColumnCount = 5
            iWorkTableColumnCount = 1
            sSheetTitle = "A1P8"
            iLineNumberOffset = 572
            sNamedRangePrefix = "A1L"

            dtaValueRegion_RR = self.dtAValueRegion_RR
            dtaValue0_RR = self.dtAValue0_RR
            dtaValue = self.dtAValue
            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            # Write values from dtaValueRegion_RR
            filtered_region_values = dtaValueRegion_RR[
                (dtaValueRegion_RR["rpt_sheet"] == sSheetTitle) &
                (dtaValueRegion_RR["year"] == iCurrentYear) &
                (dtaValueRegion_RR["code"] == "C1")
            ]
            for _, draValueRegion in filtered_region_values.iterrows():
                iROW_COUNT = int(draValueRegion["aline"]) - iLineNumberOffset
                cell = ws.cell(row=iROW_COUNT, column=5, value=draValueRegion["value"])
                named_range = f"{sNamedRangePrefix}{draValueRegion['aline']}C1"
                wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write values from dtaValue0_RR
            lines_to_check = {584, 585, 586, 587, 595}
            filtered_zero_rr = dtaValue0_RR[
                (dtaValue0_RR["rpt_sheet"] == sSheetTitle) &
                (dtaValue0_RR["year"] == iCurrentYear) &
                (dtaValue0_RR["aline"].isin(lines_to_check))
            ]
            for _, draValue0 in filtered_zero_rr.iterrows():
                iROW_COUNT = int(draValue0["aline"]) - iLineNumberOffset
                cell = ws.cell(row=iROW_COUNT, column=5, value=draValue0["value"])
                named_range = f"{sNamedRangePrefix}{draValue0['aline']}C1"
                wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write values from dtaValue (Railroad specific)
            filtered_rr_values = dtaValue[
                (dtaValue["rpt_sheet"] == sSheetTitle) &
                (dtaValue["year"] == iCurrentYear) &
                (dtaValue["aline"] == 581)
            ]
            for _, draValue in filtered_rr_values.iterrows():
                iROW_COUNT = int(draValue["aline"]) - iLineNumberOffset
                cell = ws.cell(row=iROW_COUNT, column=5, value=draValue["value"])
                named_range = f"{sNamedRangePrefix}{draValue['aline']}C1"
                wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write sources
            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                source_text = self.scrub_year(str(drSource.get("c1", "")), iCurrentYear)
                ws.cell(row=iLine, column=4, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A1P9_worksheet(self, wb):
        try:
            sTitle_WORKTABLE = "WORKTABLE A1 PART 9"
            iColumnCount = 13
            iWorkTableColumnCount = 5
            sSheetTitle = "A1P9"
            iLineNumberOffset = 893
            sNamedRangePrefix = "A1L"

            dtLineSourceText = self.dtLineSourceText
            iCurrentYear = int(self.current_year)

            print(f"Processing {sSheetTitle}")

            # Select worktable and string rows
            try:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
            except ValueError:
                part_str = sSheetTitle[sSheetTitle.index('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

            ws = wb.create_sheet(title=sSheetTitle)

            # Write titles and column headers
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            ws.freeze_panes = ws['D8']

            for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset

                for i in range(1, iWorkTableColumnCount + 1):
                    c_name = f"c{i}"
                    
                    # Source column (4, 6, 8, ...)
                    source_col = 2 * i + 2
                    source_text = self.scrub_year(to_str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=source_col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                    # Derived Value column (5, 7, 9, ...)
                    value_col = 2 * i + 3
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=value_col, value=value)
                    
                    # Conditional formatting
                    if 12 < iLine < 25:
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = "#,##0"

                    # Create named range
                    named_range_name = f"{sNamedRangePrefix}{drSource['line']}C{i}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Write range names for empty cells
            for i in range(8, 26): # Rows 8 to 25
                for j in range(5, 14): # Columns 5 to 13
                    caption_cell = ws.cell(row=6, column=j)
                    cell = ws.cell(row=i, column=j)

                    if caption_cell.value is not None and cell.value is None:
                        cell.value = "=NULL_VALUE"
                        caption_text = str(caption_cell.value).replace("(", "").replace(")", "")
                        named_range_name = f"{sNamedRangePrefix}{i + iLineNumberOffset}{caption_text}"
                        wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = "#######0"

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def A2P1_worksheet(self, wb):
        try:
            # BUILD THE WORKTABLE TITLE AND SET WORKTABLE VARS
            sTitle_WORKTABLE = "WORKTABLE A2 PART 1"
            iColumnCount = 91
            iWorkTableColumnCount = 44
            sSheetTitle = "A2P1"
            iLineNumberOffset = 93
            sNamedRangePrefix = "A2L"

            print(f"Processing {sSheetTitle}")

            # GET OUR SHEET AND A RANGE
            ws = wb.create_sheet(title=sSheetTitle)

            # WRITE OUT TITLE AND COLUMN HEADERS
            part_str = sSheetTitle[sSheetTitle.find('P')+1:]
            sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
            self.write_titles_and_column_headers(ws, self.dtTitles, sSelectWorktable, self.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)

            # WRITE OUT THE FIRST 3 COLUMNS
            sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"
            self.WriteFirst3ColumnsAndPageLayout(ws, self.dtLineSourceText, self.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

            # FREEZE THE ROWS AND COLUMNS
            ws.freeze_panes = ws['D8']

            # Helper to set cell value and named range
            def set_cell(row, col, value, name, num_format=None):
                cell = ws.cell(row=row, column=col, value=value)
                wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                cell.alignment = Alignment(horizontal="right")
                if num_format:
                    cell.number_format = num_format
                return cell

            # WRITE HARD VALUES FROM AVALUES
            iCurrentYear = int(self.current_year)
            for _, draValues in self.dtAValue[self.dtAValue["rpt_sheet"] == sSheetTitle].iterrows():
                iProcessYear = int(draValues["year"])
                aLine = int(draValues["aline"])
                acode_id = int(draValues["acode_id"])
                value_str = str(draValues["value"])

                drAnnPeriod = self.dtDataDictionary[self.dtDataDictionary["line"] == f"A2L{aLine}"]
                iROW_COUNT = aLine - iLineNumberOffset

                iCodeOffset = 898 if aLine < 142 else 1267
                iRowOffset = 8 if aLine < 142 else 82

                if iProcessYear == iCurrentYear and not drAnnPeriod.empty:
                    set_cell(iROW_COUNT, 5, drAnnPeriod.iloc[0]["annperiod"], f"A2L{aLine}C1")

                is_main_block = (aLine < 142) or (174 < aLine < 181)
                
                # Price Index helper
                def get_pi(index, year_col_name):
                    pi_row = self.dtPriceIndexes[self.dtPriceIndexes['index'] == index]
                    return "0" if value_str == "0" else pi_row.iloc[0][year_col_name] if not pi_row.empty else "0"

                # --- Start of faithful translation of the large If/ElseIf block ---

                # Condition 1
                if is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset):
                    if iProcessYear == iCurrentYear:
                        set_cell(iROW_COUNT, 9, get_pi(1, "current_year"), f"A2L{aLine}C3", "0.0000")
                        set_cell(iROW_COUNT, 7, value_str, f"A2L{aLine}C2", "#,##0")
                    elif iProcessYear == iCurrentYear - 1:
                        set_cell(iROW_COUNT, 29, get_pi(1, "current_year_minus_1"), f"A2L{aLine}C13", "0.0000")
                        set_cell(iROW_COUNT, 27, value_str, f"A2L{aLine}C12", "#,##0")
                    elif iProcessYear == iCurrentYear - 2:
                        set_cell(iROW_COUNT, 45, get_pi(1, "current_year_minus_2"), f"A2L{aLine}C21", "0.0000")
                        set_cell(iROW_COUNT, 43, value_str, f"A2L{aLine}C20", "#,##0")
                    elif iProcessYear == iCurrentYear - 3:
                        set_cell(iROW_COUNT, 61, get_pi(1, "current_year_minus_3"), f"A2L{aLine}C29", "0.0000")
                        set_cell(iROW_COUNT, 59, value_str, f"A2L{aLine}C28", "#,##0")
                    elif iProcessYear == iCurrentYear - 4:
                        set_cell(iROW_COUNT, 77, get_pi(1, "current_year_minus_4"), f"A2L{aLine}C37", "0.0000")
                        set_cell(iROW_COUNT, 75, value_str, f"A2L{aLine}C36", "#,##0")

                # Condition 2
                elif is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 1:
                    if iProcessYear == iCurrentYear:
                        set_cell(iROW_COUNT, 13, get_pi(2, "current_year"), f"A2L{aLine}C5", "0.0000")
                        set_cell(iROW_COUNT, 11, value_str, f"A2L{aLine}C4", "#,##0")
                    elif iProcessYear == iCurrentYear - 1:
                        set_cell(iROW_COUNT, 33, get_pi(2, "current_year_minus_1"), f"A2L{aLine}C15", "0.0000")
                        set_cell(iROW_COUNT, 31, value_str, f"A2L{aLine}C14", "#,##0")
                    elif iProcessYear == iCurrentYear - 2:
                        set_cell(iROW_COUNT, 49, get_pi(2, "current_year_minus_2"), f"A2L{aLine}C23", "0.0000")
                        set_cell(iROW_COUNT, 47, value_str, f"A2L{aLine}C22", "#,##0")
                    elif iProcessYear == iCurrentYear - 3:
                        set_cell(iROW_COUNT, 65, get_pi(2, "current_year_minus_3"), f"A2L{aLine}C31", "0.0000")
                        set_cell(iROW_COUNT, 63, value_str, f"A2L{aLine}C30", "#,##0")
                    elif iProcessYear == iCurrentYear - 4:
                        set_cell(iROW_COUNT, 81, get_pi(2, "current_year_minus_4"), f"A2L{aLine}C39", "0.0000")
                        set_cell(iROW_COUNT, 79, value_str, f"A2L{aLine}C38", "#,##0")

                # Condition 3
                elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 2) or \
                     acode_id in {1162, 1166, 1170, 1174, 1178, 1182, 1186, 1190, 1194, 1198, 1202, 1206, 1210, 1214, 1218, 1222, 1226, 1230, 1243, 1247, 1251, 1255, 1259, 1263}:
                    if iProcessYear == iCurrentYear:
                        set_cell(iROW_COUNT, 17, get_pi(3, "current_year"), f"A2L{aLine}C7", "0.0000")
                        set_cell(iROW_COUNT, 15, value_str, f"A2L{aLine}C6", "#,##0")
                    elif iProcessYear == iCurrentYear - 1:
                        set_cell(iROW_COUNT, 37, get_pi(3, "current_year_minus_1"), f"A2L{aLine}C17", "0.0000")
                        set_cell(iROW_COUNT, 35, value_str, f"A2L{aLine}C16", "#,##0")
                    elif iProcessYear == iCurrentYear - 2:
                        set_cell(iROW_COUNT, 53, get_pi(3, "current_year_minus_2"), f"A2L{aLine}C25", "0.0000")
                        set_cell(iROW_COUNT, 51, value_str, f"A2L{aLine}C24", "#,##0")
                    elif iProcessYear == iCurrentYear - 3:
                        set_cell(iROW_COUNT, 69, get_pi(3, "current_year_minus_3"), f"A2L{aLine}C33", "0.0000")
                        set_cell(iROW_COUNT, 67, value_str, f"A2L{aLine}C32", "#,##0")
                    elif iProcessYear == iCurrentYear - 4:
                        set_cell(iROW_COUNT, 85, get_pi(3, "current_year_minus_4"), f"A2L{aLine}C41", "0.0000")
                        set_cell(iROW_COUNT, 83, value_str, f"A2L{aLine}C40", "#,##0")

                # Condition 4
                elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 3) or \
                     acode_id in {1144, 1147, 1150, 1153, 1156, 1159, 1234, 1237, 1240}:
                    if iProcessYear == iCurrentYear:
                        set_cell(iROW_COUNT, 21, get_pi(4, "current_year"), f"A2L{aLine}C9", "0.0000")
                        set_cell(iROW_COUNT, 19, value_str, f"A2L{aLine}C8", "#,##0")
                    elif iProcessYear == iCurrentYear - 1:
                        set_cell(iROW_COUNT, 41, get_pi(4, "current_year_minus_1"), f"A2L{aLine}C19", "0.0000")
                        set_cell(iROW_COUNT, 39, value_str, f"A2L{aLine}C18", "#,##0")
                    elif iProcessYear == iCurrentYear - 2:
                        set_cell(iROW_COUNT, 57, get_pi(4, "current_year_minus_2"), f"A2L{aLine}C27", "0.0000")
                        set_cell(iROW_COUNT, 55, value_str, f"A2L{aLine}C26", "#,##0")
                    elif iProcessYear == iCurrentYear - 3:
                        set_cell(iROW_COUNT, 73, get_pi(4, "current_year_minus_3"), f"A2L{aLine}C35", "0.0000")
                        set_cell(iROW_COUNT, 71, value_str, f"A2L{aLine}C34", "#,##0")
                    elif iProcessYear == iCurrentYear - 4:
                        set_cell(iROW_COUNT, 89, get_pi(4, "current_year_minus_4"), f"A2L{aLine}C43", "0.0000")
                        set_cell(iROW_COUNT, 87, value_str, f"A2L{aLine}C42", "#,##0")

                # Condition 5
                elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 4) or \
                     acode_id in {1145, 1148, 1151, 1154, 1157, 1160, 1164, 1168, 1172, 1176, 1180, 1184, 1188, 1192, 1196, 1200, 1204, 1208, 1212, 1216, 1220, 1224, 1228, 1232, 1235, 1238, 1241, 1245, 1249, 1253, 1257, 1261, 1265}:
                    if iProcessYear == iCurrentYear:
                        set_cell(iROW_COUNT, 23, value_str, f"A2L{aLine}C10", "#,##0")

                # Condition 6
                elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 5) or \
                     acode_id in {1146, 1149, 1152, 1155, 1158, 1161, 1165, 1169, 1173, 1177, 1181, 1185, 1189, 1193, 1197, 1201, 1205, 1209, 1213, 1217, 1221, 1225, 1229, 1233, 1236, 1239, 1242, 1246, 1250, 1254, 1258, 1262, 1266}:
                    if iProcessYear == iCurrentYear:
                        set_cell(iROW_COUNT, 25, value_str, f"A2L{aLine}C11", "#,##0")

            # WRITE RANGE NAMES FOR EMPTY CELLS AND HANDLE SSAC SETTING
            SSAC_Cells = {"A2L166C8", "A2L166C10", "A2L166C11", "A2L166C18", "A2L166C26", "A2L166C34", "A2L166C42", "A2L167C8", "A2L167C10", "A2L167C11", "A2L167C18", "A2L167C26", "A2L167C34", "A2L167C42", "A2L168C8", "A2L168C10", "A2L168C11", "A2L168C18", "A2L168C26", "A2L168C34", "A2L168C42", "A2L175C2", "A2L175C4", "A2L175C6", "A2L175C8", "A2L175C10", "A2L175C11", "A2L175C12", "A2L175C14", "A2L175C16", "A2L175C18", "A2L175C20", "A2L175C22", "A2L175C24", "A2L175C26", "A2L175C28", "A2L175C30", "A2L175C32", "A2L175C34", "A2L175C36", "A2L175C38", "A2L175C40", "A2L175C42", "A2L176C2", "A2L176C4", "A2L176C6", "A2L176C8", "A2L176C10", "A2L176C11", "A2L176C12", "A2L176C14", "A2L176C16", "A2L176C18", "A2L176C20", "A2L176C22", "A2L176C24", "A2L176C26", "A2L176C28", "A2L176C30", "A2L176C32", "A2L176C34", "A2L176C36", "A2L176C38", "A2L176C40", "A2L176C42", "A2L177C2", "A2L177C4", "A2L177C6", "A2L177C8", "A2L177C10", "A2L177C11", "A2L177C12", "A2L177C14", "A2L177C16", "A2L177C18", "A2L177C20", "A2L177C22", "A2L177C24", "A2L177C26", "A2L177C28", "A2L177C30", "A2L177C32", "A2L177C34", "A2L177C36", "A2L177C38", "A2L177C40", "A2L177C42"}
            for i in range(8, 92): # Rows 8 to 91
                for j in range(5, 92): # Columns 5 to 91
                    if i == 88 and j != 5:
                        continue
                    
                    caption_cell = ws.cell(row=6, column=j)
                    cell = ws.cell(row=i, column=j)

                    if caption_cell.value:
                        named_range_name = f"A2L{i + iLineNumberOffset}{str(caption_cell.value).replace('(', '').replace(')', '')}"
                        if cell.value is None:
                            set_cell(i, j, "=NULL_VALUE", named_range_name, "#######0")
                        else:
                            # Check if the cell's named range is in SSAC_Cells
                            # This requires finding the name for the cell's location
                            current_named_range = None
                            for name, dest in wb.defined_names.items():
                                if dest.attr_text == f"'{sSheetTitle}'!${cell.column_letter}${cell.row}":
                                    current_named_range = name
                                    break
                            if current_named_range in SSAC_Cells:
                                cell.value = f'=IF(SSAC="Y",0,{cell.value})'

            # WRITE OUT THE SOURCES AND ANY VALUES THAT EXECUTE THE SOURCE
            for _, drSource in self.dtLineSourceText[self.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                iLine = int(drSource["line"]) - iLineNumberOffset
                aline_str = str(drSource["line"])

                # Sources first
                source_cols = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78, 80, 82, 84, 86, 88]
                for idx, col_num in enumerate(source_cols):
                    c_name = f"c{idx + 1}"
                    source_text = self.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                    ws.cell(row=iLine, column=col_num, value=f"'{source_text}")

                if iLine == 88: # Line 181
                    set_cell(iLine, 7, drSource["c2"], f"A2L{aline_str}C2", "#,##0")
                    set_cell(iLine, 9, "=NULL_VALUE", f"A2L{aline_str}C3")
                    set_cell(iLine, 11, drSource["c4"], f"A2L{aline_str}C4", "#,##0")
                    set_cell(iLine, 13, "=NULL_VALUE", f"A2L{aline_str}C5")
                    set_cell(iLine, 15, drSource["c6"], f"A2L{aline_str}C6", "#,##0")
                    set_cell(iLine, 17, "=NULL_VALUE", f"A2L{aline_str}C7")
                    set_cell(iLine, 19, drSource["c8"], f"A2L{aline_str}C8", "#,##0")
                    set_cell(iLine, 21, "=NULL_VALUE", f"A2L{aline_str}C9")
                    set_cell(iLine, 23, drSource["c10"], f"A2L{aline_str}C10", "#,##0")
                    set_cell(iLine, 25, drSource["c11"], f"A2L{aline_str}C11", "#,##0")
                    set_cell(iLine, 27, drSource["c12"], f"A2L{aline_str}C12", "#,##0")
                    set_cell(iLine, 29, "=NULL_VALUE", f"A2L{aline_str}C13")
                    set_cell(iLine, 31, drSource["c14"], f"A2L{aline_str}C14", "#,##0")
                    set_cell(iLine, 33, "=NULL_VALUE", f"A2L{aline_str}C15")
                    set_cell(iLine, 35, drSource["c16"], f"A2L{aline_str}C16", "#,##0")
                    set_cell(iLine, 37, "=NULL_VALUE", f"A2L{aline_str}C17")
                    set_cell(iLine, 39, drSource["c18"], f"A2L{aline_str}C18", "#,##0")
                    set_cell(iLine, 41, "=NULL_VALUE", f"A2L{aline_str}C19")
                    set_cell(iLine, 43, drSource["c20"], f"A2L{aline_str}C20", "#,##0")
                    set_cell(iLine, 45, "=NULL_VALUE", f"A2L{aline_str}C21")
                    set_cell(iLine, 47, drSource["c22"], f"A2L{aline_str}C22", "#,##0")
                    set_cell(iLine, 49, "=NULL_VALUE", f"A2L{aline_str}C23")
                    set_cell(iLine, 51, drSource["c24"], f"A2L{aline_str}C24", "#,##0")
                    set_cell(iLine, 53, "=NULL_VALUE", f"A2L{aline_str}C25")
                    set_cell(iLine, 55, drSource["c26"], f"A2L{aline_str}C26", "#,##0")
                    set_cell(iLine, 57, "=NULL_VALUE", f"A2L{aline_str}C27")
                    set_cell(iLine, 59, drSource["c28"], f"A2L{aline_str}C28", "#,##0")
                    set_cell(iLine, 61, "=NULL_VALUE", f"A2L{aline_str}C29")
                    set_cell(iLine, 63, drSource["c30"], f"A2L{aline_str}C30", "#,##0")
                    set_cell(iLine, 65, "=NULL_VALUE", f"A2L{aline_str}C31")
                    set_cell(iLine, 67, drSource["c32"], f"A2L{aline_str}C32", "#,##0")
                    set_cell(iLine, 69, "=NULL_VALUE", f"A2L{aline_str}C33")
                    set_cell(iLine, 71, drSource["c34"], f"A2L{aline_str}C34", "#,##0")
                    set_cell(iLine, 73, "=NULL_VALUE", f"A2L{aline_str}C35")
                    set_cell(iLine, 75, drSource["c36"], f"A2L{aline_str}C36", "#,##0")
                    set_cell(iLine, 77, "=NULL_VALUE", f"A2L{aline_str}C37")
                    set_cell(iLine, 79, drSource["c38"], f"A2L{aline_str}C38", "#,##0")
                    set_cell(iLine, 81, "=NULL_VALUE", f"A2L{aline_str}C39")
                    set_cell(iLine, 83, drSource["c40"], f"A2L{aline_str}C40", "#,##0")
                    set_cell(iLine, 85, "=NULL_VALUE", f"A2L{aline_str}C41")
                    set_cell(iLine, 87, drSource["c42"], f"A2L{aline_str}C42", "#,##0")
                    set_cell(iLine, 89, "=NULL_VALUE", f"A2L{aline_str}C43")
                    ws.cell(row=iLine, column=90, value=f"'{self.scrub_year(drSource.get('c44', ''), iCurrentYear)}")
                    set_cell(iLine, 91, drSource["c44"], f"A2L{aline_str}C44", "#,##0")
                else:
                    sSource = self.get_source_for_a2_summary_column(drSource["line"], iLine)
                    ws.cell(row=iLine, column=90, value=f"'{sSource}")
                    set_cell(iLine, 91, sSource, f"A2L{aline_str}C44", "#,##0")

            self.format_all_cells(ws)
            print(f"{sSheetTitle} completed")

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def get_source_for_a2_summary_column(self, line_no, row_index):
        # This is a placeholder for the complex logic in the VB.NET GetSourceForA2SummaryColumn
        # It appears to build a SUM formula based on other cells in the same row.
        # A more robust implementation would parse the dependencies from dtLineSourceText.
        try:
            if 8 <= row_index <= 87: # Corresponds to lines 101-180
                return f"=SUM(A2L{line_no}C2,A2L{line_no}C4,A2L{line_no}C6,A2L{line_no}C8)"
            return ""

        except Exception as ex:
            print(f"Error in {sSheetTitle}: {ex}")
            import traceback
            print(traceback.format_exc())

    def recalculate_excel_formulas(self, filepath):
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(filepath)
        wb.RefreshAll()
        wb.Save()
        wb.Close()
        excel.Quit()

if __name__ == "__main__":
    
    create_reports = CreateReports(2023)
    create_reports.create_reports()
