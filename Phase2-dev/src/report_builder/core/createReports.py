# create_reports_new.py (or wherever your runner lives)
import os
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
from dotenv import load_dotenv
import win32com.client


load_dotenv()

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
# Add the src directory to path to access services
src_dir = os.path.abspath(os.path.join(current_dir, '..', '..'))
if src_dir not in sys.path:
    sys.path.insert(0, src_dir)

from registry import register, get, all_names, load_all_plugins  # import the functions above
from context import ReportContext  # import the class above
from currentDF import CurrentDF  # import the CurrentDF class
from services.DBManager import DBManager
from data.db_data import db_data


from report_builder.core.discovery import discover_sheets

discover_sheets()

# load_all_plugins(os.path.join(parent_dir, "sheets"))

class CreateReports:
    def __init__(self, current_year: int, o_db: DBManager=None, fontname="Consolas", fontsize=10):
        if o_db is None:
            self.o_db = DBManager()
        self.current_year = current_year
        self.fontname = fontname
        self.fontsize = fontsize
        self.s_output_folder = os.path.join(parent_dir, "..", "..", "reports")
        self.ctx = None  # will be set by build_context()
        self.current_df = None  # will be set by update_current_df()
        self.o_db.db_data.s_current_year = str(current_year)

    def build_context(self):
        print("Initializing report context...")
        self.ctx = ReportContext.from_db(
            self.o_db,
            current_year=self.current_year,
            sTitle_RR_YEAR=f"RR {self.current_year}",
            s_output_folder=self.s_output_folder,
            fontname=self.fontname,
            fontsize=self.fontsize,
            verbose=True,
        )
        print("Context ready.")
        
    def update_variable_ctx(self, rr_no: str):
        if self.ctx is None:
            raise ValueError("Context not initialized. Call build_context() first.")
        print(f"Loading data for RR_NO={rr_no}...")
        self.ctx.variable_ctx = CurrentDF.from_db(self.o_db, rr_no, self.current_year, verbose=True)
        print("Data loaded into context.")
        
    def remove_default_sheet(self, wb):
        # Remove the default sheet if it exists
        default_sheetnames = ["Sheet", "Sheet1"]
        for sheetname in default_sheetnames:
            if sheetname in wb.sheetnames:
                std = wb[sheetname]
                wb.remove(std)
    
    def create_workbook(self, workbookname, path):

        if not os.path.exists(path):
            os.makedirs(path)
        full_path = os.path.join(path, workbookname)
        wb = Workbook()
        # The default sheet "Sheet" is created. It will be removed later.
        wb.save(full_path)
        return wb, full_path
    
    def recalculate_excel_formulas(self, filepath):
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(filepath)
        wb.RefreshAll()
        wb.Save()
        wb.Close()
        excel.Quit()
    
    def close_reports(self, wb, full_path, workbookname):
        self.remove_default_sheet(wb)
        wb.save(full_path)
        self.recalculate_excel_formulas(full_path)
        print(f"Workbook {workbookname} saved at: {full_path}")
                
    def create_a_report(self, railroad_shortname):

        workbookname = f"{railroad_shortname}-{self.o_db.db_data.s_current_year}_report.xlsx"
        wb, full_path = self.create_workbook(workbookname,self.s_output_folder)
        print(f"Workbook {workbookname} created at: {full_path}")
    
        sheets = all_names()
        sheets = ["INDEX_worksheet", "A1P1_worksheet", "A1P2A_worksheet", "A1P2B_worksheet", "A1P2C_worksheet", "A1P3A_worksheet"]
        print("Available sheets to generate:", sheets)

        for name in sheets:
            print(f"Generating sheet: {name}")
            get(name)(self.ctx, wb)
            print(f"Sheet {name} generated.")
        
        
        self.close_reports(wb, full_path, workbookname)

    def create_reports(self):

        df = self.o_db.get_class1_rail_list()
        count = 0
        self.build_context()
        for row in df.itertuples(index=False):
            count += 1
            if (count>1):
                break
            self.ctx.current_df = row
            short_name = row.short_name
            self.update_variable_ctx(row.rr_id)
            self.create_a_report(short_name)

    # If you need backward-compatibility with old self.dt... attributes:
    def mirror_ctx_to_self(self):
        """(Optional) Temporarily mirror context fields to self for legacy code paths."""
        for name, value in self.ctx.__dict__.items():
            setattr(self, name, value)
            

if __name__ == "__main__":
    current_year = int(os.getenv("CURRENT_YEAR", "2023"))
    creator = CreateReports(current_year=current_year, fontname="Consolas", fontsize=10)
    creator.create_reports()
