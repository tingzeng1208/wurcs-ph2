# context.py
from dataclasses import dataclass, field
from typing import Optional
import pandas as pd
from currentDF import CurrentDF
from openpyxl.styles import Font, Alignment, Border, Side
import os, sys

current_dir = os.path.dirname(os.path.abspath(__file__))
utils_dir = os.path.join(current_dir, "..", "..", "utils")

if utils_dir not in sys.path:
    sys.path.append(utils_dir)
    
from utility import to_str  # type: ignore

@dataclass
class ReportContext:
    current_year: int
    sTitle_RR_YEAR: str
    s_output_folder: str
    fontname: str = "Consolas"
    fontsize: int = 10

    # dataframes (use default_factory to avoid mutable-default pitfalls)
    dtTitles: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtFootnotes: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtAValue: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtAValue0_RR: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtAValueRegion_RR: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtPriceIndexes: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtPriceIndexReferences: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtRegression: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtRegressionDependentVariable: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtRegressionDistribution: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtDefinitions: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtCarTypeStatistics: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtCarTypeStatisticsPart2: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtCarTypeStatisticsPart3: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtLineSourceText: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtDataDictionary: pd.DataFrame = field(default_factory=pd.DataFrame)
    dtECodes: pd.DataFrame = field(default_factory=pd.DataFrame)
    rr_icc: Optional[object] = None
    o_db: Optional[object] = None  # Store o_db reference if needed
    variable_ctx: Optional[CurrentDF] = None
    current_df = None  # alias railroad

    @classmethod
    def from_db(
        cls,
        o_db,
        *,
        current_year: int,
        sTitle_RR_YEAR: str,
        s_output_folder: str,
        fontname: str = "Consolas",
        fontsize: int = 10,
        verbose: bool = True,
    ) -> "ReportContext":
        """Build a fully-populated context by loading all required tables from the DB layer."""
        def _log(name, numOfRows):
            if verbose:
                print(f"Loaded {name}. number of rows: {numOfRows}")

        # fetch
        dtTitles = o_db.get_custom_data("REPORT_TITLES")
        _log("Loaded dtTitles. number of rows:", len(dtTitles))
        dtFootnotes = o_db.get_custom_data("REPORT_FOOTNOTES")
        _log("Loaded dtFootnotes. number of rows:", len(dtFootnotes))
        dtAValue0_RR = o_db.get_a_value0_rr(current_year)
        _log("Loaded dtAValue0_RR. number of rows:", len(dtAValue0_RR))
        dtPriceIndexReferences = o_db.get_custom_data("PRICE_INDEX_REFERENCE")
        _log("Loaded dtPriceIndexReferences. number of rows:", len(dtPriceIndexReferences))
        dtRegression = o_db.get_custom_data("REGR")
        _log("Loaded dtRegression. number of rows:", len(dtRegression))
        dtRegressionDependentVariable = o_db.get_custom_data("REGR_DEP_VAR")
        _log("Loaded dtRegressionDependentVariable. number of rows:", len(dtRegressionDependentVariable))
        dtRegressionDistribution = o_db.get_custom_data("REGR_DISTRIB")
        _log("Loaded dtRegressionDistribution. number of rows:", len(dtRegressionDistribution))
        dtDefinitions = o_db.get_custom_data("URCSDEF")
        _log("Loaded dtDefinitions. number of rows:", len(dtDefinitions))
        dtCarTypeStatisticsPart3 = o_db.get_car_type_statistics_part3()
        _log("Loaded dtCarTypeStatisticsPart3. number of rows:", len(dtCarTypeStatisticsPart3))
        dtLineSourceText = o_db.get_line_source_text()
        _log("Loaded dtLineSourceText. number of rows:", len(dtLineSourceText))
        dtDataDictionary = o_db.get_data_dictionary(current_year)
        _log("Loaded dtDataDictionary. number of rows:", len(dtDataDictionary))
        dtECodes = o_db.get_custom_data("ECODES")
        _log("Loaded dtECodes. number of rows:", len(dtECodes))
        o_db = o_db  # keep a reference if needed

        return cls(
            current_year=current_year,
            sTitle_RR_YEAR=sTitle_RR_YEAR,
            s_output_folder=s_output_folder,
            fontname=fontname,
            fontsize=fontsize,
            dtTitles=dtTitles,
            dtFootnotes=dtFootnotes,
            dtAValue0_RR=dtAValue0_RR,
            dtPriceIndexReferences=dtPriceIndexReferences,
            dtRegression=dtRegression,
            dtRegressionDependentVariable=dtRegressionDependentVariable,
            dtRegressionDistribution=dtRegressionDistribution,
            dtDefinitions=dtDefinitions,
            dtCarTypeStatisticsPart3=dtCarTypeStatisticsPart3,
            dtLineSourceText=dtLineSourceText,
            dtDataDictionary=dtDataDictionary,
            dtECodes=dtECodes,
            o_db=o_db,
            rr_icc=None,
        )
   
    def format_all_cells(self, ws, fontname=None, fontsize=None):
        """
        Format all cells in the worksheet with the given font name and size.
        If fontname or fontsize is None, use self.fontname and self.fontsize.
        """
        if fontname is None:
            fontname = getattr(self, "fontname", "Consolas")
        if fontsize is None:
            fontsize = getattr(self, "fontsize", 10)
        font = Font(name=fontname, size=fontsize)
        for row in ws.iter_rows():
            if row and row[0].row > 5:
                for cell in row:
                    cell.font = font 
        
    def scrub_year(self, s: str, current_year: int) -> str:
        s = s.replace("#Y#", str(current_year))
        s = s.replace("#Y-1#", str(current_year - 1))
        s = s.replace("#Y-2#", str(current_year - 2))
        s = s.replace("#Y-3#", str(current_year - 3))
        s = s.replace("#Y-4#", str(current_year - 4))
        s = s.replace("#Y-5#", str(current_year - 4))
        return s
    
    def add_formula_as_text(self, ws, row, col, source_text):
        """
        Writes a formula as a string (not executable) in the specified cell.
        This is useful for displaying the formula as text in Excel, not as a calculated value.
        """
        cell = ws.cell(row=row, column=col, value=source_text if not (source_text.startswith('=') or source_text.startswith('+')) else "'" + source_text)

        cell.alignment = Alignment(horizontal="left")

    def load_current(self, o_db, rr_no: str, *, verbose: bool = True) -> None:
        """Populate/replace the active RR-scoped bundle."""
        self.variable_ctx = CurrentDF.from_db(o_db, rr_no, self.current_year, verbose=verbose)
        
    def write_titles_and_column_headers(self, ws, dtTitles, sSelectWorktable, sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle):
        # Set font for all cells
        for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=iColumnCount):
            for cell in row:
                cell.font = Font(name="Consolas", size=10)

        # Title rows
        ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
        ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)
        ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
        ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

        # Get matching title rows
        part_str = sSheetTitle[sSheetTitle.find('P')+1:]
        filtered_titles = dtTitles[(dtTitles["worktable"] == sSheetTitle[:2]) & (dtTitles["part"] == part_str)]

        for _, drTitles in filtered_titles.iterrows():
            # Worktable titles
            ws.cell(row=3, column=2, value=to_str(drTitles["title1"]))
            ws.cell(row=3, column=2).font = Font(bold=True)
            ws.cell(row=4, column=2, value=to_str(drTitles["title2"]))
            ws.cell(row=4, column=2).font = Font(bold=True)
            ws.cell(row=5, column=2, value=to_str(drTitles["title3"]))
            ws.cell(row=5, column=2).font = Font(bold=True)

            sWorktableColumn = ""
            # Column headers
            for iCurrentCol in range(1, iColumnCount + 1):
                sColumnName = f"rpt_col{iCurrentCol}"
                col_val = str(drTitles.get(sColumnName, ""))
                if "Source" in col_val:
                    sWorktableColumn = col_val.split(" ")[0]
                else:
                    if sWorktableColumn:
                        ws.cell(row=6, column=iCurrentCol, value=sWorktableColumn)
                        ws.cell(row=6, column=iCurrentCol).alignment = Alignment(horizontal="center")
                ws.cell(row=7, column=iCurrentCol, value=col_val.replace("|", "\n"))

            # Column header formatting
            # Calculate Excel column range string for formatting
            def excel_col(n):
                result = ""
                while n > 0:
                    n, rem = divmod(n - 1, 26)
                    result = chr(65 + rem) + result
                return result

            start_col = "A"
            end_col = excel_col(iColumnCount)
            sColumnRange = f"{start_col}7:{end_col}7"
            for row in ws[sColumnRange]:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    cell.border = Border(bottom=Side(style='double'), left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
                    cell.font = Font(bold=True)
            # Set row height for header row
            ws.row_dimensions[7].height = 40
            # Set column width for header columns
            for iCurrentCol in range(1, iColumnCount + 1):
                col_letter = excel_col(iCurrentCol)
                ws.column_dimensions[col_letter].width = 25

            # Setup line, code, and identification columns
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 50
            for col in ['A', 'B', 'C']:
                for cell in ws[col]:
                    cell.alignment = Alignment(horizontal="left")
            # SET PAGE LAYOUT (openpyxl syntax)
            ws.page_setup.orientation = "landscape"
            ws.page_setup.zoom = 70
            ws.oddFooter.left.text = "Page &P of &N"
            ws.print_title_rows = '1:7'
            ws.print_title_cols = 'A:C'

            # Write the line source text
            iROW_COUNT = 8

            for _, drRow in self.dtLineSourceText[self.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
                # Write The Line Number
                ws.cell(row=iROW_COUNT, column=1, value=getattr(drRow, "line", getattr(drRow, "Line", None)))
                # Write Code
                ws.cell(row=iROW_COUNT, column=2, value="'" + str(getattr(drRow, "code", getattr(drRow, "Code", ""))))
                # Write Identification
                ws.cell(row=iROW_COUNT, column=3, value=str(getattr(drRow, "identification", getattr(drRow, "Identification", ""))))
                # increment iROW_COUNT
                iROW_COUNT += 1
            iROW_COUNT += 2
            # Write footnotes
            filtered_footnotes = self.dtFootnotes[(self.dtFootnotes["worktable"] == sSheetTitle[:2]) & (self.dtFootnotes["part"] == part_str)]
            for drRow in filtered_footnotes.itertuples(index=False):
                # Write The Footnote number
                ws.cell(row=iROW_COUNT, column=2, value=str(getattr(drRow, "no", getattr(drRow, "No", ""))))
                # Write Code
                ws.cell(row=iROW_COUNT, column=3, value="'" + str(getattr(drRow, "text", getattr(drRow, "Text", ""))))
                # increment iROW_COUNT
                iROW_COUNT += 1

    def WriteFirst3ColumnsAndPageLayout(self, ws, dtLineSourceText, dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable):
        # Set page layout
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.zoom = 70
        ws.oddFooter.left.text = "Page &P of &N"
        ws.print_title_rows = '$1:$7'
        ws.print_title_cols = 'A:C'

        # Write first 3 columns from dtLineSourceText
        iROW_COUNT = 8
        for _, drRow in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            ws.cell(row=iROW_COUNT, column=1, value=to_str(drRow["line"]))
            ws.cell(row=iROW_COUNT, column=2, value=to_str(drRow["code"]))
            ws.cell(row=iROW_COUNT, column=3, value=to_str(drRow["identification"]))
            iROW_COUNT += 1

        # Leave two rows, then write footnotes
        iROW_COUNT += 2
        for _, drRow in dtFootnotes[
            (dtFootnotes["worktable"] == sSheetTitle[:2]) &
            (dtFootnotes["part"] == sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3])
].iterrows():
            ws.cell(row=iROW_COUNT, column=2, value=to_str(drRow["no"]))
            ws.cell(row=iROW_COUNT, column=3, value=to_str(drRow["text"]))
            iROW_COUNT += 1