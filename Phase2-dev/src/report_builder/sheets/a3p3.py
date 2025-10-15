from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

def to_str(value):
    """Helper function to convert value to string safely"""
    return str(value) if value is not None else ""

def apostrophe(text):
    """Add apostrophe prefix if text starts with = or +"""
    return f"'{text}" if text.startswith(('=', '+')) else text

@register("A3P3_worksheet")
def A3P3_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A3 PART 3"
        iColumnCount = 27
        iWorkTableColumnCount = 12
        sSheetTitle = "A3P3"
        iLineNumberOffset = 293
        sNamedRangePrefix = "A3L"

        print(f"Processing {sSheetTitle}")
        ws = wb.create_sheet(title=sSheetTitle)
        ws.freeze_panes = ws['D8']

        part_str = sSheetTitle[sSheetTitle.find('P')+1:]
        part_only = part_str[:2] if len(part_str) >= 2 and part_str[:2].isdigit() else part_str[:1]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_only}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable,
                                            ctx.sTitle_RR_YEAR, sTitle_WORKTABLE,
                                            iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes,
                                            sSheetTitle, sSelectStringRows, sSelectWorktable)

        def set_cell(row, col, value, name=None, num_format=None, align_right=True):
            c = ws.cell(row=row, column=col, value=value)
            if name:
                wb.defined_names[name] = DefinedName(
                    name=name, attr_text=f"'{sSheetTitle}'!${c.column_letter}${c.row}"
                )
            if align_right:
                c.alignment = Alignment(horizontal="right")
            if num_format:
                c.number_format = num_format
            return c

        def pi_from_ref(worktable, line, column_text, year_col):
            ref = ctx.dtPriceIndexReferences.query("worktable == @worktable and line == @line and column == @column_text")
            if ref.empty:
                return ""
            idx = int(ref.iloc[0]["index"])
            row = ctx.dtPriceIndexes[ctx.dtPriceIndexes["index"] == idx]
            return "" if row.empty else row.iloc[0][year_col]

        iCurrentYear = int(ctx.current_year)

        # eligible lines are NOT in these sets/ranges:
        def is_excluded_for_values(aline:int) -> bool:
            return (305 <= aline <= 307) or (aline == 326) or (aline >= 342)

        # WRITE HARD VALUES
        for _, r in ctx.variable_ctx.dtAValue[ctx.variable_ctx.dtAValue["rpt_sheet"] == sSheetTitle].iterrows():
            aLine = int(r["aline"])
            if is_excluded_for_values(aLine):
                continue

            iProcessYear = int(r["year"])
            iROW_COUNT = aLine - iLineNumberOffset
            dd = ctx.dtDataDictionary[ctx.dtDataDictionary["line"] == f"{sNamedRangePrefix}{aLine}"]

            if iProcessYear == iCurrentYear:
                if not dd.empty:
                    set_cell(iROW_COUNT, 5, dd.iloc[0]["annperiod"], f"{sNamedRangePrefix}{aLine}C1")
                set_cell(iROW_COUNT, 7, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C2", "#,##0")
                set_cell(iROW_COUNT, 9,  pi_from_ref("A3P3", str(aLine), "3", "current_year"), f"{sNamedRangePrefix}{aLine}C3", "0.0000")

            elif iProcessYear == iCurrentYear - 1:
                set_cell(iROW_COUNT, 11, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C4", "#,##0")
                set_cell(iROW_COUNT, 13, pi_from_ref("A3P3", str(aLine), "5", "current_year_minus_1"), f"{sNamedRangePrefix}{aLine}C5", "0.0000")

            elif iProcessYear == iCurrentYear - 2:
                set_cell(iROW_COUNT, 15, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C6", "#,##0")
                set_cell(iROW_COUNT, 17, pi_from_ref("A3P3", str(aLine), "7", "current_year_minus_2"), f"{sNamedRangePrefix}{aLine}C7", "0.0000")

            elif iProcessYear == iCurrentYear - 3:
                set_cell(iROW_COUNT, 19, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C8", "#,##0")
                set_cell(iROW_COUNT, 21, pi_from_ref("A3P3", str(aLine), "9", "current_year_minus_3"), f"{sNamedRangePrefix}{aLine}C9", "0.0000")

            elif iProcessYear == iCurrentYear - 4:
                set_cell(iROW_COUNT, 23, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C10", "#,##0")
                set_cell(iROW_COUNT, 25, pi_from_ref("A3P3", str(aLine), "11", "current_year_minus_4"), f"{sNamedRangePrefix}{aLine}C11", "0.0000")

        # SOURCES
        for _, dr in ctx.dtLineSourceText[ctx.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            line = int(dr["line"])
            iLine = line - iLineNumberOffset

            # Sources C1..C12 -> cols 4..26
            for idx, col in enumerate(range(4, 27, 2), start=1):
                ws.cell(row=iLine, column=col,
                        value=apostrophe(ctx.scrub_year(to_str(dr.get(f"c{idx}", "")), iCurrentYear)))

            if is_excluded_for_values(line):
                # nulls in price-index columns with numeric fills in value columns
                def put(col, cnum, is_null):
                    if is_null:
                        set_cell(iLine, col, "=NULL_VALUE", f"{sNamedRangePrefix}{line}{cnum}")
                    else:
                        set_cell(iLine, col, ctx.scrub_year(to_str(dr[cnum.lower()]), iCurrentYear),
                                f"{sNamedRangePrefix}{line}{cnum}", "#,##0")

                put(5, "C1", True); put(7, "C2", False); put(9, "C3", True)
                put(11, "C4", False); put(13, "C5", True); put(15, "C6", False)
                put(17, "C7", True); put(19, "C8", False); put(21, "C9", True)
                put(23, "C10", False); put(25, "C11", True); put(27, "C12", False)
            else:
                # Create source for C12
                sSource = ctx.get_source_for_a3p3to_p8_summary_column(line, iLine)
                ws.cell(row=iLine, column=26, value=apostrophe(sSource) if len(sSource) > 0 else "")
                set_cell(iLine, 27, sSource, f"{sNamedRangePrefix}{line}C12", "#,##0")

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")
    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
