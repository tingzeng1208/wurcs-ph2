from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

def process_numeric_value(value):
    """Convert string/numeric value to appropriate type for Excel"""
    if value is None or value == "":
        return ""
    
    # If it's already a number, return as-is
    if isinstance(value, (int, float)):
        # Convert float to int if it's a whole number
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    
    # Try to convert string to number
    try:
        str_val = str(value).strip()
        if str_val == "":
            return ""
            
        # Try integer conversion first
        if '.' not in str_val and 'e' not in str_val.lower():
            return int(float(str_val))  # Use float() first to handle scientific notation
        else:
            # If it's a decimal, check if it's actually a whole number
            float_val = float(str_val)
            if float_val.is_integer():
                return int(float_val)
            return float_val
    except (ValueError, TypeError):
        return str(value)  # Return as string if conversion fails

@register("A1P8_worksheet")
def A1P8_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 8"
        iColumnCount = 5
        iWorkTableColumnCount = 1
        sSheetTitle = "A1P8"
        iLineNumberOffset = 572
        sNamedRangePrefix = "A1L"

        dtaValueRegion_RR = ctx.variable_ctx.dtAValueRegion_RR
        dtaValue0_RR = ctx.dtAValue0_RR
        dtaValue = ctx.variable_ctx.dtAValue
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # Write values from dtaValueRegion_RR
        filtered_region_values = dtaValueRegion_RR[
            (dtaValueRegion_RR["rpt_sheet"] == sSheetTitle) &
            (dtaValueRegion_RR["year"] == iCurrentYear) &
            (dtaValueRegion_RR["code"] == "C1")
        ]
        for _, draValueRegion in filtered_region_values.iterrows():
            iROW_COUNT = int(draValueRegion["aline"]) - iLineNumberOffset
            cell_value = process_numeric_value(draValueRegion["value"])
            cell = ws.cell(row=iROW_COUNT, column=5, value=cell_value)
            named_range = f"{sNamedRangePrefix}{draValueRegion['aline']}C1"
            wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write values from dtaValue0_RR
        lines_to_check = {584, 585, 586, 587, 595}
        filtered_zero_rr = dtaValue0_RR[
            (dtaValue0_RR["rpt_sheet"] == sSheetTitle) &
            (dtaValue0_RR["year"] == iCurrentYear) &
            (dtaValue0_RR["aline"].isin(lines_to_check))
        ]
        for _, draValue0 in filtered_zero_rr.iterrows():
            iROW_COUNT = int(draValue0["aline"]) - iLineNumberOffset
            cell_value = process_numeric_value(draValue0["value"])
            cell = ws.cell(row=iROW_COUNT, column=5, value=cell_value)
            named_range = f"{sNamedRangePrefix}{draValue0['aline']}C1"
            wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write values from dtaValue (Railroad specific)
        filtered_rr_values = dtaValue[
            (dtaValue["rpt_sheet"] == sSheetTitle) &
            (dtaValue["year"] == iCurrentYear) &
            (dtaValue["aline"] == 581)
        ]
        for _, draValue in filtered_rr_values.iterrows():
            iROW_COUNT = int(draValue["aline"]) - iLineNumberOffset
            cell_value = process_numeric_value(draValue["value"])
            cell = ws.cell(row=iROW_COUNT, column=5, value=cell_value)
            named_range = f"{sNamedRangePrefix}{draValue['aline']}C1"
            wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write sources
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset
            source_text = ctx.scrub_year(str(drSource.get("c1", "")), iCurrentYear)
            ws.cell(row=iLine, column=4, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
