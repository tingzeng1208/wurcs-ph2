from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext
from utils.utility import scrub_year


@register("A3P2_worksheet")
def A3P2_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        # BUILD THE WORKTABLE TITLE AND SET WORKTABLE VARS
        sTitle_WORKTABLE = "WORKTABLE A3 PART 2"
        iColumnCount = 89
        iWorkTableColumnCount = 43
        sSheetTitle = "A3P2"
        iLineNumberOffset = 193
        sNamedRangePrefix = "A3L"

        print(f"Processing {sSheetTitle}")

        # SHEET + FREEZE (equiv. to select/split/freeze panes)
        ws = wb.create_sheet(title=sSheetTitle)
        ws.freeze_panes = ws['D8']

        # SELECT STRINGS (faithful to VB Substring/IndexOf logic)
        try:
            part_only = sSheetTitle[sSheetTitle.index('P') + 1 : sSheetTitle.index('P') + 3]
        except Exception:
            part_only = sSheetTitle[sSheetTitle.index('P') + 1 : sSheetTitle.index('P') + 2]
        sSelectWorktable   = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_only}'"
        sSelectStringRows  = f"Rpt_sheet = '{sSheetTitle}'"

        # TITLES + FIRST 3 COLS
        ctx.write_titles_and_column_headers(
            ws, ctx.dtTitles, sSelectWorktable,
            ctx.sTitle_RR_YEAR, sTitle_WORKTABLE,
            iColumnCount, sSheetTitle
        )
        ctx.WriteFirst3ColumnsAndPageLayout(
            ws, ctx.dtLineSourceText, ctx.dtFootnotes,
            sSheetTitle, sSelectStringRows, sSelectWorktable
        )

        def can_convert_to_number(s):
            """Check if string can be converted to a number"""
            try:
                float(s)
                return True
            except (ValueError, TypeError):
                return False
            
        # Helpers
        def set_cell(row, col, value, name=None, num_format=None, align_right=True):
            c = ws.cell(row=row, column=col, value=value)
            if name:
                wb.defined_names[name] = DefinedName(
                    name=name, attr_text=f"'{sSheetTitle}'!${c.column_letter}${c.row}"
                )
            if align_right:
                c.alignment = Alignment(horizontal="right")
            if num_format:
                if can_convert_to_number(value):
                    c.number_format = num_format
                    c.value = float(value)
                
            return c

        def get_pi(idx, year_col):
            df = ctx.variable_ctx.dtPriceIndexes[ctx.variable_ctx.dtPriceIndexes['index'] == idx]
            return "" if df.empty else df.iloc[0][year_col]

        def to_str(v):
            return "" if v is None else str(v)

        iCurrentYear = int(ctx.current_year)

        dtAvalue = ctx.variable_ctx.dtAValue
        # WRITE HARD VALUES FROM aValues
        # Faithfully filter by Rpt_sheet = sSheetTitle
        for _, r in dtAvalue[dtAvalue["rpt_sheet"] == sSheetTitle].iterrows():
            aLine = int(r["aline"])
            if aLine != 219 and aLine != 224:
                iProcessYear = int(r["year"])
                # drAnnPeriod (data dictionary row) by "Line = 'A3L{aLine}'"
                dd = ctx.dtDataDictionary[ctx.dtDataDictionary["line"] == f"{sNamedRangePrefix}{aLine}"]
                iROW_COUNT = aLine - iLineNumberOffset

                if iProcessYear == iCurrentYear:
                    # C1 (column 5) = AnnPeriod
                    if not dd.empty:
                        set_cell(iROW_COUNT, 5, to_str(dd.iloc[0]["annperiod"]), f"{sNamedRangePrefix}{aLine}C1")
                    # C2 (column 7) = Price Index #9 current year
                    set_cell(iROW_COUNT, 7, get_pi(9, "current_year"), f"{sNamedRangePrefix}{aLine}C2", "0.0000")

                    # C3..C8 based on aColumn - each IF should be separate to match VB logic
                    if to_str(r["acolumn"]) == "3":
                        set_cell(iROW_COUNT, 9, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C3", "#,##0")
                    if to_str(r["acolumn"]) == "4":
                        val = (f"=A3L215C4+A3L216C4+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 11, val, f"{sNamedRangePrefix}{aLine}C4", "#,##0")
                    if to_str(r["acolumn"]) == "5":
                        val = (f"=A3L215C5+A3L216C5+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 13, val, f"{sNamedRangePrefix}{aLine}C5", "#,##0")
                    if to_str(r["acolumn"]) == "6":
                        set_cell(iROW_COUNT, 15, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C6", "#,##0")
                    if to_str(r["acolumn"]) == "7":
                        val = (f"=A3L215C7+A3L216C7+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 17, val, f"{sNamedRangePrefix}{aLine}C7", "#,##0")
                    if to_str(r["acolumn"]) == "8":
                        val = (f"=A3L215C8+A3L216C8+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 19, val, f"{sNamedRangePrefix}{aLine}C8", "#,##0")

                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 21, get_pi(9, "current_year_minus_1"), f"{sNamedRangePrefix}{aLine}C9", "0.0000")
                    if to_str(r["acolumn"]) == "3":
                        set_cell(iROW_COUNT, 23, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C10", "#,##0")
                    if to_str(r["acolumn"]) == "4":
                        val = (f"=A3L215C11+A3L216C11+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 25, val, f"{sNamedRangePrefix}{aLine}C11", "#,##0")
                    if to_str(r["acolumn"]) == "5":
                        val = (f"=A3L215C12+A3L216C12+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 27, val, f"{sNamedRangePrefix}{aLine}C12", "#,##0")
                    if to_str(r["acolumn"]) == "6":
                        set_cell(iROW_COUNT, 29, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C13", "#,##0")
                    if to_str(r["acolumn"]) == "7":
                        val = (f"=A3L215C14+A3L216C14+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 31, val, f"{sNamedRangePrefix}{aLine}C14", "#,##0")
                    if to_str(r["acolumn"]) == "8":
                        val = (f"=A3L215C15+A3L216C15+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 33, val, f"{sNamedRangePrefix}{aLine}C15", "#,##0")

                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 35, get_pi(9, "current_year_minus_2"), f"{sNamedRangePrefix}{aLine}C16", "0.0000")
                    if to_str(r["acolumn"]) == "3":
                        set_cell(iROW_COUNT, 37, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C17", "#,##0")
                    if to_str(r["acolumn"]) == "4":
                        val = (f"=A3L215C18+A3L216C18+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 39, val, f"{sNamedRangePrefix}{aLine}C18", "#,##0")
                    if to_str(r["acolumn"]) == "5":
                        val = (f"=A3L215C19+A3L216C19+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 41, val, f"{sNamedRangePrefix}{aLine}C19", "#,##0")
                    if to_str(r["acolumn"]) == "6":
                        set_cell(iROW_COUNT, 43, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C20", "#,##0")
                    if to_str(r["acolumn"]) == "7":
                        val = (f"=A3L215C21+A3L216C21+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 45, val, f"{sNamedRangePrefix}{aLine}C21", "#,##0")
                    if to_str(r["acolumn"]) == "8":
                        val = (f"=A3L215C22+A3L216C22+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 47, val, f"{sNamedRangePrefix}{aLine}C22", "#,##0")

                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 49, get_pi(9, "current_year_minus_3"), f"{sNamedRangePrefix}{aLine}C23", "0.0000")
                    if to_str(r["acolumn"]) == "3":
                        set_cell(iROW_COUNT, 51, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C24", "#,##0")
                    if to_str(r["acolumn"]) == "4":
                        val = (f"=A3L215C25+A3L216C25+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 53, val, f"{sNamedRangePrefix}{aLine}C25", "#,##0")
                    if to_str(r["acolumn"]) == "5":
                        val = (f"=A3L215C26+A3L216C26+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 55, val, f"{sNamedRangePrefix}{aLine}C26", "#,##0")
                    if to_str(r["acolumn"]) == "6":
                        set_cell(iROW_COUNT, 57, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C27", "#,##0")
                    if to_str(r["acolumn"]) == "7":
                        val = (f"=A3L215C28+A3L216C28+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 59, val, f"{sNamedRangePrefix}{aLine}C28", "#,##0")
                    if to_str(r["acolumn"]) == "8":
                        val = (f"=A3L215C29+A3L216C29+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 61, val, f"{sNamedRangePrefix}{aLine}C29", "#,##0")

                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 63, get_pi(9, "current_year_minus_4"), f"{sNamedRangePrefix}{aLine}C30", "0.0000")
                    if to_str(r["acolumn"]) == "3":
                        set_cell(iROW_COUNT, 65, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C31", "#,##0")
                    if to_str(r["acolumn"]) == "4":
                        val = (f"=A3L215C32+A3L216C32+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 67, val, f"{sNamedRangePrefix}{aLine}C32", "#,##0")
                    if to_str(r["acolumn"]) == "5":
                        val = (f"=A3L215C33+A3L216C33+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 69, val, f"{sNamedRangePrefix}{aLine}C33", "#,##0")
                    if to_str(r["acolumn"]) == "6":
                        set_cell(iROW_COUNT, 71, to_str(r["value"]), f"{sNamedRangePrefix}{aLine}C34", "#,##0")
                    if to_str(r["acolumn"]) == "7":
                        val = (f"=A3L215C35+A3L216C35+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 73, val, f"{sNamedRangePrefix}{aLine}C35", "#,##0")
                    if to_str(r["acolumn"]) == "8":
                        val = (f"=A3L215C36+A3L216C36+{to_str(r['value'])}" if iROW_COUNT == 25 else to_str(r["value"]))
                        set_cell(iROW_COUNT, 75, val, f"{sNamedRangePrefix}{aLine}C36", "#,##0")

        # WRITE OUT THE SOURCES AND ANY VALUES THAT EXECUTE THE SOURCE
        for _, dr in ctx.dtLineSourceText[ctx.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            line = int(dr["line"])
            iLine = line - iLineNumberOffset

            # Sources first: C1..C43 -> cols 4,6,8,...88 (even cols), as text with leading apostrophe
            source_columns = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78, 80, 82, 84, 86, 88]
            for idx, col in enumerate(source_columns, start=1):
                key = f"c{idx}" if f"c{idx}" in dr else f"C{idx}"
                src = scrub_year(to_str(dr.get(key, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=f"'{src}")

            if line == 219 or line == 224:
                # Special: write NULL_VALUE or numeric per VB
                set_cell(iLine, 5, "=NULL_VALUE", f"{sNamedRangePrefix}{line}C1").alignment = Alignment(horizontal="right")
                c = set_cell(iLine, 7, "=NULL_VALUE", f"{sNamedRangePrefix}{line}C2")
                c.number_format = "#,##0"; c.alignment = Alignment(horizontal="right")

                def put_num(col, cname):
                    c = set_cell(iLine, col, ctx.scrub_year(to_str(dr[cname.lower() if cname.lower() in dr else cname]), iCurrentYear),
                                f"{sNamedRangePrefix}{line}{cname}", "#,##0")
                    return c

                for col, cname in [(9,"C3"),(11,"C4"),(13,"C5"),(15,"C6"),(17,"C7"),(19,"C8")]:
                    put_num(col, cname)

                set_cell(iLine, 21, "=NULL_VALUE", f"{sNamedRangePrefix}{line}C9").alignment = Alignment(horizontal="right")

                for col, cname in [(23,"C10"),(25,"C11"),(27,"C12"),(29,"C13"),(31,"C14"),(33,"C15")]:
                    put_num(col, cname)

                set_cell(iLine, 35, "=NULL_VALUE", f"{sNamedRangePrefix}{line}C16").alignment = Alignment(horizontal="right")

                for col, cname in [(37,"C17"),(39,"C18"),(41,"C19"),(43,"C20"),(45,"C21"),(47,"C22")]:
                    put_num(col, cname)

                set_cell(iLine, 49, "=NULL_VALUE", f"{sNamedRangePrefix}{line}C23").alignment = Alignment(horizontal="right")

                for col, cname in [(51,"C24"),(53,"C25"),(55,"C26"),(57,"C27"),(59,"C28"),(61,"C29")]:
                    put_num(col, cname)

                set_cell(iLine, 63, "=NULL_VALUE", f"{sNamedRangePrefix}{line}C30").alignment = Alignment(horizontal="right")

                for col, cname in [(65,"C31"),(67,"C32"),(69,"C33"),(71,"C34"),(73,"C35"),(75,"C36"),
                                (77,"C37"),(79,"C38"),(81,"C39"),(83,"C40"),(85,"C41"),(87,"C42"),(89,"C43")]:
                    put_num(col, cname)

            else:
                # Create source for C37..C42 (notes in even cols 76,78,.. and values in odd cols 77,79,..)
                for k, (note_col, val_col, cname, arg_col) in enumerate(
                    [(76,77,"C37",3),(78,79,"C38",4),(80,81,"C39",5),
                    (82,83,"C40",6),(84,85,"C41",7),(86,87,"C42",8)], start=0
                ):
                    sSource = ctx.get_source_for_a3p2_summary_column(ws, line, iLine, arg_col)
                    ws.cell(row=iLine, column=note_col, value=(f"'{sSource}" if len(sSource) > 0 else ""))
                    set_cell(iLine, val_col, sSource, f"{sNamedRangePrefix}{line}{cname}", "#,##0")

                # C43 numeric (col 89) from source table (with scrub), as in VB
                c43 = ctx.scrub_year(to_str(dr["c43" if "c43" in dr else "C43"]), iCurrentYear)
                set_cell(iLine, 89, c43, f"{sNamedRangePrefix}{line}C43", "#,##0")

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback; print(traceback.format_exc())
        # faithful error handler call analog (if you want parity with VB's DB logging):
        # ctx.oDB.HandleError(str(ctx.current_year), "Error", str(ex), traceback.format_exc(), "A3P2")

def get_source_for_a3p2_summary_column(self, ws, line_number, cell_line_number, column_number):
        sReturn = ""
        
        sline_number = str(line_number)
        oCell = ws.cell(row=cell_line_number, column=5)
        try:
            iPeriod = int(oCell.value) if oCell.value is not None else 0
        except (ValueError, TypeError):
            iPeriod = 0

        if iPeriod > 0:
            sReturn = "=SUM(PRODUCT(A3L" + sline_number + "C2,A3L" + sline_number + "C" + str(column_number) + ")"

            if iPeriod > 1:
                sReturn = sReturn + ",PRODUCT(A3L" + sline_number + "C9,A3L" + sline_number + "C" + str(column_number + 7) + ")"
            if iPeriod > 2:
                sReturn = sReturn + ",PRODUCT(A3L" + sline_number + "C16,A3L" + sline_number + "C" + str(column_number + 14) + ")"
            if iPeriod > 3:
                sReturn = sReturn + ",PRODUCT(A3L" + sline_number + "C23,A3L" + sline_number + "C" + str(column_number + 21) + ")"
            if iPeriod > 4:
                sReturn = sReturn + ",PRODUCT(A3L" + sline_number + "C30,A3L" + sline_number + "C" + str(column_number + 28) + ")"

            sReturn = sReturn + ")/A3L" + sline_number + "C1"
        else:
            sReturn = ""

        return sReturn
