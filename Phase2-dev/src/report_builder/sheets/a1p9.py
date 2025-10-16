from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext
from utils.utility import excel_cell_to_col_row, to_str, apostrophe





different_cells = "E8, F8, G8, H8, I8, J8, K8, L8, M8, E9, F9, G9, H9, I9, J9, K9, L9, M9"

# Convert different_cells string to (col, row) pairs
cell_refs = [cell.strip() for cell in different_cells.split(',')]
identified_set_cells = {excel_cell_to_col_row(cell_ref) for cell_ref in cell_refs}

@register("A1P9_worksheet")
def A1P9_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 9"
        iColumnCount = 13
        iWorkTableColumnCount = 5
        sSheetTitle = "A1P9"
        iLineNumberOffset = 893
        sNamedRangePrefix = "A1L"

        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset

            for i in range(1, iWorkTableColumnCount + 1):
                c_name = f"c{i}"
                
                # Source column (4, 6, 8, ...)
                source_col = 2 * i + 2
                source_text = ctx.scrub_year(to_str(drSource.get(c_name, "")), iCurrentYear)
                source_text = f"'{source_text}" if source_text.startswith(('=', '+')) else source_text
                set_cell(ws, iLine, source_col, source_text)

                # Derived Value column (5, 7, 9, ...)
                value_col = 2 * i + 3
                value = drSource.get(c_name, "")
                cell = set_cell(ws, iLine, value_col, value)

                # Conditional formatting
                if 12 < iLine < 25:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0"

                # Create named range
                named_range_name = f"{sNamedRangePrefix}{drSource['line']}C{i}"
                wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write range names for empty cells
        for i in range(8, 26): # Rows 8 to 25
            for j in range(5, 14): # Columns 5 to 13
                caption_cell = ws.cell(row=6, column=j)
                cell = ws.cell(row=i, column=j)

                if caption_cell.value is not None and cell.value is None:
                    cell.value = "=NULL_VALUE"
                    caption_text = str(caption_cell.value).replace("(", "").replace(")", "")
                    named_range_name = f"{sNamedRangePrefix}{i + iLineNumberOffset}{caption_text}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#######0"

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
        
def set_cell(ws, row, col, value):

    # if value is None or (isinstance(value, str) and value.strip() == ""):
    #     value = ""
    # if (col, row) in identified_set_cells:
    #     print(f"Setting cell at Row: {row}, Column: {col} with value: {value}")
    cell = ws.cell(row=row, column=col, value=value)
    # cell.alignment = Alignment(horizontal="right")
    # cell.number_format = "#######0"
    
    # if (col, row) in identified_set_cells:
    #     print(f"Set cell at Row: {row}, Column: {col} with value: {cell.value}")
    return cell