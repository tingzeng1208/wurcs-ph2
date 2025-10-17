from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext
from utils.utility import to_str, apostrophe, scrub_year

@register("A3P1_worksheet")
def A3P1_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        # BUILD THE WORKTABLE TITLE AND SET WORKTABLE VARS
        sTitle_WORKTABLE = "WORKTABLE A3 PART 1"
        iColumnCount = 27
        iWorkTableColumnCount = 12
        sSheetTitle = "A3P1"
        iLineNumberOffset = 93
        sNamedRangePrefix = "A3L"

        print(f"Processing {sSheetTitle}")

        # SHEET + FREEZE
        ws = wb.create_sheet(title=sSheetTitle)
        ws.freeze_panes = ws['D8']

        # SELECT STRINGS
        part_str = sSheetTitle[sSheetTitle.find('P')+1:]
        part_only = part_str[:2] if len(part_str) >= 2 and part_str[:2].isdigit() else part_str[:1]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_only}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        # TITLES + FIRST 3 COLS
        ctx.write_titles_and_column_headers(
            ws, ctx.dtTitles, sSelectWorktable,
            ctx.sTitle_RR_YEAR, sTitle_WORKTABLE,
            iColumnCount, sSheetTitle
        )
        ctx.WriteFirst3ColumnsAndPageLayout(
            ws, ctx.dtLineSourceText, ctx.dtFootnotes,
            sSheetTitle, sSelectStringRows, sSelectWorktable
        )

        # Helpers
        def set_cell(row, col, value, name=None, num_format=None, align_right=True):
            c = ws.cell(row=row, column=col, value=value)
            if name:
                wb.defined_names[name] = DefinedName(
                    name=name, attr_text=f"'{sSheetTitle}'!${c.column_letter}${c.row}"
                )
            if align_right:
                c.alignment = Alignment(horizontal="right")
            if num_format:
                c.number_format = num_format
            return c

        def get_pi(index, year_col_name, value_str):
            if str(value_str) in ("0", "0.0") or value_str == 0:
                return "0"
            df = ctx.variable_ctx.dtPriceIndexes[ctx.variable_ctx.dtPriceIndexes['index'] == index]
            return "0" if df.empty else df.iloc[0][year_col_name]

        iCurrentYear = int(ctx.current_year)
        excluded_lines = {110, 120, 130, 140, 150, 160, 168, 178}

        # WRITE HARD VALUES
        for _, r in ctx.variable_ctx.dtAValue[ctx.variable_ctx.dtAValue["rpt_sheet"] == sSheetTitle].iterrows():
            aLine = int(r["aline"])
            if aLine in excluded_lines:
                continue

            iProcessYear = int(r["year"])
            iROW_COUNT = aLine - iLineNumberOffset
            dd = ctx.dtDataDictionary[ctx.dtDataDictionary["line"] == f"{sNamedRangePrefix}{aLine}"]

            # Current year
            if iProcessYear == iCurrentYear:
                if not dd.empty:
                    set_cell(iROW_COUNT, 5, dd.iloc[0]["annperiod"], f"{sNamedRangePrefix}{aLine}C1")
                set_cell(iROW_COUNT, 9,  get_pi(22, "current_year", r["value"]), f"{sNamedRangePrefix}{aLine}C3", "0.0000")
                set_cell(iROW_COUNT, 7,  to_str(r["value"]),                      f"{sNamedRangePrefix}{aLine}C2", "#,##0")

            # CY-1, CY-2, CY-3, CY-4 follow similar pattern
            elif iProcessYear == iCurrentYear - 1:
                set_cell(iROW_COUNT, 13, get_pi(22, "current_year_minus_1", r["value"]), f"{sNamedRangePrefix}{aLine}C5", "0.0000")
                set_cell(iROW_COUNT, 11, to_str(r["value"]),                             f"{sNamedRangePrefix}{aLine}C4", "#,##0")

            elif iProcessYear == iCurrentYear - 2:
                set_cell(iROW_COUNT, 17, get_pi(22, "current_year_minus_2", r["value"]), f"{sNamedRangePrefix}{aLine}C7", "0.0000")
                set_cell(iROW_COUNT, 15, to_str(r["value"]),                             f"{sNamedRangePrefix}{aLine}C6", "#,##0")

            elif iProcessYear == iCurrentYear - 3:
                set_cell(iROW_COUNT, 21, get_pi(22, "current_year_minus_3", r["value"]), f"{sNamedRangePrefix}{aLine}C9", "0.0000")
                set_cell(iROW_COUNT, 19, to_str(r["value"]),                             f"{sNamedRangePrefix}{aLine}C8", "#,##0")

            elif iProcessYear == iCurrentYear - 4:
                set_cell(iROW_COUNT, 25, get_pi(22, "current_year_minus_4", r["value"]), f"{sNamedRangePrefix}{aLine}C11", "0.0000")
                set_cell(iROW_COUNT, 23, to_str(r["value"]),                             f"{sNamedRangePrefix}{aLine}C10", "#,##0")

        # SOURCES + special excluded lines
        for _, dr in ctx.dtLineSourceText[ctx.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            line = int(dr["line"])
            iLine = line - iLineNumberOffset

            # Sources C1..C12 -> cols 4..26 (even columns)
            for idx, col in enumerate(range(4, 27, 2), start=1):
                key = f"c{idx}"
                ws.cell(row=iLine, column=col, value=apostrophe(ctx.scrub_year(to_str(dr.get(key, "")), iCurrentYear)))

            if line in excluded_lines:
                # Alternating NULL/value pattern
                pairs = [
                    (5, "C1",  True),  (7,  "C2", False),
                    (9, "C3",  True),  (11, "C4", False),
                    (13,"C5",  True),  (15, "C6", False),
                    (17,"C7",  True),  (19, "C8", False),
                    (21,"C9",  True),  (23, "C10", False),
                    (25,"C11", True),  (27, "C12", False),
                ]
                for col, cnum, is_null in pairs:
                    if is_null:
                        set_cell(iLine, col, "=NULL_VALUE", f"{sNamedRangePrefix}{line}{cnum}")
                    else:
                        # scrubbed numeric text, formatted as number
                        set_cell(iLine, col, ctx.scrub_year(to_str(dr[cnum.lower()]), iCurrentYear),
                                f"{sNamedRangePrefix}{line}{cnum}", "#,##0")
            else:
                # Create source for C12
                sSource = ctx.get_source_for_a3p1_summary_column(ws, line, iLine)
                ws.cell(row=iLine, column=26, value=apostrophe(sSource) if len(sSource) > 0 else "")
                set_cell(iLine, 27, sSource, f"{sNamedRangePrefix}{line}C12", "#,##0")

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")
    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
