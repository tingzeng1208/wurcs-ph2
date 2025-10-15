from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P1_worksheet")
def A1P1_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 1"
        iColumnCount = 21
        sSheetTitle = "A1P1"
        iLineNumberOffset = 93

        dtaValue = ctx.variable_ctx.dtAValue
        dtaValue0_RR = ctx.dtAValue0_RR
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        # Update status
        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        # Get worksheet
        ws = wb.create_sheet(title=sSheetTitle)

        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)  
        # Freeze panes
        ws.freeze_panes = ws['D8']

        # Write hard values from dtaValue for all lines but 158
        col_map = {
                    iCurrentYear: 5,
                    iCurrentYear - 1: 7,
                    iCurrentYear - 2: 9,
                    iCurrentYear - 3: 11,
                    iCurrentYear - 4: 13
                }
        for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
            if draValues["aline"] != 158:
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
                if iProcessYear in col_map:
                    col = col_map[iProcessYear]
                    # Calculate the 'C' number for the named range, similar to the original VB
                    c_num = iCurrentYear - iProcessYear + 1
                    ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                    cell = ws.cell(row=iROW_COUNT, column=col)
                    sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                    cell.alignment = Alignment(horizontal="right")
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    
                    cell.number_format = "#,##0"

        # Write hard values from dtaValue0_RR for line 158
        for _, draValues in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
            if draValues["aline"] == 158:
                iProcessYear = int(draValues["year"])
                iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset

                if iProcessYear in col_map:
                    c_num = iCurrentYear - iProcessYear + 1
                    col = col_map[iProcessYear]
                    value = draValues["value"] if iProcessYear == iCurrentYear else "0"
                    ws.cell(row=iROW_COUNT, column=col, value=value)
                    cell = ws.cell(row=iROW_COUNT, column=col)
                    sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                    cell.alignment = Alignment(horizontal="right")
                    wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    
                    cell.number_format = "#,##0"

        # Write sources and values
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset
            
            # Sources first
            source_cols = {
                4: "c1", 6: "c2", 8: "c3", 10: "c4", 12: "c5",
                14: "c6", 16: "c7", 18: "c8", 20: "c9"
            }
            for col, c_name in source_cols.items():
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), ctx.current_year)
                if source_text.startswith('=') or source_text.startswith('+'):
                    ctx.add_formula_as_text(ws, iLine, col, source_text)
                else:
                    ws.cell(row=iLine, column=col, value=source_text)

            # C6 - C9 Values
            value_cols_c6_c9 = {15: "c6", 17: "c7", 19: "c8",  21: "c9"}
            for col, c_name in value_cols_c6_c9.items():
                value = "0" if drSource["line"] == 158 else str(drSource[c_name])
                c_num = (col - 15) // 2 + 6
                ws.cell(row=iLine, column=col, value=value)
                cell = ws.cell(row=iLine, column=col)
                cell.number_format = "#,##0"
                named_range_name = f"A1L{drSource['line']}C{c_num}"
                wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                

            # Random values not from Data Dictionary
            random_lines = {111, 122, 131, 133, 137, 144, 145, 146, 151, 154, 155}
            if int(drSource["line"]) in random_lines:
                random_value_cols = {5: "c1", 7: "c2", 9: "c3", 11: "c4", 13: "c5"}
                for col, c_name in random_value_cols.items():
                    c_num = (col - 5) // 2 + 1
                    ws.cell(row=iLine, column=col, value=drSource.get(c_name, ""))
                    cell = ws.cell(row=iLine, column=col)
                    cell.number_format = "#,##0"
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                    

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in A1P1: {ex}")
        import traceback
        print(traceback.format_exc())
