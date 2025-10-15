from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
import openpyxl.styles.numbers
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P5A_worksheet")
def A1P5A_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 5A"
        iColumnCount = 25
        iWorkTableColumnCount = 11
        sSheetTitle = "A1P5A"
        iLineNumberOffset = 493
        sNamedRangePrefix = "A1L"

        dtCarTypeStatistics = ctx.variable_ctx.dtCarTypeStatistics
        dtaValueRegion_RR = ctx.variable_ctx.dtAValueRegion_RR
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # Write hard values from dtCarTypeStatistics
        for _, drCarStats in dtCarTypeStatistics.iterrows():
            iROW_COUNT = int(drCarStats["line"]) - iLineNumberOffset
            col_map = { 5: "c1", 7: "c2", 9: "c3", 13: "c5", 15: "c6", 17: "c7", 19: "c8", 21: "c9", 23: "c10", 25: "c11" }
            c_num_map = { 5: 1, 7: 2, 9: 3, 13: 5, 15: 6, 17: 7, 19: 8, 21: 9, 23: 10, 25: 11 }
            for col, c_name in col_map.items():
                cell = ws.cell(row=iROW_COUNT, column=col, value=drCarStats.get(c_name, ""))
                cell.number_format = "#,##0.#####"
                if c_name == "c6":
                    cell.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
                named_range = f"A1L{drCarStats['line']}C{c_num_map[col]}"
                wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write hard values from dtaValueRegion_RR
        filtered_region_values = dtaValueRegion_RR[
            (dtaValueRegion_RR["rpt_sheet"] == sSheetTitle) &
            (dtaValueRegion_RR["year"] == iCurrentYear) &
            (dtaValueRegion_RR["code"] == "C4")
        ]
        for _, draValueRegion in filtered_region_values.iterrows():
            iROW_COUNT = int(draValueRegion["aline"]) - iLineNumberOffset
            cell = ws.cell(row=iROW_COUNT, column=11, value=draValueRegion["value"])
            named_range = f"A1L{draValueRegion['aline']}C4"
            wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write sources
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset
            for n in range(1, iWorkTableColumnCount + 1):
                col = 2 * n + 2
                c_name = f"c{n}"
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)
        
        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
