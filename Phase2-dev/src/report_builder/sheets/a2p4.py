from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext
from utils.utility import to_str, apostrophe, scrub_year

@register("A2P4_worksheet")
def A2P4_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        # BUILD THE WORKTABLE TITLE AND SET WORKTABLE VARS
        sTitle_WORKTABLE = "WORKTABLE A2 PART 4"
        iColumnCount = 91
        iWorkTableColumnCount = 44
        sSheetTitle = "A2P4"
        iLineNumberOffset = 393
        sNamedRangePrefix = "A2L"

        print(f"Processing {sSheetTitle}")

        # SHEET + FREEZE
        ws = wb.create_sheet(title=sSheetTitle)
        ws.freeze_panes = ws['D8']

        # SELECT STRINGS
        part_str = sSheetTitle[sSheetTitle.find('P')+1:]
        if len(part_str) >= 2 and part_str[:2].isdigit():
            part_only = part_str[:2]
        else:
            part_only = part_str[:1]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_only}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        # TITLES + FIRST 3 COLS
        ctx.write_titles_and_column_headers(
            ws, ctx.dtTitles, sSelectWorktable,
            ctx.sTitle_RR_YEAR, sTitle_WORKTABLE,
            iColumnCount, sSheetTitle
        )
        ctx.WriteFirst3ColumnsAndPageLayout(
            ws, ctx.dtLineSourceText, ctx.dtFootnotes,
            sSheetTitle, sSelectStringRows, sSelectWorktable
        )

        # Helpers
        def set_cell(row, col, value, name, num_format=None, align_right=True):
            cell = ws.cell(row=row, column=col, value=value)
            wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
            if align_right:
                cell.alignment = Alignment(horizontal="right")
            if num_format:
                cell.number_format = num_format
            return cell

        def get_pi(index, year_col_name, value_str):
            if str(value_str) in ("0", "0.0") or value_str == 0:
                return "0"
            pi_row = ctx.variable_ctx.dtPriceIndexes[ctx.variable_ctx.dtPriceIndexes['index'] == index]
            return "0" if pi_row.empty else pi_row.iloc[0][year_col_name]

        def is_main_block(aline: int) -> bool:
            # VB condition: (aLine < 411) OR (aLine == 418)
            return (aline < 411) or (aline == 418)

        # --- WRITE HARD VALUES FROM AVALUES ---
        dtAvalue = ctx.variable_ctx.dtAValue  # for brevity
        iCurrentYear = int(ctx.current_year)
        for _, draValues in dtAvalue[dtAvalue["rpt_sheet"] == sSheetTitle].iterrows():
            iProcessYear = int(draValues["year"])
            aLine = int(draValues["aline"])
            acode_id = int(draValues["acode_id"])
            value_str = to_str(draValues["value"])

            drAnnPeriod = ctx.dtDataDictionary[ctx.dtDataDictionary["line"] == f"A2L{aLine}"]
            iROW_COUNT = aLine - iLineNumberOffset

            # iCodeOffset / iRowOffset mapping
            iCodeOffset = None
            iRowOffset = None
            if aLine < 411:
                iCodeOffset, iRowOffset = 1883, 8
            elif aLine == 418:
                iCodeOffset, iRowOffset = 1966, 25

            # AnnPeriod into C1 except for lines 419 and 420
            if aLine not in (419, 420):
                if iProcessYear == iCurrentYear and not drAnnPeriod.empty:
                    set_cell(iROW_COUNT, 5, drAnnPeriod.iloc[0]["annperiod"], f"A2L{aLine}C1")

            # Condition 1 (Index 1)
            if is_main_block(aLine) and iCodeOffset is not None and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset):
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 9,  get_pi(1, "current_year", value_str),           f"A2L{aLine}C3",  "0.0000")
                    set_cell(iROW_COUNT, 7,  value_str,                                       f"A2L{aLine}C2",  "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 29, get_pi(1, "current_year_minus_1", value_str),    f"A2L{aLine}C13", "0.0000")
                    set_cell(iROW_COUNT, 27, value_str,                                       f"A2L{aLine}C12", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 45, get_pi(1, "current_year_minus_2", value_str),    f"A2L{aLine}C21", "0.0000")
                    set_cell(iROW_COUNT, 43, value_str,                                       f"A2L{aLine}C20", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 61, get_pi(1, "current_year_minus_3", value_str),    f"A2L{aLine}C29", "0.0000")
                    set_cell(iROW_COUNT, 59, value_str,                                       f"A2L{aLine}C28", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 77, get_pi(1, "current_year_minus_4", value_str),    f"A2L{aLine}C37", "0.0000")
                    set_cell(iROW_COUNT, 75, value_str,                                       f"A2L{aLine}C36", "#,##0")

            # Condition 2 (Index 2)
            elif is_main_block(aLine) and iCodeOffset is not None and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 1:
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 13, get_pi(2, "current_year", value_str),            f"A2L{aLine}C5",  "0.0000")
                    set_cell(iROW_COUNT, 11, value_str,                                       f"A2L{aLine}C4",  "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 33, get_pi(2, "current_year_minus_1", value_str),    f"A2L{aLine}C15", "0.0000")
                    set_cell(iROW_COUNT, 31, value_str,                                       f"A2L{aLine}C14", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 49, get_pi(2, "current_year_minus_2", value_str),    f"A2L{aLine}C23", "0.0000")
                    set_cell(iROW_COUNT, 47, value_str,                                       f"A2L{aLine}C22", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 65, get_pi(2, "current_year_minus_3", value_str),    f"A2L{aLine}C31", "0.0000")
                    set_cell(iROW_COUNT, 63, value_str,                                       f"A2L{aLine}C30", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 81, get_pi(2, "current_year_minus_4", value_str),    f"A2L{aLine}C39", "0.0000")
                    set_cell(iROW_COUNT, 79, value_str,                                       f"A2L{aLine}C38", "#,##0")

            # Condition 3 (Index 3 OR explicit IDs)
            elif (
                (is_main_block(aLine) and iCodeOffset is not None and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 2)
                or acode_id in {1958, 1962}
            ):
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 17, get_pi(3, "current_year", value_str),            f"A2L{aLine}C7",  "0.0000")
                    set_cell(iROW_COUNT, 15, value_str,                                       f"A2L{aLine}C6",  "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 37, get_pi(3, "current_year_minus_1", value_str),    f"A2L{aLine}C17", "0.0000")
                    set_cell(iROW_COUNT, 35, value_str,                                       f"A2L{aLine}C16", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 53, get_pi(3, "current_year_minus_2", value_str),    f"A2L{aLine}C25", "0.0000")
                    set_cell(iROW_COUNT, 51, value_str,                                       f"A2L{aLine}C24", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 69, get_pi(3, "current_year_minus_3", value_str),    f"A2L{aLine}C33", "0.0000")
                    set_cell(iROW_COUNT, 67, value_str,                                       f"A2L{aLine}C32", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 85, get_pi(3, "current_year_minus_4", value_str),    f"A2L{aLine}C41", "0.0000")
                    set_cell(iROW_COUNT, 83, value_str,                                       f"A2L{aLine}C40", "#,##0")

            # Condition 4 (Index 4 OR explicit IDs; special sIndex when aLine > 420)
            elif (
                (is_main_block(aLine) and iCodeOffset is not None and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 3)
                or acode_id in {1943,1946,1949,1952,1955,1959,1963,1974,1977}
            ):
                sIndex = 23 if aLine > 420 else 4
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 21, get_pi(sIndex, "current_year", value_str),       f"A2L{aLine}C9",  "0.0000")
                    set_cell(iROW_COUNT, 19, value_str,                                       f"A2L{aLine}C8",  "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 41, get_pi(sIndex, "current_year_minus_1", value_str), f"A2L{aLine}C19", "0.0000")
                    set_cell(iROW_COUNT, 39, value_str,                                         f"A2L{aLine}C18", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 57, get_pi(sIndex, "current_year_minus_2", value_str), f"A2L{aLine}C27", "0.0000")
                    set_cell(iROW_COUNT, 55, value_str,                                         f"A2L{aLine}C26", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 73, get_pi(sIndex, "current_year_minus_3", value_str), f"A2L{aLine}C35", "0.0000")
                    set_cell(iROW_COUNT, 71, value_str,                                         f"A2L{aLine}C34", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 89, get_pi(sIndex, "current_year_minus_4", value_str), f"A2L{aLine}C43", "0.0000")
                    set_cell(iROW_COUNT, 87, value_str,                                         f"A2L{aLine}C42", "#,##0")

            # Condition 5 (C10) OR explicit IDs
            elif (
                (is_main_block(aLine) and iCodeOffset is not None and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 4)
                or acode_id in {1944,1947,1950,1953,1956,1960,1964,1975,1978}
            ):
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 23, value_str, f"A2L{aLine}C10", "#,##0")

            # Condition 6 (C11) OR explicit IDs
            elif (
                (is_main_block(aLine) and iCodeOffset is not None and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 5)
                or acode_id in {1945,1948,1951,1954,1957,1961,1965,1976,1979}
            ):
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 25, value_str, f"A2L{aLine}C11", "#,##0")

        # --- WRITE RANGE NAMES FOR EMPTY CELLS ---
        for i in range(8, 30):  # 8..29 inclusive
            for j in range(5, 92):  # 5..91 inclusive
                if ((i not in (26, 42)) or j == 5):
                    caption_cell = ws.cell(row=6, column=j)
                    cell = ws.cell(row=i, column=j)
                    if caption_cell.value is not None and (cell.value is None):
                        cell.value = "=NULL_VALUE"
                        name = f"A2L{i + iLineNumberOffset}{str(caption_cell.value).replace('(', '').replace(')', '')}"
                        wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = "#######0"

        # --- WRITE OUT THE SOURCES AND ANY VALUES THAT EXECUTE THE SOURCE ---
        for _, drSource in ctx.dtLineSourceText[ctx.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset

            # Sources first (C1..C43 -> cols 4..88), apostrophe + scrub
            for idx, col in enumerate(range(4, 89, 2), start=1):
                key = f"c{idx}"
                src_text = scrub_year(to_str(drSource.get(key, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=apostrophe(src_text))

            # Derived rows: 26, 27
            if iLine in (26, 27):
                def put_val(c, key, cnum):
                    set_cell(iLine, c, to_str(drSource.get(key, "")), f"A2L{int(drSource['line'])}C{cnum}", "#,##0")
                def put_null(c, cnum):
                    set_cell(iLine, c, "=NULL_VALUE", f"A2L{int(drSource['line'])}C{cnum}")

                put_val(7,  "c2",  2);  put_null(9,  3)
                put_val(11, "c4",  4);  put_null(13, 5)
                put_val(15, "c6",  6);  put_null(17, 7)
                put_val(19, "c8",  8);  put_null(21, 9)
                put_val(23, "c10", 10); put_val(25, "c11", 11)
                put_val(27, "c12", 12); put_null(29, 13)
                put_val(31, "c14", 14); put_null(33, 15)
                put_val(35, "c16", 16); put_null(37, 17)
                put_val(39, "c18", 18); put_null(41, 19)
                put_val(43, "c20", 20); put_null(45, 21)
                put_val(47, "c22", 22); put_null(49, 23)
                put_val(51, "c24", 24); put_null(53, 25)
                put_val(55, "c26", 26); put_null(57, 27)
                put_val(59, "c28", 28); put_null(61, 29)
                put_val(63, "c30", 30); put_null(65, 31)
                put_val(67, "c32", 32); put_null(69, 33)
                put_val(71, "c34", 34); put_null(73, 35)
                put_val(75, "c36", 36); put_null(77, 37)
                put_val(79, "c38", 38); put_null(81, 39)
                put_val(83, "c40", 40); put_null(85, 41)
                put_val(87, "c42", 42); put_null(89, 43)

                # Column 90 (raw source text with apostrophe), Column 91 numeric
                c44_src = scrub_year(to_str(drSource.get("c44", "")), iCurrentYear)
                ws.cell(row=iLine, column=90, value=apostrophe(c44_src))
                set_cell(iLine, 91, drSource.get("c44", ""), f"A2L{int(drSource['line'])}C44", "#,##0")
            else:
                # Create source for C44
                sSource = ctx.get_source_for_a2_summary_column(ws, drSource["line"], iLine)
                ws.cell(row=iLine, column=90, value=apostrophe(sSource) if len(sSource) > 0 else "")
                set_cell(iLine, 91, sSource, f"A2L{int(drSource['line'])}C44", "#,##0")

        # FINAL FORMAT
        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())

