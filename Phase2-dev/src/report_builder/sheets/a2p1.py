from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext
from utils.utility import to_str, apostrophe, scrub_year

def process_numeric_value(value):
    """Convert string/numeric value to appropriate type for Excel"""
    if value is None or value == "":
        return ""
    
    # If it's already a number, return as-is
    if isinstance(value, (int, float)):
        # Convert float to int if it's a whole number
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    
    # Try to convert string to number
    try:
        str_val = str(value).strip()
        if str_val == "":
            return ""
            
        # Try integer conversion first
        if '.' not in str_val and 'e' not in str_val.lower():
            return int(float(str_val))  # Use float() first to handle scientific notation
        else:
            # If it's a decimal, check if it's actually a whole number
            float_val = float(str_val)
            if float_val.is_integer():
                return int(float_val)
            return float_val
    except (ValueError, TypeError):
        return str(value)  # Return as string if conversion fails


@register("A2P1_worksheet")
def A2P1_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        # BUILD THE WORKTABLE TITLE AND SET WORKTABLE VARS
        sTitle_WORKTABLE = "WORKTABLE A2 PART 1"
        iColumnCount = 91
        iWorkTableColumnCount = 44
        sSheetTitle = "A2P1"
        iLineNumberOffset = 93
        sNamedRangePrefix = "A2L"

        print(f"Processing {sSheetTitle}")

        # GET OUR SHEET AND A RANGE
        ws = wb.create_sheet(title=sSheetTitle)

        # WRITE OUT TITLE AND COLUMN HEADERS
        part_str = sSheetTitle[sSheetTitle.find('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)

        # WRITE OUT THE FIRST 3 COLUMNS
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        # FREEZE THE ROWS AND COLUMNS
        ws.freeze_panes = ws['D8']

        # Helper to set cell value and named range
        def set_cell(row, col, value, name, num_format=None):
            cell = ws.cell(row=row, column=col, value=value)
            wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
            cell.alignment = Alignment(horizontal="right")
            if num_format:
                cell.number_format = num_format
            return cell

        # WRITE HARD VALUES FROM AVALUES
        iCurrentYear = int(ctx.current_year)
        dtAValue = ctx.variable_ctx.dtAValue
        
        for _, draValues in dtAValue[dtAValue["rpt_sheet"] == sSheetTitle].iterrows():
            iProcessYear = int(draValues["year"])
            aLine = int(draValues["aline"])
            acode_id = int(draValues["acode_id"])
            value_raw = draValues["value"]
            value_str = to_str(value_raw)  # Keep string version for comparisons
            value_numeric = process_numeric_value(value_raw)  # Use for cell values
            # print(f"Processing Line {aLine}, Code {acode_id}, Year {iProcessYear}, Value {value_str}")

            if value_str == 1.399175 or value_str == '1.399175':
                
                print(f"Debug: Line {aLine}, Code {acode_id}, Value {value_str}")

            drAnnPeriod = ctx.dtDataDictionary[ctx.dtDataDictionary["line"] == f"A2L{aLine}"]

            iROW_COUNT = aLine - iLineNumberOffset

            iCodeOffset = 898 if aLine < 142 else 1267
            iRowOffset = 8 if aLine < 142 else 82

            if iProcessYear == iCurrentYear and not drAnnPeriod.empty:
                # print(f"Setting annperiod for line {aLine} to {drAnnPeriod.iloc[0]['annperiod']}")
                set_cell(iROW_COUNT, 5, drAnnPeriod.iloc[0]["annperiod"], f"A2L{aLine}C1")

            is_main_block = (aLine < 142) or (174 < aLine < 181)
            
            # Price Index helper
            def get_pi(index, year_col_name):
                pi_row = ctx.variable_ctx.dtPriceIndexes[ctx.variable_ctx.dtPriceIndexes['index'] == index]
                # print(f"Value_str is {value_str} Fetching PI for index {index}, year column {year_col_name}: Found {len(pi_row)} rows")
                return "0" if value_str == 0 else pi_row.iloc[0][year_col_name] if not pi_row.empty else "0"

            # --- Start of faithful translation of the large If/ElseIf block ---

            # Condition 1
            if is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset):
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 9, get_pi(1, "current_year"), f"A2L{aLine}C3", "0.0000")
                    set_cell(iROW_COUNT, 7, value_numeric, f"A2L{aLine}C2", "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 29, get_pi(1, "current_year_minus_1"), f"A2L{aLine}C13", "0.0000")
                    set_cell(iROW_COUNT, 27, value_numeric, f"A2L{aLine}C12", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 45, get_pi(1, "current_year_minus_2"), f"A2L{aLine}C21", "0.0000")
                    set_cell(iROW_COUNT, 43, value_numeric, f"A2L{aLine}C20", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 61, get_pi(1, "current_year_minus_3"), f"A2L{aLine}C29", "0.0000")
                    set_cell(iROW_COUNT, 59, value_numeric, f"A2L{aLine}C28", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 77, get_pi(1, "current_year_minus_4"), f"A2L{aLine}C37", "0.0000")
                    set_cell(iROW_COUNT, 75, value_numeric, f"A2L{aLine}C36", "#,##0")

            # Condition 2
            elif is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 1:
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 13, get_pi(2, "current_year"), f"A2L{aLine}C5", "0.0000")
                    set_cell(iROW_COUNT, 11, value_numeric, f"A2L{aLine}C4", "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 33, get_pi(2, "current_year_minus_1"), f"A2L{aLine}C15", "0.0000")
                    set_cell(iROW_COUNT, 31, value_numeric, f"A2L{aLine}C14", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 49, get_pi(2, "current_year_minus_2"), f"A2L{aLine}C23", "0.0000")
                    set_cell(iROW_COUNT, 47, value_numeric, f"A2L{aLine}C22", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 65, get_pi(2, "current_year_minus_3"), f"A2L{aLine}C31", "0.0000")
                    set_cell(iROW_COUNT, 63, value_numeric, f"A2L{aLine}C30", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 81, get_pi(2, "current_year_minus_4"), f"A2L{aLine}C39", "0.0000")
                    set_cell(iROW_COUNT, 79, value_numeric, f"A2L{aLine}C38", "#,##0")

            # Condition 3
            elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 2) or \
                    acode_id in {1162, 1166, 1170, 1174, 1178, 1182, 1186, 1190, 1194, 1198, 1202, 1206, 1210, 1214, 1218, 1222, 1226, 1230, 1243, 1247, 1251, 1255, 1259, 1263}:
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 17, get_pi(3, "current_year"), f"A2L{aLine}C7", "0.0000")
                    set_cell(iROW_COUNT, 15, value_str, f"A2L{aLine}C6", "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 37, get_pi(3, "current_year_minus_1"), f"A2L{aLine}C17", "0.0000")
                    set_cell(iROW_COUNT, 35, value_str, f"A2L{aLine}C16", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 53, get_pi(3, "current_year_minus_2"), f"A2L{aLine}C25", "0.0000")
                    set_cell(iROW_COUNT, 51, value_str, f"A2L{aLine}C24", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 69, get_pi(3, "current_year_minus_3"), f"A2L{aLine}C33", "0.0000")
                    set_cell(iROW_COUNT, 67, value_str, f"A2L{aLine}C32", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 85, get_pi(3, "current_year_minus_4"), f"A2L{aLine}C41", "0.0000")
                    set_cell(iROW_COUNT, 83, value_str, f"A2L{aLine}C40", "#,##0")

            # Condition 4
            elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 3) or \
                    acode_id in {1144, 1147, 1150, 1153, 1156, 1159, 1234, 1237, 1240}:
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 21, get_pi(4, "current_year"), f"A2L{aLine}C9", "0.0000")
                    set_cell(iROW_COUNT, 19, value_str, f"A2L{aLine}C8", "#,##0")
                elif iProcessYear == iCurrentYear - 1:
                    set_cell(iROW_COUNT, 41, get_pi(4, "current_year_minus_1"), f"A2L{aLine}C19", "0.0000")
                    set_cell(iROW_COUNT, 39, value_str, f"A2L{aLine}C18", "#,##0")
                elif iProcessYear == iCurrentYear - 2:
                    set_cell(iROW_COUNT, 57, get_pi(4, "current_year_minus_2"), f"A2L{aLine}C27", "0.0000")
                    set_cell(iROW_COUNT, 55, value_str, f"A2L{aLine}C26", "#,##0")
                elif iProcessYear == iCurrentYear - 3:
                    set_cell(iROW_COUNT, 73, get_pi(4, "current_year_minus_3"), f"A2L{aLine}C35", "0.0000")
                    set_cell(iROW_COUNT, 71, value_str, f"A2L{aLine}C34", "#,##0")
                elif iProcessYear == iCurrentYear - 4:
                    set_cell(iROW_COUNT, 89, get_pi(4, "current_year_minus_4"), f"A2L{aLine}C43", "0.0000")
                    set_cell(iROW_COUNT, 87, value_str, f"A2L{aLine}C42", "#,##0")

            # Condition 5
            elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 4) or \
                    acode_id in {1145, 1148, 1151, 1154, 1157, 1160, 1164, 1168, 1172, 1176, 1180, 1184, 1188, 1192, 1196, 1200, 1204, 1208, 1212, 1216, 1220, 1224, 1228, 1232, 1235, 1238, 1241, 1245, 1249, 1253, 1257, 1261, 1265}:
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 23, value_str, f"A2L{aLine}C10", "#,##0")

            # Condition 6
            elif (is_main_block and acode_id == iCodeOffset + 6 * (iROW_COUNT - iRowOffset) + 5) or \
                    acode_id in {1146, 1149, 1152, 1155, 1158, 1161, 1165, 1169, 1173, 1177, 1181, 1185, 1189, 1193, 1197, 1201, 1205, 1209, 1213, 1217, 1221, 1225, 1229, 1233, 1236, 1239, 1242, 1246, 1250, 1254, 1258, 1262, 1266}:
                if iProcessYear == iCurrentYear:
                    set_cell(iROW_COUNT, 25, value_str, f"A2L{aLine}C11", "#,##0")

        # WRITE RANGE NAMES FOR EMPTY CELLS AND HANDLE SSAC SETTING
        SSAC_Cells = {"A2L166C8", "A2L166C10", "A2L166C11", "A2L166C18", "A2L166C26", "A2L166C34", "A2L166C42", "A2L167C8", "A2L167C10", "A2L167C11", "A2L167C18", "A2L167C26", "A2L167C34", "A2L167C42", "A2L168C8", "A2L168C10", "A2L168C11", "A2L168C18", "A2L168C26", "A2L168C34", "A2L168C42", "A2L175C2", "A2L175C4", "A2L175C6", "A2L175C8", "A2L175C10", "A2L175C11", "A2L175C12", "A2L175C14", "A2L175C16", "A2L175C18", "A2L175C20", "A2L175C22", "A2L175C24", "A2L175C26", "A2L175C28", "A2L175C30", "A2L175C32", "A2L175C34", "A2L175C36", "A2L175C38", "A2L175C40", "A2L175C42", "A2L176C2", "A2L176C4", "A2L176C6", "A2L176C8", "A2L176C10", "A2L176C11", "A2L176C12", "A2L176C14", "A2L176C16", "A2L176C18", "A2L176C20", "A2L176C22", "A2L176C24", "A2L176C26", "A2L176C28", "A2L176C30", "A2L176C32", "A2L176C34", "A2L176C36", "A2L176C38", "A2L176C40", "A2L176C42", "A2L177C2", "A2L177C4", "A2L177C6", "A2L177C8", "A2L177C10", "A2L177C11", "A2L177C12", "A2L177C14", "A2L177C16", "A2L177C18", "A2L177C20", "A2L177C22", "A2L177C24", "A2L177C26", "A2L177C28", "A2L177C30", "A2L177C32", "A2L177C34", "A2L177C36", "A2L177C38", "A2L177C40", "A2L177C42"}
        for i in range(8, 92): # Rows 8 to 91
            for j in range(5, 92): # Columns 5 to 91
                if i == 88 and j != 5:
                    continue
                
                caption_cell = ws.cell(row=6, column=j)
                cell = ws.cell(row=i, column=j)

                if caption_cell.value:
                    named_range_name = f"A2L{i + iLineNumberOffset}{str(caption_cell.value).replace('(', '').replace(')', '')}"
                    if cell.value is None:
                        set_cell(i, j, "=NULL_VALUE", named_range_name, "#######0")
                    else:
                        # Check if the cell's named range is in SSAC_Cells
                        # This requires finding the name for the cell's location
                        current_named_range = None
                        for name, dest in wb.defined_names.items():
                            if dest.attr_text == f"'{sSheetTitle}'!${cell.column_letter}${cell.row}":
                                current_named_range = name
                                break
                        if current_named_range in SSAC_Cells:
                            cell.value = f'=IF(SSAC="Y",0,{cell.value})'

        # WRITE OUT THE SOURCES AND ANY VALUES THAT EXECUTE THE SOURCE
        for _, drSource in ctx.dtLineSourceText[ctx.dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset
            # print(f"Processing line {iLine} , source line {drSource['line']} and iLineNumberOffset {iLineNumberOffset}")
            aline_str = str(drSource["line"])
            # print(f"Processing sources for line {aline_str}")

            # Sources first
            source_cols = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78, 80, 82, 84, 86, 88]
            for idx, col_num in enumerate(source_cols):
                c_name = f"c{idx + 1}"
                source_text = scrub_year(to_str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=col_num, value=apostrophe(source_text))

            if iLine == 88: # Line 181
                set_cell(iLine, 7, drSource["c2"], f"A2L{aline_str}C2", "#,##0")
                set_cell(iLine, 9, "=NULL_VALUE", f"A2L{aline_str}C3")
                set_cell(iLine, 11, drSource["c4"], f"A2L{aline_str}C4", "#,##0")
                set_cell(iLine, 13, "=NULL_VALUE", f"A2L{aline_str}C5")
                set_cell(iLine, 15, drSource["c6"], f"A2L{aline_str}C6", "#,##0")
                set_cell(iLine, 17, "=NULL_VALUE", f"A2L{aline_str}C7")
                set_cell(iLine, 19, drSource["c8"], f"A2L{aline_str}C8", "#,##0")
                set_cell(iLine, 21, "=NULL_VALUE", f"A2L{aline_str}C9")
                set_cell(iLine, 23, drSource["c10"], f"A2L{aline_str}C10", "#,##0")
                set_cell(iLine, 25, drSource["c11"], f"A2L{aline_str}C11", "#,##0")
                set_cell(iLine, 27, drSource["c12"], f"A2L{aline_str}C12", "#,##0")
                set_cell(iLine, 29, "=NULL_VALUE", f"A2L{aline_str}C13")
                set_cell(iLine, 31, drSource["c14"], f"A2L{aline_str}C14", "#,##0")
                set_cell(iLine, 33, "=NULL_VALUE", f"A2L{aline_str}C15")
                set_cell(iLine, 35, drSource["c16"], f"A2L{aline_str}C16", "#,##0")
                set_cell(iLine, 37, "=NULL_VALUE", f"A2L{aline_str}C17")
                set_cell(iLine, 39, drSource["c18"], f"A2L{aline_str}C18", "#,##0")
                set_cell(iLine, 41, "=NULL_VALUE", f"A2L{aline_str}C19")
                set_cell(iLine, 43, drSource["c20"], f"A2L{aline_str}C20", "#,##0")
                set_cell(iLine, 45, "=NULL_VALUE", f"A2L{aline_str}C21")
                set_cell(iLine, 47, drSource["c22"], f"A2L{aline_str}C22", "#,##0")
                set_cell(iLine, 49, "=NULL_VALUE", f"A2L{aline_str}C23")
                set_cell(iLine, 51, drSource["c24"], f"A2L{aline_str}C24", "#,##0")
                set_cell(iLine, 53, "=NULL_VALUE", f"A2L{aline_str}C25")
                set_cell(iLine, 55, drSource["c26"], f"A2L{aline_str}C26", "#,##0")
                set_cell(iLine, 57, "=NULL_VALUE", f"A2L{aline_str}C27")
                set_cell(iLine, 59, drSource["c28"], f"A2L{aline_str}C28", "#,##0")
                set_cell(iLine, 61, "=NULL_VALUE", f"A2L{aline_str}C29")
                set_cell(iLine, 63, drSource["c30"], f"A2L{aline_str}C30", "#,##0")
                set_cell(iLine, 65, "=NULL_VALUE", f"A2L{aline_str}C31")
                set_cell(iLine, 67, drSource["c32"], f"A2L{aline_str}C32", "#,##0")
                set_cell(iLine, 69, "=NULL_VALUE", f"A2L{aline_str}C33")
                set_cell(iLine, 71, drSource["c34"], f"A2L{aline_str}C34", "#,##0")
                set_cell(iLine, 73, "=NULL_VALUE", f"A2L{aline_str}C35")
                set_cell(iLine, 75, drSource["c36"], f"A2L{aline_str}C36", "#,##0")
                set_cell(iLine, 77, "=NULL_VALUE", f"A2L{aline_str}C37")
                set_cell(iLine, 79, drSource["c38"], f"A2L{aline_str}C38", "#,##0")
                set_cell(iLine, 81, "=NULL_VALUE", f"A2L{aline_str}C39")
                set_cell(iLine, 83, drSource["c40"], f"A2L{aline_str}C40", "#,##0")
                set_cell(iLine, 85, "=NULL_VALUE", f"A2L{aline_str}C41")
                set_cell(iLine, 87, drSource["c42"], f"A2L{aline_str}C42", "#,##0")
                set_cell(iLine, 89, "=NULL_VALUE", f"A2L{aline_str}C43")
                ws.cell(row=iLine, column=90, value=apostrophe(ctx.scrub_year(to_str(drSource.get('c44', '')), iCurrentYear)))
                # print(f"C44 source is at row: {iLine}, column: 90 is {drSource.get('c44', '')}")
                set_cell(iLine, 91, drSource["c44"], f"A2L{aline_str}C44", "#,##0")
            else:
                sSource = ctx.get_source_for_a2_summary_column(ws, drSource["line"], iLine)
                ws.cell(row=iLine, column=90, value=apostrophe(sSource if len(sSource) > 0 else ''))
                # print(f"row {iLine} column 90 source is {sSource}")
                set_cell(iLine, 91, sSource, f"A2L{aline_str}C44", "#,##0")

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())

