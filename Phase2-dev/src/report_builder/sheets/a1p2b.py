from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P2B_worksheet")
def A1P2B_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 2B"
        iColumnCount = 57
        sSheetTitle = "A1P2B"
        iLineNumberOffset = 209

        dtaValue = ctx.variable_ctx.dtAValue
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # Write hard values from dtaValue
        for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
            iProcessYear = int(draValues["year"])
            iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
            aCode_id = int(draValues.get("acode_id") or draValues.get("aCode_id", 0))

            # Even/odd column mapping (same as A1P2A)
            if aCode_id % 2 == 0:
                col_map = {
                    iCurrentYear: 5, iCurrentYear - 1: 11, iCurrentYear - 2: 17,
                    iCurrentYear - 3: 23, iCurrentYear - 4: 29
                }
                c_map = {
                    iCurrentYear: 1, iCurrentYear - 1: 4, iCurrentYear - 2: 7,
                    iCurrentYear - 3: 10, iCurrentYear - 4: 13
                }
            else:
                col_map = {
                    iCurrentYear: 7, iCurrentYear - 1: 13, iCurrentYear - 2: 19,
                    iCurrentYear - 3: 25, iCurrentYear - 4: 31
                }
                c_map = {
                    iCurrentYear: 2, iCurrentYear - 1: 5, iCurrentYear - 2: 8,
                    iCurrentYear - 3: 11, iCurrentYear - 4: 14
                }

            if iProcessYear in col_map:
                col = col_map[iProcessYear]
                c_num = c_map[iProcessYear]
                cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
                sNamedRange = f"A1L{draValues['aline']}C{c_num}"
                wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write sources and derived values
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset

            # Write all 27 source columns (C1 to C27) to columns 4,6,8,...,56
            for n in range(1, 28):
                col = 2 * n + 2
                c_name = f"c{n}"
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

            # Write derived values (same as A1P2A)
            derived_cols = [
                (9, 3), (15, 6), (21, 9), (27, 12), (33, 15), (35, 16), (37, 17), (39, 18),
                (41, 19), (43, 20), (45, 21), (47, 22), (49, 23), (51, 24), (53, 25), (55, 26), (57, 27)
            ]
            for col, c_num in derived_cols:
                c_name = f"c{c_num}"
                value = drSource.get(c_name, "")
                cell = ws.cell(row=iLine, column=col, value=value)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
                named_range_name = f"A1L{drSource['line']}C{c_num}"
                wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Random Values for line 234
            if int(drSource["line"]) == 234:
                random_value_cols = {
                    5: "c1", 7: "c2", 11: "c4", 13: "c5", 17: "c7", 19: "c8",
                    23: "c10", 25: "c11", 29: "c13", 31: "c14"
                }
                c_num_map = {
                    5: 1, 7: 2, 11: 4, 13: 5, 17: 7, 19: 8,
                    23: 10, 25: 11, 29: 13, 31: 14
                }
                for col, c_name in random_value_cols.items():
                    c_num = c_num_map[col]
                    value = drSource.get(c_name, "")
                    cell = ws.cell(row=iLine, column=col, value=value)
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    named_range_name = f"A1L{drSource['line']}C{c_num}"
                    wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
