from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P2C_worksheet")
def A1P2C_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 2C"
        iColumnCount = 57
        iWorkTableColumnCount = 27
        sSheetTitle = "A1P2C"
        iLineNumberOffset = 228
        sNamedRangePrefix = "A1L"

        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # This sheet has no values from the Data Dictionary (dtAValue)
        # It writes sources and values from dtLineSourceText
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset
            
            # Loop through C1 to C27 columns from dtLineSourceText
            for i in range(1, iWorkTableColumnCount + 1):
                c_name = f"c{i}"
                
                # Source column (4, 6, 8, ...)
                source_col = 2 * i + 2
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=source_col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

                # Value column (5, 7, 9, ...)
                value_col = source_col + 1
                value = drSource.get(c_name, "")
                cell = ws.cell(row=iLine, column=value_col, value=value)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

                # Create named range
                named_range_name = f"{sNamedRangePrefix}{drSource['line']}{c_name.upper()}"
                wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")
        
        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
