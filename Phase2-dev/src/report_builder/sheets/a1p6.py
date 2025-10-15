from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P6_worksheet")
def A1P6_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 6"
        iColumnCount = 165
        iWorkTableColumnCount = 81
        sSheetTitle = "A1P6"
        iLineNumberOffset = 533
        sNamedRangePrefix = "A1L"

        dtaValue = ctx.variable_ctx.dtAValue
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)
        iCodeOffset = 738

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # Write hard values from dtaValue
        for _, draValues in dtaValue[dtaValue["rpt_sheet"] == sSheetTitle].iterrows():
            iProcessYear = int(draValues["year"])
            iROW_COUNT = int(draValues["aline"]) - iLineNumberOffset
            iLineAdjustment = int(draValues["aline"]) - iLineNumberOffset - 8
            acode_id = int(draValues.get("acode_id", 0))

            year_diff = iCurrentYear - iProcessYear
            if not (0 <= year_diff <= 4):
                continue

            col_offset = year_diff * 18
            c_offset = year_diff * 9

            acode_check = acode_id - (iCodeOffset + iLineAdjustment * 4)

            if acode_check == 0: # C1, C10, C19, C28, C37
                col, c_num = 5 + col_offset, 1 + c_offset
            elif acode_check == 1: # C2, C11, C20, C29, C38
                col, c_num = 7 + col_offset, 2 + c_offset
            elif acode_check == 2: # C3, C12, C21, C30, C39
                col, c_num = 9 + col_offset, 3 + c_offset
            elif acode_check == 3: # C4, C13, C22, C31, C40
                col, c_num = 11 + col_offset, 4 + c_offset
            else:
                continue

            cell = ws.cell(row=iROW_COUNT, column=col, value=draValues["value"])
            cell.alignment = Alignment(horizontal="right")
            sNamedRange = f"{sNamedRangePrefix}{draValues['aline']}C{c_num}"
            wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write sources and derived values
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset

            # Write source columns (C1 to C81)
            for n in range(1, iWorkTableColumnCount + 1):
                col = 2 * n + 2
                c_name = f"c{n}"
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

            # Write derived values
            derived_cols = [
                (13, 5), (15, 6), (17, 7), (19, 8), (21, 9), (31, 14), (33, 15), (35, 16), (37, 17), (39, 18),
                (49, 23), (51, 24), (53, 25), (55, 26), (57, 27), (67, 32), (69, 33), (71, 34), (73, 35), (75, 36),
                (85, 41), (87, 42), (89, 43), (91, 44), (93, 45), (95, 46), (97, 47), (99, 48), (101, 49), (103, 50),
                (105, 51), (107, 52), (109, 53), (111, 54), (113, 55), (115, 56), (117, 57), (119, 58), (121, 59),
                (123, 60), (125, 61), (127, 62), (129, 63), (131, 64), (133, 65), (135, 66), (137, 67), (139, 68),
                (141, 69), (143, 70), (145, 71), (147, 72), (149, 73), (151, 74), (153, 75), (155, 76), (157, 77),
                (159, 78), (161, 79), (163, 80), (165, 81)
            ]
            for col, c_num in derived_cols:
                c_name = f"c{c_num}"
                value = drSource.get(c_name, "")
                cell = ws.cell(row=iLine, column=col, value=value)
                if c_num in {8, 9, 17, 18, 26, 27, 35, 36} or c_num >= 44:
                    cell.number_format = "#,##0.#####"
                named_range_name = f"{sNamedRangePrefix}{drSource['line']}C{c_num}"
                wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
