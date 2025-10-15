from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P7_worksheet")
def A1P7_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 7"
        iColumnCount = 13
        iWorkTableColumnCount = 5
        sSheetTitle = "A1P7"
        iLineNumberOffset = 553
        sNamedRangePrefix = "A1L"

        dtCarTypeStatisticsPart3 = ctx.variable_ctx.dtCarTypeStatisticsPart3
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # Write hard values from R_OP_STATS_BY_CAR_TYPE_3
        for _, drCarStats in dtCarTypeStatisticsPart3.iterrows():
            iROW_COUNT = int(drCarStats["line"]) - iLineNumberOffset
            for i in range(1, iWorkTableColumnCount + 1):
                col = 2 * i + 3 # 5, 7, 9, ...
                c_name = f"c{i}"
                cell = ws.cell(row=iROW_COUNT, column=col, value=drCarStats.get(c_name, ""))
                named_range = f"{sNamedRangePrefix}{drCarStats['line']}C{i}"
                wb.defined_names[named_range] = DefinedName(name=named_range, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write sources
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset
            for n in range(1, iWorkTableColumnCount + 1):
                col = 2 * n + 2 # 4, 6, 8, ...
                c_name = f"c{n}"
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
