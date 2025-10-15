from datetime import datetime
from report_builder.core.registry import register
from report_builder.core.context import ReportContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName

@register("INDEX_worksheet")
def INDEX_worksheet(ctx: ReportContext, wb: Workbook):
        """
        Builds the 'INDEX' worksheet in the provided workbook, sets titles, and populates user input values with formatting.
        Args:
            workbook (openpyxl.Workbook): The workbook to add the worksheet to.
        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The created worksheet object.
        """
        try:
            sTitle_WORKTABLE = "User Inputs and Index"
            sSheetTitle = "INDEX"
            ws = wb.create_sheet(title=sSheetTitle)

            # Set font for all cells
            default_font = Font(name="Consolas", size=10)
            for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=2):
                for cell in row:
                    cell.font = default_font
            sTitle_RR_YEAR = f"{ctx.current_df.short_name} {ctx.current_year} Run: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"
            ctx.sTitle_RR_YEAR = sTitle_RR_YEAR
            ws.cell(row=1, column=1, value=sTitle_RR_YEAR)
            ws.cell(row=1, column=1).font = Font(name="Consolas", size=10, italic=True)

            # Worktable title
            ws.cell(row=2, column=1, value=sTitle_WORKTABLE)
            ws.cell(row=2, column=1).font = Font(name="Consolas", size=12, bold=True)

            # User Input Table header
            ws.cell(row=4, column=1, value="User Input values:")
            ws.cell(row=4, column=1).font = Font(name="Consolas", size=12, bold=True)
            ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
            ws.merge_cells("A4:B4")
            # Add border and double bottom border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
            for cell in ws["A4:B4"]:
                for c in cell:
                    c.border = border

            # User input values
            user_inputs = [
                ("RRICC:", str(ctx.current_df.rricc), "RRICC"),
                ("Railroad Short Name:", str(ctx.current_df.short_name), "RR_SHORT_NAME"),
                ("Railroad Name:", str(ctx.current_df.name), "RR_NAME"),
                ("Railroad ID:", str(ctx.current_df.rr_id), "RR_ID"),
                ("Current Year:", str(ctx.o_db.db_data.s_current_year), "CURRENT_YEAR"),
                ("Run Years:", str(ctx.o_db.db_data.RUN_YEARS), "RUN_YEARS"),
                ("Embedded Cost of Capital for ROI:", "Y" if ctx.o_db.db_data.CapitalCost else "N", "EMBEDDED_COC"),
                ("100% Variability Flow Thru:", "Y" if ctx.o_db.db_data.VarialabilityFlow else "N", "FLOW_THRU_100PCT"),
                ("Include Account 76:", "Y" if ctx.o_db.db_data.Account76 else "N", "INCL_ACCT_76"),
                ("Include Account 80:", "Y" if ctx.o_db.db_data.Account80 else "N", "INCL_ACCT_80"),
                ("Include Account 90:", "Y" if ctx.o_db.db_data.Account90 else "N", "INCL_ACCT_90"),
                ("SSAC:", "Y" if ctx.o_db.db_data.SSAC else "N", "SSAC"),
                ("Replace Null Values With:", str(ctx.o_db.db_data.NullStringReplace), "NULL_VALUE"),
            ]
            for idx, (label, value, name) in enumerate(user_inputs, start=5):
                ws.cell(row=idx, column=1, value=label)
                cell = ws.cell(row=idx, column=2, value=value)
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(horizontal="right")
                ws.column_dimensions[cell.column_letter].width = 20
                # Add named range for this cell
                wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Set column width for A2 (worktable title)
            ws.column_dimensions['A'].width = 33

            return ws

        except Exception as ex:
            print(f"Error in build_index_worksheet: {ex}")
            return None