from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from report_builder.core.registry import register
from report_builder.core.context import ReportContext

@register("A1P4_worksheet")
def A1P4_worksheet(ctx: ReportContext, wb: Workbook):
    try:
        sTitle_WORKTABLE = "WORKTABLE A1 PART 4"
        iColumnCount = 9
        iWorkTableColumnCount = 3
        sSheetTitle = "A1P4"
        iLineNumberOffset = 393
        sNamedRangePrefix = "A1L"

        dtaValue0_RR = ctx.variable_ctx.dtAValue0_RR
        dtLineSourceText = ctx.dtLineSourceText
        iCurrentYear = int(ctx.current_year)

        print(f"Processing {sSheetTitle}")

        # Select worktable and string rows
        try:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:sSheetTitle.index('P')+3]
        except ValueError:
            part_str = sSheetTitle[sSheetTitle.index('P')+1:]
        sSelectWorktable = f"Worktable = '{sSheetTitle[:2]}' And Part = '{part_str}'"
        sSelectStringRows = f"Rpt_sheet = '{sSheetTitle}'"

        ws = wb.create_sheet(title=sSheetTitle)

        # Write titles and column headers
        ctx.write_titles_and_column_headers(ws, ctx.dtTitles, sSelectWorktable, ctx.sTitle_RR_YEAR, sTitle_WORKTABLE, iColumnCount, sSheetTitle)
        ctx.WriteFirst3ColumnsAndPageLayout(ws, ctx.dtLineSourceText, ctx.dtFootnotes, sSheetTitle, sSelectStringRows, sSelectWorktable)

        ws.freeze_panes = ws['D8']

        # Write hard values from dtaValue0_RR
        for _, draValue0_RR in dtaValue0_RR[dtaValue0_RR["rpt_sheet"] == sSheetTitle].iterrows():
            iProcessYear = int(draValue0_RR["year"])
            if iProcessYear == iCurrentYear:
                iROW_COUNT = int(draValue0_RR["aline"]) - iLineNumberOffset
                aCode_id = int(draValue0_RR.get("acode_id", 0))

                if aCode_id % 2 == 0: # Even
                    col, c_num = 5, 1
                else: # Odd
                    col, c_num = 7, 2

                cell = ws.cell(row=iROW_COUNT, column=col, value=draValue0_RR["value"])
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
                sNamedRange = f"A1L{draValue0_RR['aline']}C{c_num}"
                wb.defined_names[sNamedRange] = DefinedName(name=sNamedRange, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

        # Write sources and derived values
        for _, drSource in dtLineSourceText[dtLineSourceText["rpt_sheet"] == sSheetTitle].iterrows():
            iLine = int(drSource["line"]) - iLineNumberOffset

            # Write source columns (C1 to C3)
            for n in range(1, iWorkTableColumnCount + 1):
                col = 2 * n + 2
                c_name = f"c{n}"
                source_text = ctx.scrub_year(str(drSource.get(c_name, "")), iCurrentYear)
                ws.cell(row=iLine, column=col, value=f"'{source_text}" if source_text.startswith(('=', '+')) else source_text)

            # Write derived value for C3
            cell = ws.cell(row=iLine, column=9, value=drSource.get("c3", ""))
            named_range_name = f"A1L{drSource['line']}C3"
            wb.defined_names[named_range_name] = DefinedName(name=named_range_name, attr_text=f"'{sSheetTitle}'!${cell.column_letter}${cell.row}")

            # Random Values for specific lines
            if int(drSource["line"]) in {406, 426, 431, 439, 445, 449, 453, 457, 460, 465, 470, 475, 480}:
                # C1 value
                cell1 = ws.cell(row=iLine, column=5, value=drSource.get("c1", ""))
                cell1.number_format = "#,##0"
                wb.defined_names[f"A1L{drSource['line']}C1"] = DefinedName(name=f"A1L{drSource['line']}C1", attr_text=f"'{sSheetTitle}'!${cell1.column_letter}${cell1.row}")
                # C2 value
                cell2 = ws.cell(row=iLine, column=7, value=drSource.get("c2", ""))
                cell2.number_format = "#,##0"
                wb.defined_names[f"A1L{drSource['line']}C2"] = DefinedName(name=f"A1L{drSource['line']}C2", attr_text=f"'{sSheetTitle}'!${cell2.column_letter}${cell2.row}")
        
        ctx.format_all_cells(ws)
        print(f"{sSheetTitle} completed")

    except Exception as ex:
        print(f"Error in {sSheetTitle}: {ex}")
        import traceback
        print(traceback.format_exc())
